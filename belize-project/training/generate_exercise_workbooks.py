#!/usr/bin/env python3
"""Generate exercise workbook and answer key for the condition account training."""

import math
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ---- Style constants ----
GREEN = "3B9C7B"
TEAL = "0A5455"
GRAY = "404040"
TEXT = "30302F"
WHITE = "FFFFFF"
YELLOW = "FFF2CC"
LIGHT_GRAY = "F2F2F2"

green_fill = PatternFill(start_color=GREEN, end_color=GREEN, fill_type="solid")
teal_fill = PatternFill(start_color=TEAL, end_color=TEAL, fill_type="solid")
yellow_fill = PatternFill(start_color=YELLOW, end_color=YELLOW, fill_type="solid")
gray_fill = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")
bdr = Border(
    left=Side("thin", GRAY), right=Side("thin", GRAY),
    top=Side("thin", GRAY), bottom=Side("thin", GRAY),
)
hdr_font = Font(name="Arial", size=10, bold=True, color=WHITE)
data_font = Font(name="Arial", size=10, color=TEXT)
label_font = Font(name="Arial", size=10, bold=True, color=TEAL)
title_font = Font(name="Arial", size=14, bold=True, color=TEAL)
input_font = Font(name="Arial", size=10, color=TEXT)
answer_font = Font(name="Arial", size=10, bold=True, color="006100")
answer_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")


def apply_header(cell):
    cell.font = hdr_font
    cell.fill = green_fill
    cell.border = bdr
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def apply_row_header(cell):
    cell.font = Font(name="Arial", size=10, bold=True, color=WHITE)
    cell.fill = teal_fill
    cell.border = bdr
    cell.alignment = Alignment(horizontal="left", vertical="center")


def apply_data(cell, align="left"):
    cell.font = data_font
    cell.border = bdr
    cell.alignment = Alignment(horizontal=align, vertical="center")


def apply_input_cell(cell):
    cell.fill = yellow_fill
    cell.font = input_font
    cell.border = bdr
    cell.alignment = Alignment(horizontal="right", vertical="center")


def apply_answer(cell):
    cell.font = answer_font
    cell.fill = answer_fill
    cell.border = bdr
    cell.alignment = Alignment(horizontal="right", vertical="center")


# ---- Data ----
indicators = [
    {"name": "Live Coral Cover", "unit": "%", "ref": 50, "dir": "better", "col_idx": 2},
    {"name": "Fleshy Macroalgae Cover", "unit": "%", "ref": 50, "dir": "worse", "col_idx": 3},
    {"name": "Coral Recruit Density", "unit": "per m2", "ref": 15, "dir": "better", "col_idx": 4},
    {"name": "Coral Diversity (Shannon H')", "unit": "index", "ref": 2.0, "dir": "better", "col_idx": 5},
    {"name": "Reef Relief", "unit": "cm", "ref": 150, "dir": "better", "col_idx": 6},
]

site_data = [
    # Year, Site, LCC, FMA, RD, CD, Relief
    [2023, "BZHCCD01", 8.55, 28.63, 0.32, None, 45.50],
    [2023, "BZHCCD02", 6.17, 32.50, 0.20, None, 38.00],
    [2023, "BZ1231",   11.33, 30.10, 0.28, None, 62.30],
    [2023, "BZ1077",   5.83, 38.45, 0.18, None, 35.80],
    [2025, "BZHCCD01", 2.10, 40.15, 0.28, 1.12, 52.30],
    [2025, "BZHCCD02", 1.50, 28.30, 0.15, 0.95, 40.10],
    [2025, "BZ1231",   5.80, 32.50, 0.30, 1.18, 58.50],
    [2025, "BZ1077",   3.25, 45.20, 0.20, 0.88, 42.80],
]

ref_data = [
    ["Live Coral Cover", 50, "%", "Higher is better"],
    ["Fleshy Macroalgae Cover", 50, "%", "Higher is worse"],
    ["Coral Recruit Density", 15, "per m2", "Higher is better"],
    ["Coral Diversity (Shannon H')", 2.0, "index", "Higher is better"],
    ["Reef Relief", 150, "cm", "Higher is better"],
]


def calc_stats(values):
    """Calculate mean, SD, SE from a list of numbers."""
    vals = [v for v in values if v is not None]
    n = len(vals)
    if n == 0:
        return None, None, None, 0
    mean = sum(vals) / n
    if n > 1:
        var = sum((v - mean) ** 2 for v in vals) / (n - 1)
        sd = math.sqrt(var)
        se = sd / math.sqrt(n)
    else:
        sd = None
        se = None
    return mean, sd, se, n


def calc_ci(value, ref, direction):
    if value is None:
        return None
    if direction == "better":
        return min(value / ref, 1.0)
    else:
        return max(1.0 - value / ref, 0.0)


def compute_answers():
    """Compute the full answer key."""
    answers = []
    for ind in indicators:
        idx = ind["col_idx"]
        ref = ind["ref"]
        d = ind["dir"]

        open_vals = [row[idx] for row in site_data if row[0] == 2023]
        close_vals = [row[idx] for row in site_data if row[0] == 2025]

        o_mean, o_sd, o_se, o_n = calc_stats(open_vals)
        c_mean, c_sd, c_se, c_n = calc_stats(close_vals)

        o_ci = calc_ci(o_mean, ref, d) if o_mean is not None else None
        c_ci = calc_ci(c_mean, ref, d) if c_mean is not None else None

        o_se_ci = (o_se / ref) if o_se is not None else None
        c_se_ci = (c_se / ref) if c_se is not None else None

        if o_ci is not None and c_ci is not None:
            ci_change = c_ci - o_ci
            pct_change = (ci_change / o_ci * 100) if o_ci != 0 else None

            ses = [s for s in [o_se_ci, c_se_ci] if s is not None]
            se_thresh = sum(ses) / len(ses) if ses else 0

            if se_thresh > 0 and abs(ci_change) <= se_thresh:
                interp = "Stable"
            elif ci_change > 0:
                interp = "Improving"
            elif ci_change < 0:
                interp = "Declining"
            else:
                interp = "Stable"
        elif o_n > 0 and c_n == 0:
            ci_change, pct_change, interp = None, None, "Opening only"
        elif o_n == 0 and c_n > 0:
            ci_change, pct_change, interp = None, None, "Closing only"
        else:
            ci_change, pct_change, interp = None, None, "Insufficient data"

        n_min = min(o_n, c_n)
        if o_n == 0 or c_n == 0:
            n_avail = max(o_n, c_n)
            dq = ("Good (one period)" if n_avail >= 5
                  else "Moderate (one period)" if n_avail >= 3
                  else "Limited (one period)")
        else:
            dq = ("Good" if n_min >= 5
                  else "Moderate" if n_min >= 3
                  else "Limited")

        answers.append({
            "indicator": ind["name"],
            "unit": ind["unit"],
            "ref": ref,
            "o_mean": o_mean, "o_se": o_se, "o_n": o_n, "o_ci": o_ci,
            "c_mean": c_mean, "c_se": c_se, "c_n": c_n, "c_ci": c_ci,
            "ci_change": ci_change, "pct_change": pct_change,
            "interp": interp, "n_sites": f"{o_n}/{c_n}", "dq": dq,
        })
    return answers


def create_workbook(include_answers=False):
    wb = openpyxl.Workbook()

    # ---- INSTRUCTIONS ----
    ws = wb.active
    ws.title = "Instructions"
    ws.sheet_properties.tabColor = TEAL
    ws.column_dimensions["A"].width = 80

    if include_answers:
        title = "Answer Key: Hol Chan Marine Reserve Condition Account"
        lines = [
            (title, title_font),
            ("", None),
            ("This workbook contains the completed calculations for the exercise.", data_font),
            ("Green cells show the correct answers.", data_font),
        ]
    else:
        title = "Exercise: Hol Chan Marine Reserve Condition Account"
        lines = [
            (title, title_font),
            ("", None),
            ("This workbook contains site-level indicator data for 4 survey sites", data_font),
            ("in Hol Chan Marine Reserve (2023 and 2025). Your task is to:", data_font),
            ("", None),
            ("  1. Open the Your_Calculations sheet.", data_font),
            ("  2. For each indicator, calculate the mean and SE for each year.", data_font),
            ("  3. Normalize to condition indices using the reference levels.", data_font),
            ("  4. Calculate change and classify as Stable, Improving, or Declining.", data_font),
            ("  5. Fill in the Condition_Account sheet.", data_font),
            ("", None),
            ("Yellow cells are where you enter your calculations.", data_font),
            ("The Site_Data and Reference_Levels sheets are read-only inputs.", data_font),
            ("", None),
            ("Tip: start with live coral cover (the first indicator). Once you", data_font),
            ("have done one indicator, the others follow the same pattern.", data_font),
        ]

    for i, (text, font) in enumerate(lines, 1):
        cell = ws.cell(row=i, column=1, value=text)
        if font:
            cell.font = font

    # ---- REFERENCE LEVELS ----
    ws_ref = wb.create_sheet("Reference_Levels")
    ws_ref.sheet_properties.tabColor = GREEN
    for j, h in enumerate(["Indicator", "Reference Level", "Unit", "Direction"], 1):
        apply_header(ws_ref.cell(row=1, column=j, value=h))
    for i, row in enumerate(ref_data, 2):
        for j, val in enumerate(row, 1):
            c = ws_ref.cell(row=i, column=j, value=val)
            apply_data(c, "right" if j == 2 else "left")
    ws_ref.column_dimensions["A"].width = 30
    ws_ref.column_dimensions["B"].width = 16
    ws_ref.column_dimensions["C"].width = 10
    ws_ref.column_dimensions["D"].width = 18

    # ---- SITE DATA ----
    ws_data = wb.create_sheet("Site_Data")
    ws_data.sheet_properties.tabColor = GREEN
    headers = ["Year", "Site", "Live_Coral_Cover", "Fleshy_Macroalgae_Cover",
               "Recruit_Density", "Coral_Diversity_H", "Reef_Relief"]
    for j, h in enumerate(headers, 1):
        apply_header(ws_data.cell(row=1, column=j, value=h))
    for i, row in enumerate(site_data, 2):
        for j, val in enumerate(row, 1):
            if val is not None:
                c = ws_data.cell(row=i, column=j, value=val)
                apply_data(c, "right" if j >= 3 else "left")
                if j >= 3:
                    c.number_format = "0.00"
    ws_data.column_dimensions["A"].width = 8
    ws_data.column_dimensions["B"].width = 14
    for col_l in ["C", "D", "E", "F", "G"]:
        ws_data.column_dimensions[col_l].width = 24
    ws_data.freeze_panes = "A2"

    # ---- YOUR CALCULATIONS / ANSWERS ----
    ws_calc = wb.create_sheet("Your_Calculations")
    ws_calc.sheet_properties.tabColor = YELLOW if not include_answers else GREEN

    answers = compute_answers()

    r = 1
    ws_calc.cell(row=r, column=1, value="Calculations Worksheet").font = title_font
    r += 1
    ws_calc.cell(row=r, column=1,
                 value="For each indicator below, fill in the yellow cells.").font = data_font
    r += 2

    for a_idx, ans in enumerate(answers):
        ind = indicators[a_idx]
        col_idx = ind["col_idx"]

        # Indicator header
        c = ws_calc.cell(row=r, column=1, value=ans["indicator"])
        c.font = Font(name="Arial", size=12, bold=True, color=TEAL)
        c = ws_calc.cell(row=r, column=3,
                         value=f"Reference level: {ans['ref']} {ans['unit']}")
        c.font = label_font
        c = ws_calc.cell(row=r, column=6,
                         value=f"Direction: {ref_data[a_idx][3]}")
        c.font = label_font
        r += 1

        # Sub-table headers
        sub_hdrs = ["Statistic", "Opening (2023)", "Closing (2025)"]
        for j, h in enumerate(sub_hdrs, 1):
            apply_header(ws_calc.cell(row=r, column=j, value=h))
        r += 1

        # Rows: N, Mean, SE, CI
        stat_rows = [
            ("N (sites with data)", "o_n", "c_n", "0"),
            ("Mean", "o_mean", "c_mean", "0.00"),
            ("Standard Error", "o_se", "c_se", "0.00"),
            ("Condition Index", "o_ci", "c_ci", "0.0000"),
        ]
        for label, o_key, c_key, fmt in stat_rows:
            c = ws_calc.cell(row=r, column=1, value=label)
            apply_data(c)
            c.fill = gray_fill

            for col, key in [(2, o_key), (3, c_key)]:
                c = ws_calc.cell(row=r, column=col)
                if include_answers and ans[key] is not None:
                    c.value = round(ans[key], 4) if "ci" in key else (
                        round(ans[key], 2) if isinstance(ans[key], float) else ans[key])
                    apply_answer(c)
                    c.number_format = fmt
                else:
                    apply_input_cell(c)
                    c.number_format = fmt
            r += 1

        # Change row
        change_rows = [
            ("CI Change", "ci_change", "0.0000"),
            ("Percent Change (%)", "pct_change", "0.00"),
            ("Interpretation", "interp", None),
            ("Data Quality", "dq", None),
        ]
        for label, key, fmt in change_rows:
            c = ws_calc.cell(row=r, column=1, value=label)
            apply_data(c)
            c.fill = gray_fill

            c = ws_calc.cell(row=r, column=2)
            if include_answers and ans[key] is not None:
                val = ans[key]
                if isinstance(val, float):
                    val = round(val, 4) if "ci" in key else round(val, 2)
                c.value = val
                apply_answer(c)
                if fmt:
                    c.number_format = fmt
            else:
                apply_input_cell(c)
                if fmt:
                    c.number_format = fmt
            r += 1

        r += 1  # blank row between indicators

    ws_calc.column_dimensions["A"].width = 24
    ws_calc.column_dimensions["B"].width = 18
    ws_calc.column_dimensions["C"].width = 18
    ws_calc.column_dimensions["D"].width = 4
    ws_calc.column_dimensions["E"].width = 4
    ws_calc.column_dimensions["F"].width = 30

    # ---- CONDITION ACCOUNT TABLE ----
    ws_acct = wb.create_sheet("Condition_Account")
    ws_acct.sheet_properties.tabColor = TEAL

    ws_acct.cell(row=1, column=1,
                 value="SEEA EA Condition Account: Hol Chan Marine Reserve").font = title_font
    ws_acct.cell(row=2, column=1,
                 value="Ecosystem type: Coral reef (M1.3)").font = label_font
    ws_acct.cell(row=3, column=1,
                 value="Accounting period: 2023 (Opening) to 2025 (Closing)").font = label_font

    acct_hdrs = [
        "Indicator", "Unit", "Reference\nLevel", "Opening\nValue",
        "Opening\nSE", "Opening\nN", "Opening\nCI", "Closing\nValue",
        "Closing\nSE", "Closing\nN", "Closing\nCI", "CI\nChange",
        "Pct Change\n(%)", "Interpretation", "N Sites", "Data\nQuality",
    ]
    r = 5
    for j, h in enumerate(acct_hdrs, 1):
        apply_header(ws_acct.cell(row=r, column=j, value=h))

    for a_idx, ans in enumerate(answers):
        r = 6 + a_idx
        apply_row_header(ws_acct.cell(row=r, column=1, value=ans["indicator"]))
        apply_data(ws_acct.cell(row=r, column=2, value=ans["unit"]))
        c = ws_acct.cell(row=r, column=3, value=ans["ref"])
        apply_data(c, "right")

        # Columns 4-16: answers or empty
        fields = [
            (4, "o_mean", "0.00"), (5, "o_se", "0.00"), (6, "o_n", "0"),
            (7, "o_ci", "0.0000"), (8, "c_mean", "0.00"), (9, "c_se", "0.00"),
            (10, "c_n", "0"), (11, "c_ci", "0.0000"), (12, "ci_change", "0.0000"),
            (13, "pct_change", "0.00"), (14, "interp", None),
            (15, "n_sites", None), (16, "dq", None),
        ]
        for col, key, fmt in fields:
            c = ws_acct.cell(row=r, column=col)
            if include_answers and ans[key] is not None:
                val = ans[key]
                if isinstance(val, float):
                    val = round(val, 4) if "ci" in key else round(val, 2)
                c.value = val
                apply_answer(c)
            else:
                apply_input_cell(c)
            if fmt:
                c.number_format = fmt

    ws_acct.column_dimensions["A"].width = 30
    ws_acct.column_dimensions["B"].width = 8
    for col_l in "CDEFGHIJKLM":
        ws_acct.column_dimensions[col_l].width = 12
    ws_acct.column_dimensions["N"].width = 16
    ws_acct.column_dimensions["O"].width = 10
    ws_acct.column_dimensions["P"].width = 18
    ws_acct.freeze_panes = "A6"

    wb.active = wb.sheetnames.index("Your_Calculations" if not include_answers
                                     else "Condition_Account")
    return wb


# ---- Generate both workbooks ----
wb_exercise = create_workbook(include_answers=False)
wb_exercise.save("exercise_workbook.xlsx")
print("Created: exercise_workbook.xlsx")

wb_answer = create_workbook(include_answers=True)
wb_answer.save("answer_key.xlsx")
print("Created: answer_key.xlsx")
