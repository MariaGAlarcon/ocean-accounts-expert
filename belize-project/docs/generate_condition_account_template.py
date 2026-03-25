#!/usr/bin/env python3
"""Generate the SEEA EA Condition Account Excel Template.

Creates a reusable Excel workbook where Belize statisticians can:
- Paste site-level indicator values
- Edit reference levels
- Change opening/closing years and geographic unit
- See condition indices, change metrics, and interpretations update automatically

Output: docs/condition_account_template.xlsx
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ---- Style constants (GOAP formatting) ----
GREEN = "3B9C7B"
TEAL = "0A5455"
GRAY = "404040"
TEXT = "30302F"
WHITE = "FFFFFF"
YELLOW = "FFF2CC"

green_fill = PatternFill(start_color=GREEN, end_color=GREEN, fill_type="solid")
teal_fill = PatternFill(start_color=TEAL, end_color=TEAL, fill_type="solid")
yellow_fill = PatternFill(start_color=YELLOW, end_color=YELLOW, fill_type="solid")
cell_border = Border(
    left=Side("thin", GRAY), right=Side("thin", GRAY),
    top=Side("thin", GRAY), bottom=Side("thin", GRAY),
)
hdr_font = Font(name="Arial", size=10, bold=True, color=WHITE)
data_font = Font(name="Arial", size=10, color=TEXT)
label_font = Font(name="Arial", size=10, bold=True, color=TEAL)
title_font = Font(name="Arial", size=14, bold=True, color=TEAL)
input_font = Font(name="Arial", size=11, bold=True, color=TEXT)

MAX_ROW = 500  # Formula range end row

# ---- Indicator definitions ----
indicators = [
    {"name": "Live Coral Cover", "unit": "%", "ref": 50, "dir": "Higher is better", "col": "D"},
    {"name": "Fleshy Macroalgae Cover", "unit": "%", "ref": 50, "dir": "Higher is worse", "col": "E"},
    {"name": "Coral Recruit Density", "unit": "per m2", "ref": 15, "dir": "Higher is better", "col": "F"},
    {"name": "Coral Diversity (Shannon H')", "unit": "index", "ref": 2.0, "dir": "Higher is better", "col": "G"},
    {"name": "Reef Relief", "unit": "cm", "ref": 150, "dir": "Higher is better", "col": "H"},
    {"name": "Bleaching Prevalence", "unit": "%", "ref": 100, "dir": "Higher is worse", "col": "I"},
]

# ---- Example data (Ambergris Caye, approximate site-level values) ----
example_data = [
    [2023, "BZHCCD01", "Ambergris Caye (Pilot)", 8.55, 28.63, 0.32, None, 45.50, None],
    [2023, "BZHCCD02", "Ambergris Caye (Pilot)", 6.17, 32.50, 0.20, None, 38.00, None],
    [2023, "BZ1231",   "Ambergris Caye (Pilot)", 11.33, 30.10, 0.28, None, 62.30, None],
    [2023, "BZ1077",   "Ambergris Caye (Pilot)", 5.83, 38.45, 0.18, None, 35.80, None],
    [2023, "BZ1079",   "Ambergris Caye (Pilot)", 3.50, 42.30, None, None, None, None],
    [2023, "MXRCK",    "Ambergris Caye (Pilot)", 14.50, 25.80, 0.35, None, 55.10, None],
    [2023, "BZ1081",   "Ambergris Caye (Pilot)", 6.28, 41.62, 0.17, None, 46.62, None],
    [2025, "BZHCCD01", "Ambergris Caye (Pilot)", 2.10, 40.15, 0.28, 1.12, 52.30, 12.5],
    [2025, "BZHCCD02", "Ambergris Caye (Pilot)", 1.50, 28.30, 0.15, 0.95, 40.10, 8.3],
    [2025, "BZ1231",   "Ambergris Caye (Pilot)", 5.80, 32.50, 0.30, 1.18, 58.50, 15.2],
    [2025, "BZ1077",   "Ambergris Caye (Pilot)", 3.25, 45.20, 0.20, 0.88, 42.80, 22.1],
    [2025, "BZ1079",   "Ambergris Caye (Pilot)", 4.50, 36.00, 0.22, 1.17, 54.40, 10.0],
]


# ---- Helper functions ----
def apply_header(cell):
    cell.font = hdr_font
    cell.fill = green_fill
    cell.border = cell_border
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def apply_row_header(cell):
    cell.font = Font(name="Arial", size=10, bold=True, color=WHITE)
    cell.fill = teal_fill
    cell.border = cell_border
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)


def apply_data(cell, align="left"):
    cell.font = data_font
    cell.border = cell_border
    cell.alignment = Alignment(horizontal=align, vertical="center")


def apply_input(cell):
    cell.font = input_font
    cell.fill = yellow_fill
    cell.border = cell_border
    cell.alignment = Alignment(horizontal="left", vertical="center")


# ---- Create workbook ----
wb = openpyxl.Workbook()

# =========================================================================
# INSTRUCTIONS SHEET
# =========================================================================
ws = wb.active
ws.title = "Instructions"
ws.sheet_properties.tabColor = TEAL

lines = [
    ("Condition Account Template", title_font),
    ("", None),
    ("This workbook calculates SEEA EA coral reef condition accounts from", data_font),
    ("site-level indicator values. It has four working sheets:", data_font),
    ("", None),
    ("  1. Reference_Levels -- edit reference level values for each indicator", data_font),
    ("  2. Site_Data -- paste your site-level indicator values here", data_font),
    ("  3. Calculations -- intermediate statistics (do not edit)", data_font),
    ("  4. Condition_Account -- the final account table with formulas", data_font),
    ("", None),
    ("HOW TO USE", label_font),
    ("", None),
    ("Step 1. Open the Site_Data sheet. Delete the example data rows (keep the", data_font),
    ("header row). Paste your site-level indicator values. The columns must be:", data_font),
    ("Year, Site, Geographic_Unit, Live_Coral_Cover, Fleshy_Macroalgae_Cover,", data_font),
    ("Recruit_Density, Coral_Diversity_H, Reef_Relief, Bleaching_Prevalence.", data_font),
    ("Leave cells blank where data is missing. Do not enter zero for missing data.", data_font),
    ("The Year column must contain numbers (2023, 2025), not text.", data_font),
    ("", None),
    ("Step 2. Open the Condition_Account sheet. In the yellow cells at the top,", data_font),
    ("enter the opening year, closing year, and geographic unit name. The name", data_font),
    ("must match exactly what you entered in the Geographic_Unit column of Site_Data.", data_font),
    ("", None),
    ("Step 3. The condition account table below updates automatically.", data_font),
    ("", None),
    ("Step 4. To change reference levels, edit them on the Reference_Levels sheet.", data_font),
    ("", None),
    ("Step 5. To produce accounts for a different area, change the geographic unit", data_font),
    ("name in cell B4 of the Condition_Account sheet.", data_font),
    ("", None),
    ("NOTES", label_font),
    ("", None),
    ("Yellow cells are input cells you can edit. White cells contain formulas.", data_font),
    ("The example data on Site_Data is from the Ambergris Caye pilot. Means match", data_font),
    ("the actual condition account output; individual site values are approximations.", data_font),
    ("Standard errors may differ slightly from the R script output for this reason.", data_font),
    ("Replace the example data with your own data before use.", data_font),
    ("", None),
    ("Requires Excel 2016 or later.", data_font),
]

ws.column_dimensions["A"].width = 80
for i, (text, font) in enumerate(lines, 1):
    cell = ws.cell(row=i, column=1, value=text)
    if font:
        cell.font = font

# =========================================================================
# REFERENCE LEVELS SHEET
# =========================================================================
ws_ref = wb.create_sheet("Reference_Levels")
ws_ref.sheet_properties.tabColor = GREEN

for j, h in enumerate(["Indicator", "Reference Level", "Unit", "Direction"], 1):
    apply_header(ws_ref.cell(row=1, column=j, value=h))

for i, ind in enumerate(indicators, 2):
    c = ws_ref.cell(row=i, column=1, value=ind["name"])
    apply_data(c)
    c = ws_ref.cell(row=i, column=2, value=ind["ref"])
    apply_input(c)
    c.alignment = Alignment(horizontal="right", vertical="center")
    c = ws_ref.cell(row=i, column=3, value=ind["unit"])
    apply_data(c)
    c = ws_ref.cell(row=i, column=4, value=ind["dir"])
    apply_data(c)

ws_ref.column_dimensions["A"].width = 30
ws_ref.column_dimensions["B"].width = 16
ws_ref.column_dimensions["C"].width = 10
ws_ref.column_dimensions["D"].width = 18

# =========================================================================
# SITE DATA SHEET
# =========================================================================
ws_data = wb.create_sheet("Site_Data")
ws_data.sheet_properties.tabColor = GREEN

headers_data = [
    "Year", "Site", "Geographic_Unit",
    "Live_Coral_Cover", "Fleshy_Macroalgae_Cover",
    "Recruit_Density", "Coral_Diversity_H", "Reef_Relief",
    "Bleaching_Prevalence",
]
for j, h in enumerate(headers_data, 1):
    apply_header(ws_data.cell(row=1, column=j, value=h))

for i, row_data in enumerate(example_data, 2):
    for j, val in enumerate(row_data, 1):
        if val is not None:
            c = ws_data.cell(row=i, column=j, value=val)
            apply_data(c, "right" if j >= 4 else "left")
            if j >= 4:
                c.number_format = "0.00"

ws_data.column_dimensions["A"].width = 8
ws_data.column_dimensions["B"].width = 14
ws_data.column_dimensions["C"].width = 26
for col_l in ["D", "E", "F", "G", "H", "I"]:
    ws_data.column_dimensions[col_l].width = 24
ws_data.freeze_panes = "A2"

# =========================================================================
# CALCULATIONS SHEET
# =========================================================================
ws_calc = wb.create_sheet("Calculations")
ws_calc.sheet_properties.tabColor = GRAY

calc_hdrs = ["Indicator", "Period", "Data Col", "N", "Mean",
             "Sum of Squares", "Variance", "SD", "SE"]
for j, h in enumerate(calc_hdrs, 1):
    apply_header(ws_calc.cell(row=1, column=j, value=h))

calc_row = 2
for ind in indicators:
    col = ind["col"]
    for period, yr_cell in [("Opening", 2), ("Closing", 3)]:
        r = calc_row
        ws_calc.cell(row=r, column=1, value=ind["name"]).font = data_font
        ws_calc.cell(row=r, column=1).border = cell_border
        ws_calc.cell(row=r, column=2, value=period).font = data_font
        ws_calc.cell(row=r, column=2).border = cell_border
        ws_calc.cell(row=r, column=3, value=col).font = data_font
        ws_calc.cell(row=r, column=3).border = cell_border

        yr = f"Condition_Account!$B${yr_cell}"
        geo = "Condition_Account!$B$4"
        yr_rng = f"Site_Data!$A$2:$A${MAX_ROW}"
        geo_rng = f"Site_Data!$C$2:$C${MAX_ROW}"
        val_rng = f"Site_Data!{col}$2:{col}${MAX_ROW}"

        # N
        c = ws_calc.cell(row=r, column=4)
        c.value = f'=COUNTIFS({yr_rng},{yr},{geo_rng},{geo},{val_rng},"<>")'
        apply_data(c, "right")

        # Mean
        c = ws_calc.cell(row=r, column=5)
        c.value = f'=IFERROR(AVERAGEIFS({val_rng},{yr_rng},{yr},{geo_rng},{geo}),"")'
        apply_data(c, "right")
        c.number_format = "0.00"

        # Sum of squares (conditional)
        c = ws_calc.cell(row=r, column=6)
        c.value = (
            f"=SUMPRODUCT(({yr_rng}={yr})*({geo_rng}={geo})"
            f"*(ISNUMBER({val_rng}))*({val_rng})^2)"
        )
        apply_data(c, "right")
        c.number_format = "0.00"

        # Variance = (SumSq - N * Mean^2) / (N - 1)
        c = ws_calc.cell(row=r, column=7)
        c.value = f'=IFERROR((F{r}-D{r}*E{r}^2)/(D{r}-1),"")'
        apply_data(c, "right")
        c.number_format = "0.0000"

        # SD
        c = ws_calc.cell(row=r, column=8)
        c.value = f'=IFERROR(SQRT(G{r}),"")'
        apply_data(c, "right")
        c.number_format = "0.00"

        # SE = SD / sqrt(N)
        c = ws_calc.cell(row=r, column=9)
        c.value = f'=IFERROR(H{r}/SQRT(D{r}),"")'
        apply_data(c, "right")
        c.number_format = "0.00"

        calc_row += 1

ws_calc.column_dimensions["A"].width = 28
ws_calc.column_dimensions["B"].width = 10
ws_calc.column_dimensions["C"].width = 10
for col_l in ["D", "E", "F", "G", "H", "I"]:
    ws_calc.column_dimensions[col_l].width = 14

# =========================================================================
# CONDITION ACCOUNT SHEET
# =========================================================================
ws_acct = wb.create_sheet("Condition_Account")
ws_acct.sheet_properties.tabColor = TEAL

# Settings area
ws_acct.cell(row=1, column=1, value="Condition Account Settings").font = title_font
for i, (lbl, val) in enumerate(
    [("Opening Year", 2023), ("Closing Year", 2025),
     ("Geographic Unit", "Ambergris Caye (Pilot)")], 2
):
    ws_acct.cell(row=i, column=1, value=lbl).font = label_font
    apply_input(ws_acct.cell(row=i, column=2, value=val))

# Table headers (row 6)
acct_hdrs = [
    "Indicator", "Unit", "Reference\nLevel", "Opening\nValue", "Opening\nSE",
    "Opening\nN", "Opening\nCI", "Closing\nValue", "Closing\nSE", "Closing\nN",
    "Closing\nCI", "CI\nChange", "Pct Change\n(%)", "Interpretation",
    "N Sites", "Data\nQuality",
]
for j, h in enumerate(acct_hdrs, 1):
    apply_header(ws_acct.cell(row=6, column=j, value=h))

# Data rows (rows 7-11)
for i, ind in enumerate(indicators):
    r = 7 + i
    co = 2 + i * 2       # Calculations opening row
    cc = 2 + i * 2 + 1   # Calculations closing row
    rr = 2 + i            # Reference_Levels row

    # A: Indicator
    apply_row_header(ws_acct.cell(row=r, column=1, value=ind["name"]))

    # B: Unit
    apply_data(ws_acct.cell(row=r, column=2, value=ind["unit"]))

    # C: Reference Level
    c = ws_acct.cell(row=r, column=3)
    c.value = f"=Reference_Levels!B{rr}"
    apply_data(c, "right")
    c.number_format = "0.00" if ind["ref"] < 100 else "0"

    # D: Opening Value
    c = ws_acct.cell(row=r, column=4)
    c.value = f'=IFERROR(IF(Calculations!D{co}=0,"",Calculations!E{co}),"")'
    apply_data(c, "right")
    c.number_format = "0.00"

    # E: Opening SE
    c = ws_acct.cell(row=r, column=5)
    c.value = f'=IFERROR(IF(Calculations!D{co}<=1,"",Calculations!I{co}),"")'
    apply_data(c, "right")
    c.number_format = "0.00"

    # F: Opening N
    c = ws_acct.cell(row=r, column=6)
    c.value = f"=Calculations!D{co}"
    apply_data(c, "right")

    # G: Opening CI
    if ind["dir"] == "Higher is better":
        ci_f = f'=IFERROR(IF(D{r}="","",MIN(D{r}/C{r},1)),"")'
    else:
        ci_f = f'=IFERROR(IF(D{r}="","",MAX(1-D{r}/C{r},0)),"")'
    c = ws_acct.cell(row=r, column=7)
    c.value = ci_f
    apply_data(c, "right")
    c.number_format = "0.0000"

    # H: Closing Value
    c = ws_acct.cell(row=r, column=8)
    c.value = f'=IFERROR(IF(Calculations!D{cc}=0,"",Calculations!E{cc}),"")'
    apply_data(c, "right")
    c.number_format = "0.00"

    # I: Closing SE
    c = ws_acct.cell(row=r, column=9)
    c.value = f'=IFERROR(IF(Calculations!D{cc}<=1,"",Calculations!I{cc}),"")'
    apply_data(c, "right")
    c.number_format = "0.00"

    # J: Closing N
    c = ws_acct.cell(row=r, column=10)
    c.value = f"=Calculations!D{cc}"
    apply_data(c, "right")

    # K: Closing CI
    if ind["dir"] == "Higher is better":
        ci_f = f'=IFERROR(IF(H{r}="","",MIN(H{r}/C{r},1)),"")'
    else:
        ci_f = f'=IFERROR(IF(H{r}="","",MAX(1-H{r}/C{r},0)),"")'
    c = ws_acct.cell(row=r, column=11)
    c.value = ci_f
    apply_data(c, "right")
    c.number_format = "0.0000"

    # L: CI Change
    c = ws_acct.cell(row=r, column=12)
    c.value = f'=IFERROR(IF(OR(G{r}="",K{r}=""),"",K{r}-G{r}),"")'
    apply_data(c, "right")
    c.number_format = "0.0000"

    # M: Pct Change
    c = ws_acct.cell(row=r, column=13)
    c.value = f'=IFERROR(IF(OR(G{r}="",K{r}="",G{r}=0),"",(K{r}-G{r})/G{r}*100),"")'
    apply_data(c, "right")
    c.number_format = "0.00"

    # N: Interpretation
    interp = (
        f'=IF(AND(ISNUMBER(G{r}),ISNUMBER(K{r})),'
        f'IF(AND(ISNUMBER(E{r}),ISNUMBER(I{r}),C{r}<>0),'
        f'IF(ABS(L{r})<=(E{r}/C{r}+I{r}/C{r})/2,"Stable",'
        f'IF(L{r}>0,"Improving","Declining")),'
        f'IF(L{r}>0,"Improving",IF(L{r}<0,"Declining","Stable"))),'
        f'IF(AND(ISNUMBER(G{r}),NOT(ISNUMBER(K{r}))),"Opening only",'
        f'IF(AND(NOT(ISNUMBER(G{r})),ISNUMBER(K{r})),"Closing only",'
        f'"Insufficient data")))'
    )
    c = ws_acct.cell(row=r, column=14)
    c.value = interp
    apply_data(c)

    # O: N Sites
    c = ws_acct.cell(row=r, column=15)
    c.value = f'=F{r}&"/"&J{r}'
    apply_data(c, "center")

    # P: Data Quality
    dq = (
        f'=IF(OR(F{r}=0,J{r}=0),'
        f'IF(MAX(F{r},J{r})>=5,"Good (one period)",'
        f'IF(MAX(F{r},J{r})>=3,"Moderate (one period)","Limited (one period)")),'
        f'IF(MIN(F{r},J{r})>=5,"Good",'
        f'IF(MIN(F{r},J{r})>=3,"Moderate","Limited")))'
    )
    c = ws_acct.cell(row=r, column=16)
    c.value = dq
    apply_data(c)

# Column widths
ws_acct.column_dimensions["A"].width = 30
ws_acct.column_dimensions["B"].width = 8
for col_l in "CDEFGHIJKLM":
    ws_acct.column_dimensions[col_l].width = 12
ws_acct.column_dimensions["N"].width = 16
ws_acct.column_dimensions["O"].width = 10
ws_acct.column_dimensions["P"].width = 20
ws_acct.freeze_panes = "A7"

# Set active sheet
wb.active = wb.sheetnames.index("Condition_Account")

# Save
output_path = "docs/condition_account_template.xlsx"
wb.save(output_path)
print(f"Created: {output_path}")
