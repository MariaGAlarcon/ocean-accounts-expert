#!/usr/bin/env python3
"""Regenerate services Excel files with Part A / Part B structure."""

from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

GREEN = "3B9C7B"
TEAL = "0A5455"
GRAY = "404040"
TEXT = "30302F"
WHITE = "FFFFFF"
YELLOW = "FFF2CC"
MINT = "D4EEE5"

green_fill = PatternFill(start_color=GREEN, end_color=GREEN, fill_type="solid")
teal_fill = PatternFill(start_color=TEAL, end_color=TEAL, fill_type="solid")
yellow_fill = PatternFill(start_color=YELLOW, end_color=YELLOW, fill_type="solid")
mint_fill = PatternFill(start_color=MINT, end_color=MINT, fill_type="solid")
bdr = Border(left=Side("thin", GRAY), right=Side("thin", GRAY),
             top=Side("thin", GRAY), bottom=Side("thin", GRAY))
hdr_font = Font(name="Arial", size=10, bold=True, color=WHITE)
data_font = Font(name="Arial", size=10, color=TEXT)
label_font = Font(name="Arial", size=10, bold=True, color=TEAL)
title_font = Font(name="Arial", size=14, bold=True, color=TEAL)
subtitle_font = Font(name="Arial", size=11, bold=False, color=GREEN)
input_font = Font(name="Arial", size=10, color=TEXT)
answer_font = Font(name="Arial", size=10, bold=True, color="006100")
answer_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")


def apply_header(cell):
    cell.font = hdr_font; cell.fill = green_fill; cell.border = bdr
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def apply_row_header(cell):
    cell.font = Font(name="Arial", size=10, bold=True, color=WHITE)
    cell.fill = teal_fill; cell.border = bdr
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

def apply_data(cell, align="left"):
    cell.font = data_font; cell.border = bdr
    cell.alignment = Alignment(horizontal=align, vertical="center")

def apply_input(cell):
    cell.fill = yellow_fill; cell.font = input_font; cell.border = bdr
    cell.alignment = Alignment(horizontal="right", vertical="center")

def apply_answer(cell):
    cell.font = answer_font; cell.fill = answer_fill; cell.border = bdr
    cell.alignment = Alignment(horizontal="right", vertical="center")

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def section_title(ws, row, col, text):
    ws.cell(row=row, column=col, value=text).font = Font(
        name="Arial", size=12, bold=True, color=TEAL)


def generate_services_template():
    wb = openpyxl.Workbook()

    # ---- Sheet 1: Part A Physical Supply ----
    ws = wb.active
    ws.title = "Part A - Physical Supply"
    ws.sheet_properties.tabColor = GREEN

    ws["A1"] = "SEEA EA Ecosystem Service Account -- Part A: Physical Supply"
    ws["A1"].font = title_font
    ws["A2"] = "Part A is a complete account on its own. Monetary valuation (Part B) is optional."
    ws["A2"].font = subtitle_font
    ws["A4"] = "Accounting area:"; ws["A4"].font = label_font
    ws["B4"].fill = yellow_fill; ws["B4"].border = bdr
    ws["A5"] = "Accounting year:"; ws["A5"].font = label_font
    ws["B5"].fill = yellow_fill; ws["B5"].border = bdr

    headers = ["Service", "Unit", "Coral reefs", "Seagrass", "Mangroves", "Other", "Total"]
    for c, h in enumerate(headers, 1):
        apply_header(ws.cell(row=7, column=c, value=h))

    services = [
        ("Fish provisioning", "kg/yr"),
        ("Carbon sequestration", "Mg CO2/yr"),
        ("Coastal protection", "m coastline"),
        ("Nursery habitat", "kg biomass/yr"),
        ("Recreation", "visitors/yr"),
        ("Gleaning", "hours/yr"),
        ("Gleaning harvest", "kg/yr"),
    ]
    for r, (svc, unit) in enumerate(services, 8):
        apply_row_header(ws.cell(row=r, column=1, value=svc))
        apply_data(ws.cell(row=r, column=2, value=unit))
        for c in range(3, 7):
            apply_input(ws.cell(row=r, column=c))
        ws.cell(row=r, column=7).value = f"=SUM(C{r}:F{r})"
        apply_data(ws.cell(row=r, column=7), "right")

    set_col_widths(ws, [24, 16, 16, 16, 16, 16, 16])

    # ---- Sheet 2: Part B Monetary Supply ----
    ws2 = wb.create_sheet("Part B - Monetary Supply")
    ws2.sheet_properties.tabColor = TEAL

    ws2["A1"] = "SEEA EA Ecosystem Service Account -- Part B: Monetary Supply (USD/yr)"
    ws2["A1"].font = title_font
    ws2["A2"] = "Part B builds on Part A. Only fill in services where you have economic data."
    ws2["A2"].font = subtitle_font

    m_headers = ["Service", "Valuation method", "Value type", "Coral reefs", "Seagrass", "Mangroves", "Other", "Total"]
    for c, h in enumerate(m_headers, 1):
        apply_header(ws2.cell(row=4, column=c, value=h))

    m_services = [
        ("Fish provisioning", "Resource rent", "Market"),
        ("Carbon sequestration", "Social cost of carbon", "Non-market"),
        ("Coastal protection", "Replacement cost", "Non-market"),
        ("Nursery habitat", "Productivity change", "Market (indirect)"),
        ("Recreation", "Direct expenditure", "Market"),
        ("Gleaning", "Equivalent wage + market", "Mixed"),
        ("TOTAL", "", ""),
    ]
    for r, (svc, method, vtype) in enumerate(m_services, 5):
        apply_row_header(ws2.cell(row=r, column=1, value=svc))
        apply_data(ws2.cell(row=r, column=2, value=method))
        apply_data(ws2.cell(row=r, column=3, value=vtype))
        if svc == "TOTAL":
            for c in range(4, 9):
                ws2.cell(row=r, column=c).value = f"=SUM({get_column_letter(c)}5:{get_column_letter(c)}10)"
                apply_data(ws2.cell(row=r, column=c), "right")
                ws2.cell(row=r, column=c).font = Font(name="Arial", size=10, bold=True, color=TEXT)
        else:
            for c in range(4, 8):
                apply_input(ws2.cell(row=r, column=c))
            ws2.cell(row=r, column=8).value = f"=SUM(D{r}:G{r})"
            apply_data(ws2.cell(row=r, column=8), "right")

    set_col_widths(ws2, [24, 22, 16, 16, 16, 16, 16, 16])

    # ---- Sheet 3: Carbon Calculator ----
    ws3 = wb.create_sheet("Carbon Calculator")
    ws3["A1"] = "Carbon Sequestration Calculator"
    ws3["A1"].font = title_font
    ws3["A2"] = "Part A: fill extent and NCP rate. Part B: set carbon price."
    ws3["A2"].font = subtitle_font

    section_title(ws3, 4, 1, "Part A: Physical Quantification")
    calc_h = ["Ecosystem", "Extent (ha)", "NCP rate\n(Mg CO2/ha/yr)", "Physical supply\n(Mg CO2/yr)"]
    for c, h in enumerate(calc_h, 1):
        apply_header(ws3.cell(row=5, column=c, value=h))

    for r, (eco, ncp) in enumerate([("Mangroves", 3.7), ("Seagrass", 1.3)], 6):
        apply_row_header(ws3.cell(row=r, column=1, value=eco))
        apply_input(ws3.cell(row=r, column=2))
        ws3.cell(row=r, column=3, value=ncp); apply_data(ws3.cell(row=r, column=3), "right")
        ws3.cell(row=r, column=4).value = f"=B{r}*C{r}"; apply_data(ws3.cell(row=r, column=4), "right")
    apply_row_header(ws3.cell(row=8, column=1, value="Total"))
    ws3.cell(row=8, column=4).value = "=D6+D7"; apply_data(ws3.cell(row=8, column=4), "right")
    ws3.cell(row=8, column=4).font = Font(name="Arial", size=10, bold=True, color=TEXT)

    section_title(ws3, 10, 1, "Part B: Monetary Valuation")
    val_h = ["Ecosystem", "Physical supply\n(Mg CO2/yr)", "Carbon price\n(USD/Mg CO2)", "Value\n(USD/yr)"]
    for c, h in enumerate(val_h, 1):
        apply_header(ws3.cell(row=11, column=c, value=h))
    for r, eco in enumerate(["Mangroves", "Seagrass"], 12):
        apply_row_header(ws3.cell(row=r, column=1, value=eco))
        ws3.cell(row=r, column=2).value = f"=D{r-6}"; apply_data(ws3.cell(row=r, column=2), "right")
        ws3.cell(row=r, column=3, value=51); apply_data(ws3.cell(row=r, column=3), "right")
        ws3.cell(row=r, column=4).value = f"=B{r}*C{r}"; apply_data(ws3.cell(row=r, column=4), "right")
    apply_row_header(ws3.cell(row=14, column=1, value="Total"))
    ws3.cell(row=14, column=4).value = "=D12+D13"
    apply_data(ws3.cell(row=14, column=4), "right")
    ws3.cell(row=14, column=4).font = Font(name="Arial", size=10, bold=True, color=TEXT)

    set_col_widths(ws3, [20, 18, 20, 20])

    # ---- Sheet 4: Resource Rent Calculator ----
    ws4 = wb.create_sheet("Resource Rent Calculator")
    ws4["A1"] = "Fish Provisioning -- Resource Rent Calculator"
    ws4["A1"].font = title_font

    section_title(ws4, 3, 1, "Part A: Physical Quantification")
    for r, lb in enumerate(["Total catch (kg/yr)", "Catch by species 1 (kg/yr)",
                            "Catch by species 2 (kg/yr)", "Catch by species 3 (kg/yr)"], 4):
        ws4.cell(row=r, column=1, value=lb).font = label_font; ws4.cell(row=r, column=1).border = bdr
        apply_input(ws4.cell(row=r, column=2))

    section_title(ws4, 9, 1, "Part B: Monetary Valuation (Resource Rent)")
    for r, lb in enumerate([
        "Average price (USD/kg)", "Gross revenue (USD/yr)", "",
        "Labour costs (USD/yr)", "Fuel costs (USD/yr)",
        "Gear and maintenance (USD/yr)", "Capital depreciation (USD/yr)",
        "Total costs (USD/yr)", "", "Resource rent (USD/yr)"
    ], 10):
        ws4.cell(row=r, column=1, value=lb).font = label_font; ws4.cell(row=r, column=1).border = bdr
        if lb:
            apply_input(ws4.cell(row=r, column=2))

    ws4.cell(row=11, column=2).value = "=B4*B10"  # Revenue
    ws4.cell(row=17, column=2).value = "=SUM(B13:B16)"  # Total costs
    ws4.cell(row=19, column=2).value = "=MAX(B11-B17,0)"  # Rent

    set_col_widths(ws4, [34, 20])

    # ---- Sheet 5: Instructions ----
    ws5 = wb.create_sheet("Instructions")
    ws5["A1"] = "How to use this template"; ws5["A1"].font = title_font
    instructions = [
        "This workbook has two main sheets:",
        "",
        "Part A - Physical Supply: quantify services in physical units (kg, Mg CO2, visitors, m).",
        "   This is a complete account on its own. Fill yellow cells with your data.",
        "",
        "Part B - Monetary Supply: assign economic values (USD/yr) to physical quantities.",
        "   Requires Part A first. Only fill in services where you have economic data.",
        "   It is normal to have Part B for some services and not others.",
        "",
        "Calculator sheets:",
        "   Carbon Calculator: enter extent and NCP rate (Part A), then set carbon price (Part B).",
        "   Resource Rent Calculator: enter catch (Part A), then prices and costs (Part B).",
        "",
        "Yellow cells are for your input. White cells with formulas calculate automatically.",
        "Adapt ecosystem type names and add columns as needed for your area.",
    ]
    for i, txt in enumerate(instructions, 3):
        ws5.cell(row=i, column=1, value=txt).font = data_font
    ws5.column_dimensions["A"].width = 80

    path = Path(__file__).parent / "services" / "services_account_template.xlsx"
    wb.save(path)
    print(f"  Created: {path}")


def generate_services_exercise():
    wb = openpyxl.Workbook()

    # ---- Exercise sheet ----
    ws = wb.active
    ws.title = "Exercise"
    ws.sheet_properties.tabColor = GREEN

    ws["A1"] = "Ecosystem Services Exercise"; ws["A1"].font = title_font
    ws["A2"] = "Part A: Physical quantification | Part B: Monetary valuation"; ws["A2"].font = subtitle_font

    # Given data
    section_title(ws, 4, 1, "Given Data")
    given = [
        ("Coral reef extent", "1,200 ha"), ("Mangrove extent", "350 ha"),
        ("Seagrass extent", "800 ha"), ("", ""),
        ("Reef fish catch", "180,000 kg/yr"), ("Average fish price", "USD 3.50/kg"),
        ("Total fishing costs", "USD 420,000/yr"), ("", ""),
        ("Mangrove NCP rate", "3.7 Mg CO2/ha/yr"), ("Seagrass NCP rate", "1.3 Mg CO2/ha/yr"),
        ("Social cost of carbon", "USD 51/Mg CO2"), ("", ""),
        ("Reef visitors", "15,000/yr"), ("Reef spending per visitor", "USD 85"),
        ("", ""),
        ("Protected coastline", "12,000 m"), ("Seawall cost", "USD 5,000/m"),
        ("Seawall lifespan", "50 years"),
    ]
    for r, (lb, val) in enumerate(given, 5):
        if lb:
            ws.cell(row=r, column=1, value=lb).font = data_font
            ws.cell(row=r, column=1).border = bdr
            ws.cell(row=r, column=2, value=val).font = data_font
            ws.cell(row=r, column=2).border = bdr

    row = 25
    section_title(ws, row, 1, "PART A: Physical Quantification")
    row += 1
    ws.cell(row=row, column=1, value="A1: Carbon (Mg CO2/yr)").font = label_font
    row += 1
    for lb in ["Mangrove: 350 x 3.7 =", "Seagrass: 800 x 1.3 =", "Total carbon:"]:
        ws.cell(row=row, column=1, value=lb).font = data_font; ws.cell(row=row, column=1).border = bdr
        apply_input(ws.cell(row=row, column=2)); row += 1

    row += 1
    ws.cell(row=row, column=1, value="Physical Supply Table (Part A)").font = label_font
    row += 1
    for c, h in enumerate(["Service", "Unit", "Value"], 1):
        apply_header(ws.cell(row=row, column=c, value=h))
    row += 1
    for svc, unit in [("Carbon sequestration", "Mg CO2/yr"), ("Fish provisioning", "kg/yr"),
                       ("Coastal protection", "m"), ("Recreation", "visitors/yr")]:
        apply_row_header(ws.cell(row=row, column=1, value=svc))
        apply_data(ws.cell(row=row, column=2, value=unit))
        apply_input(ws.cell(row=row, column=3))
        row += 1

    row += 2
    section_title(ws, row, 1, "PART B: Monetary Valuation")
    row += 1
    for lb in ["B1: Fish revenue (180,000 x 3.50) =", "B1: Resource rent (revenue - 420,000) =",
               "", "B2: Carbon value (total Mg CO2 x 51) =",
               "", "B3: Recreation (15,000 x 85) =",
               "", "B4: Coastal total (12,000 x 5,000) =", "B4: Annualized (total / 50) ="]:
        if lb:
            ws.cell(row=row, column=1, value=lb).font = data_font
            ws.cell(row=row, column=1).border = bdr
            apply_input(ws.cell(row=row, column=2))
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Monetary Supply Table (Part B)").font = label_font
    row += 1
    for c, h in enumerate(["Service", "Method", "USD/yr"], 1):
        apply_header(ws.cell(row=row, column=c, value=h))
    row += 1
    for svc, method in [("Fish provisioning", "Resource rent"), ("Carbon sequestration", "SCC"),
                         ("Recreation", "Direct expenditure"), ("Coastal protection", "Replacement cost"),
                         ("TOTAL", "")]:
        apply_row_header(ws.cell(row=row, column=1, value=svc))
        apply_data(ws.cell(row=row, column=2, value=method))
        apply_input(ws.cell(row=row, column=3))
        row += 1

    set_col_widths(ws, [40, 20, 20])

    # ---- Answer sheet ----
    ws_a = wb.create_sheet("Answers")
    ws_a.sheet_properties.tabColor = "006100"
    ws_a["A1"] = "Answer Key"; ws_a["A1"].font = title_font

    section_title(ws_a, 3, 1, "Part A: Physical Supply")
    answers_a = [
        ("Carbon (mangrove)", "1,295 Mg CO2/yr"),
        ("Carbon (seagrass)", "1,040 Mg CO2/yr"),
        ("Carbon total", "2,335 Mg CO2/yr"),
        ("Fish provisioning", "180,000 kg/yr"),
        ("Coastal protection", "12,000 m"),
        ("Recreation", "15,000 visitors/yr"),
    ]
    for r, (lb, val) in enumerate(answers_a, 4):
        ws_a.cell(row=r, column=1, value=lb).font = label_font; ws_a.cell(row=r, column=1).border = bdr
        cell = ws_a.cell(row=r, column=2, value=val); apply_answer(cell)

    section_title(ws_a, 12, 1, "Part B: Monetary Valuation")
    answers_b = [
        ("Fish: gross revenue", "USD 630,000"),
        ("Fish: resource rent", "USD 210,000"),
        ("Carbon: mangrove", "USD 66,045"),
        ("Carbon: seagrass", "USD 53,040"),
        ("Carbon: total", "USD 119,085"),
        ("Recreation", "USD 1,275,000"),
        ("Coastal: total replacement", "USD 60,000,000"),
        ("Coastal: annualized", "USD 1,200,000"),
        ("", ""),
        ("TOTAL (all services)", "USD 2,804,085"),
    ]
    for r, (lb, val) in enumerate(answers_b, 13):
        if lb:
            ws_a.cell(row=r, column=1, value=lb).font = label_font; ws_a.cell(row=r, column=1).border = bdr
            cell = ws_a.cell(row=r, column=2, value=val); apply_answer(cell)

    set_col_widths(ws_a, [34, 24])

    path = Path(__file__).parent / "services" / "services_exercise.xlsx"
    wb.save(path)
    print(f"  Created: {path}")


if __name__ == "__main__":
    print("Regenerating services files with Part A / Part B structure...")
    generate_services_template()
    generate_services_exercise()
    print("Done!")
