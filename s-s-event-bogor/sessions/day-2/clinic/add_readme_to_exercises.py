#!/usr/bin/env python3
"""Add README and Glossary sheets to all exercise and template Excel files.

Fixes:
- Adds a README tab explaining the exercise purpose and how to use the file
- Adds a Glossary tab defining all abbreviations and technical terms
- Updates abbreviated units (M → million, IC → Intermediate Consumption, etc.)
"""

from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

GREEN = "3B9C7B"
TEAL = "0A5455"
GRAY = "404040"
TEXT = "30302F"
WHITE = "FFFFFF"

green_fill = PatternFill(start_color=GREEN, end_color=GREEN, fill_type="solid")
teal_fill = PatternFill(start_color=TEAL, end_color=TEAL, fill_type="solid")
bdr = Border(left=Side("thin", GRAY), right=Side("thin", GRAY),
             top=Side("thin", GRAY), bottom=Side("thin", GRAY))
hdr_font = Font(name="Arial", size=10, bold=True, color=WHITE)
data_font = Font(name="Arial", size=10, color=TEXT)
label_font = Font(name="Arial", size=10, bold=True, color=TEAL)
title_font = Font(name="Arial", size=14, bold=True, color=TEAL)
subtitle_font = Font(name="Arial", size=11, color=GREEN)


def add_readme_sheet(wb, account_type, is_exercise=True):
    """Add a README sheet as the first sheet."""
    ws = wb.create_sheet("README", 0)
    ws.sheet_properties.tabColor = TEAL

    ws["A1"] = f"Ocean Accounting Clinic: {account_type}"
    ws["A1"].font = title_font

    if is_exercise:
        ws["A2"] = "Exercise Workbook"
        ws["A2"].font = subtitle_font
        ws["A4"] = "How to use this file"
        ws["A4"].font = Font(name="Arial", size=12, bold=True, color=TEAL)

        instructions = {
            "Extent": [
                "This exercise helps you practice building an ecosystem extent account.",
                "",
                "The Exercise sheet contains:",
                "  - Given data: ecosystem areas for two time periods and a change matrix",
                "  - Yellow cells: fill these in with your calculations",
                "  - You will calculate: additions, reductions, net change, closing extent, and percentage change",
                "",
                "The Answers sheet shows the correct values in green cells.",
                "",
                "What you are calculating:",
                "  - Additions: area GAINED by each ecosystem type (from other types converting to it)",
                "  - Reductions: area LOST by each ecosystem type (converting to other types)",
                "  - Net change: additions minus reductions",
                "  - Closing extent: opening extent plus net change",
                "  - The total accounting area must be the same for opening and closing (area is conserved)",
            ],
            "Condition": [
                "This exercise helps you practice normalizing ecological indicators to a Condition Index (CI).",
                "",
                "The Exercise sheet contains:",
                "  - Reference levels for each indicator (the benchmark for 'healthy')",
                "  - Raw monitoring data from 4 sites for two time periods",
                "  - Yellow cells: fill these in with your CI calculations",
                "",
                "The Answers sheet shows the correct values in green cells.",
                "",
                "What you are calculating:",
                "  - Condition Index (CI): a score from 0 (degraded) to 1 (reference condition)",
                "  - For 'higher-is-better' indicators (e.g., coral cover): CI = Measured / Reference",
                "  - For 'higher-is-worse' indicators (e.g., macroalgae): CI = 1 - (Measured / Reference)",
                "  - CI is capped between 0 and 1",
                "  - Mean CI: average across all sites for each indicator",
            ],
            "Services": [
                "This exercise helps you practice calculating ecosystem service values.",
                "",
                "The exercise is split into two parts:",
                "  Part A (Physical Quantification): calculate how much service the ecosystem provides",
                "  Part B (Monetary Valuation): assign economic values to physical quantities",
                "  Part A is a complete account on its own. Part B builds on Part A.",
                "",
                "Yellow cells: fill these in with your calculations.",
                "The Answers sheet shows the correct values in green cells.",
                "",
                "What you are calculating:",
                "  - Physical supply: quantities in natural units (kilograms, tonnes CO2, metres, visitors)",
                "  - Monetary value: economic values in USD per year",
                "  - Resource rent: revenue minus all costs (labour, fuel, gear, capital)",
                "  - SCC value: physical carbon sequestration multiplied by the Social Cost of Carbon price",
            ],
            "OESA": [
                "This exercise helps you practice compiling an Ocean Economy Satellite Account.",
                "",
                "The Exercise sheet contains:",
                "  - National economic data for 6 industries (from Supply-Use Tables)",
                "  - Ocean economy partials (the share of each industry that is ocean-related)",
                "  - Yellow cells: fill these in with your calculations",
                "",
                "The Answers sheet shows the correct values in green cells.",
                "",
                "What you are calculating:",
                "  - Ocean Output: the ocean-related portion of each industry's total output",
                "    Formula: Ocean Output = National Output x Ocean Partial",
                "  - Ocean IC (Intermediate Consumption): the cost of inputs used in ocean production",
                "    Formula: Ocean IC = National IC x Ocean Partial",
                "  - Ocean GVA (Gross Value Added): the value the ocean economy adds",
                "    Formula: Ocean GVA = Ocean Output - Ocean IC",
                "  - Ocean Economy % of GDP: how big the ocean economy is relative to the whole economy",
                "    Formula: Total Ocean GVA / National GDP x 100",
            ],
            "OTSA": [
                "This exercise helps you practice calculating an Ocean Tourism Satellite Account.",
                "",
                "The Exercise sheet contains:",
                "  - National tourism data (arrivals, expenditure, GVA, employment)",
                "  - Visitor survey data on coastal/marine motivation",
                "  - Yellow cells: fill these in with your calculations",
                "",
                "The Answers sheet shows the correct values in green cells.",
                "",
                "What you are calculating:",
                "  - Tourism partial: the share of total tourism that is coastal/ocean-related",
                "  - Weighted partial: accounts for different spending levels between visitor types",
                "    (international visitors spend 2.5x more per trip than domestic visitors)",
                "  - Coastal tourism indicators: arrivals, expenditure, GVA, and employment",
                "    Formula: Coastal value = National value x Tourism partial",
            ],
        }

        for r, line in enumerate(instructions.get(account_type, ["No specific instructions."]), 5):
            ws.cell(row=r, column=1, value=line).font = data_font
    else:
        ws["A2"] = "Account Template"
        ws["A2"].font = subtitle_font
        ws["A4"] = "How to use this file"
        ws["A4"].font = Font(name="Arial", size=12, bold=True, color=TEAL)
        template_instructions = [
            "Fill in the yellow cells with your country's data.",
            "White cells with formulas will calculate automatically.",
            "See the Glossary sheet for definitions of all terms and abbreviations.",
            "See the Instructions sheet (if present) for detailed step-by-step guidance.",
        ]
        for r, line in enumerate(template_instructions, 5):
            ws.cell(row=r, column=1, value=line).font = data_font

    ws.column_dimensions["A"].width = 80
    return ws


def add_glossary_sheet(wb, account_type):
    """Add a Glossary sheet defining all terms."""
    ws = wb.create_sheet("Glossary")
    ws.sheet_properties.tabColor = GREEN

    ws["A1"] = "Glossary of Terms and Abbreviations"
    ws["A1"].font = title_font

    for c, h in enumerate(["Term / Abbreviation", "Full Name", "Definition", "Unit"], 1):
        cell = ws.cell(row=3, column=c, value=h)
        cell.font = hdr_font
        cell.fill = green_fill
        cell.border = bdr
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Common terms
    terms = [
        ("GVA", "Gross Value Added", "Total output minus intermediate consumption. The value an industry adds to the economy.", "Currency (e.g., USD, local currency)"),
        ("GDP", "Gross Domestic Product", "Total value of goods and services produced in a country. Sum of GVA across all industries plus taxes minus subsidies.", "Currency"),
        ("IC", "Intermediate Consumption", "The value of goods and services consumed as inputs during production (e.g., fuel, raw materials, purchased services). Does not include fixed assets.", "Currency"),
        ("Output", "Gross Output", "Total value of goods and services produced by an industry before deducting input costs. Also called gross sales or turnover.", "Currency"),
        ("SUT", "Supply-Use Tables", "Statistical tables from the national accounts showing what each industry produces (supply) and what each industry consumes (use). Published by the National Statistical Office.", ""),
        ("SNA", "System of National Accounts", "The international statistical standard for measuring economic activity. Published by the UN. The basis for GDP calculation.", ""),
        ("SEEA-EA", "System of Environmental-Economic Accounting - Ecosystem Accounting", "The international statistical standard for measuring ecosystems, their condition, and the services they provide.", ""),
        ("OESA", "Ocean Economy Satellite Account", "A satellite account that reorganizes national accounts to identify the market-based ocean economy. Reports GDP-compatible indicators.", ""),
        ("OTSA", "Ocean Tourism Satellite Account", "A satellite account focused on coastal and marine tourism's economic contribution. A subset of OESA.", ""),
        ("Partial", "Ocean Economy Partial", "The share (0 to 1) of a mixed industry's output that is ocean-related. For example, if 30% of construction is coastal, the partial is 0.30.", "Proportion (0 to 1)"),
        ("CI", "Condition Index", "A normalized score from 0 (fully degraded) to 1 (reference/healthy condition). Used in SEEA-EA condition accounts.", "Index (0 to 1)"),
        ("NCP", "Net Carbon Production", "The rate at which an ecosystem sequesters carbon dioxide per unit area per year.", "Mg CO2/ha/yr"),
        ("SCC", "Social Cost of Carbon", "An estimate of the economic damage caused by emitting one additional tonne of CO2. Used to value carbon sequestration.", "USD per Mg CO2"),
        ("Mg CO2", "Megagrams of CO2", "Metric tonnes of carbon dioxide. 1 Mg = 1 tonne = 1,000 kg.", "Mass"),
        ("ha", "Hectares", "Unit of area. 1 hectare = 10,000 square metres = 0.01 square kilometres.", "Area"),
        ("kg/yr", "Kilograms per year", "Physical quantity measured in kilograms over one year.", "Mass per time"),
        ("USD million", "United States Dollars, millions", "Monetary values expressed in millions of US dollars. For example, USD 132.5 million = USD 132,500,000.", "Currency"),
        ("R million", "South African Rand, millions", "Monetary values in the South Africa example are in millions of Rand.", "Currency"),
        ("ISIC", "International Standard Industrial Classification", "The UN classification system for economic activities. Used to map industries to ocean economy groups.", ""),
        ("SIC", "Standard Industrial Classification", "National classification system for economic activities. Varies by country but based on ISIC.", ""),
        ("EEZ", "Exclusive Economic Zone", "Maritime zone extending 200 nautical miles from a country's coastline, within which the country has rights to marine resources.", ""),
        ("RAG", "Red-Amber-Green", "A confidence rating system. Green = high confidence (direct evidence). Amber = moderate (proxy data). Red = low (sparse evidence).", ""),
    ]

    for r, (term, full, defn, unit) in enumerate(terms, 4):
        cell_t = ws.cell(row=r, column=1, value=term)
        cell_t.font = Font(name="Arial", size=10, bold=True, color=TEAL)
        cell_t.border = bdr
        ws.cell(row=r, column=2, value=full).font = data_font
        ws.cell(row=r, column=2).border = bdr
        ws.cell(row=r, column=3, value=defn).font = data_font
        ws.cell(row=r, column=3).border = bdr
        ws.cell(row=r, column=3).alignment = Alignment(wrap_text=True)
        ws.cell(row=r, column=4, value=unit).font = data_font
        ws.cell(row=r, column=4).border = bdr

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 60
    ws.column_dimensions["D"].width = 20

    return ws


def fix_units_in_sheet(ws):
    """Replace abbreviated units with clear versions throughout a sheet."""
    replacements = {
        "(M)": "(USD million)",
        " M)": " million)",
        "(R'million": "(Rand million",
        "Mg CO2": "tonnes CO2",
    }
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                for old, new in replacements.items():
                    if old in cell.value:
                        cell.value = cell.value.replace(old, new)


def process_file(filepath, account_type, is_exercise=True):
    """Add README and Glossary to an existing Excel file."""
    wb = openpyxl.load_workbook(filepath)

    # Remove existing README/Glossary if they exist (to avoid duplicates)
    for name in ["README", "Glossary"]:
        if name in wb.sheetnames:
            del wb[name]

    # Fix units in all existing sheets
    for ws_name in wb.sheetnames:
        fix_units_in_sheet(wb[ws_name])

    # Add new sheets
    add_readme_sheet(wb, account_type, is_exercise)
    add_glossary_sheet(wb, account_type)

    wb.save(filepath)
    print(f"  Updated: {filepath.name}")


def main():
    base = Path("/Users/mariaalarcon/GitHub/ocean-accounts-expert/s-s-event-bogor/sessions/day-2/clinic")

    files = [
        (base / "extent/extent_exercise.xlsx", "Extent", True),
        (base / "extent/extent_account_template.xlsx", "Extent", False),
        (base / "condition/condition_exercise.xlsx", "Condition", True),
        (base / "condition/condition_account_template.xlsx", "Condition", False),
        (base / "services/services_exercise.xlsx", "Services", True),
        (base / "services/services_account_template.xlsx", "Services", False),
        (base / "economic/oesa_exercise.xlsx", "OESA", True),
        (base / "economic/oesa_template.xlsx", "OESA", False),
        (base / "tourism/otsa_exercise.xlsx", "OTSA", True),
        (base / "tourism/otsa_template.xlsx", "OTSA", False),
    ]

    print("Adding README and Glossary sheets to all Excel files...")
    for filepath, account_type, is_exercise in files:
        if filepath.exists():
            process_file(filepath, account_type, is_exercise)
        else:
            print(f"  SKIPPED (not found): {filepath}")

    print("\nDone! All files updated.")


if __name__ == "__main__":
    main()
