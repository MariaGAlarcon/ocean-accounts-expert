#!/usr/bin/env python3
"""Generate GOAP-styled waste account Excel template and exercise."""

from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

GREEN = "3B9C7B"
TEAL = "0A5455"
GRAY = "404040"
TEXT = "30302F"
WHITE = "FFFFFF"
YELLOW = "FFF2CC"
RED_LIGHT = "FCE4EC"

green_fill = PatternFill(start_color=GREEN, end_color=GREEN, fill_type="solid")
teal_fill = PatternFill(start_color=TEAL, end_color=TEAL, fill_type="solid")
yellow_fill = PatternFill(start_color=YELLOW, end_color=YELLOW, fill_type="solid")
bdr = Border(left=Side("thin", GRAY), right=Side("thin", GRAY),
             top=Side("thin", GRAY), bottom=Side("thin", GRAY))
hdr_font = Font(name="Arial", size=10, bold=True, color=WHITE)
data_font = Font(name="Arial", size=10, color=TEXT)
label_font = Font(name="Arial", size=10, bold=True, color=TEAL)
title_font = Font(name="Arial", size=14, bold=True, color=TEAL)
subtitle_font = Font(name="Arial", size=11, color=GREEN)
answer_font = Font(name="Arial", size=10, bold=True, color="006100")
answer_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")


def apply_header(cell):
    cell.font = hdr_font; cell.fill = green_fill; cell.border = bdr
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def apply_row_header(cell):
    cell.font = Font(name="Arial", size=10, bold=True, color=WHITE)
    cell.fill = teal_fill; cell.border = bdr
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

def apply_data(cell, align="right"):
    cell.font = data_font; cell.border = bdr
    cell.alignment = Alignment(horizontal=align, vertical="center")

def apply_input(cell):
    cell.fill = yellow_fill; cell.font = Font(name="Arial", size=10, color=TEXT); cell.border = bdr
    cell.alignment = Alignment(horizontal="right", vertical="center")

def apply_answer(cell):
    cell.font = answer_font; cell.fill = answer_fill; cell.border = bdr
    cell.alignment = Alignment(horizontal="right", vertical="center")

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def section_title(ws, row, col, text):
    ws.cell(row=row, column=col, value=text).font = Font(
        name="Arial", size=12, bold=True, color=TEAL)


def generate_waste_template():
    wb = openpyxl.Workbook()

    # README
    ws0 = wb.active
    ws0.title = "README"
    ws0.sheet_properties.tabColor = TEAL
    ws0["A1"] = "Ocean Waste and Emissions Account Template"
    ws0["A1"].font = title_font
    ws0["A2"] = "SEEA-CF Physical Flow Accounts for the Marine Environment"
    ws0["A2"].font = subtitle_font
    instructions = [
        "",
        "This template covers three types of accounts tracking what the economy puts into the ocean:",
        "",
        "1. Water Emissions Account: substances discharged to coastal waters and seas",
        "   (nutrients, chemicals, sewage, heavy metals)",
        "",
        "2. Solid Waste Account: waste entering the marine environment",
        "   (plastics, fishing gear, ship waste, construction debris)",
        "",
        "3. Air Emissions Account: greenhouse gases and pollutants from ocean industries",
        "   (shipping, fishing fleets, offshore energy, port operations)",
        "",
        "How to use:",
        "  - Fill in yellow cells with your country's data",
        "  - White cells with formulas calculate automatically",
        "  - Each sheet has a 'Source' column to document where your data comes from",
        "  - See the Glossary sheet for definitions of all terms",
        "",
        "These accounts complement SEEA-EA (ecosystem health) and OESA (ocean economy).",
        "Together they show: what the ocean provides (SEEA-EA), what economic activity it supports",
        "(OESA), and what pressures the economy puts on it (these waste accounts).",
    ]
    for r, line in enumerate(instructions, 4):
        ws0.cell(row=r, column=1, value=line).font = data_font
    ws0.column_dimensions["A"].width = 80

    # Sheet 1: Water Emissions
    ws1 = wb.create_sheet("Water Emissions")
    ws1.sheet_properties.tabColor = GREEN

    ws1["A1"] = "Water Emissions Account: Substances Discharged to Sea"
    ws1["A1"].font = title_font
    ws1["A2"] = "Country:"; ws1["A2"].font = label_font
    ws1["B2"].fill = yellow_fill; ws1["B2"].border = bdr
    ws1["A3"] = "Year:"; ws1["A3"].font = label_font
    ws1["B3"].fill = yellow_fill; ws1["B3"].border = bdr

    section_title(ws1, 5, 1, "Supply: Generation of Water Emissions by Source (tonnes/year)")
    w_headers = ["Source sector", "Nitrogen\n(tonnes N/yr)", "Phosphorus\n(tonnes P/yr)",
                 "BOD\n(tonnes/yr)", "Heavy metals\n(tonnes/yr)", "Other\nsubstances\n(tonnes/yr)",
                 "Total\n(tonnes/yr)", "Data source"]
    for c, h in enumerate(w_headers, 1):
        apply_header(ws1.cell(row=6, column=c, value=h))

    sources = ["Agriculture (runoff)", "Manufacturing", "Mining and quarrying",
               "Sewerage (treated effluent)", "Sewerage (untreated discharge)",
               "Households (direct)", "Other industries", "TOTAL"]
    for r, s in enumerate(sources, 7):
        apply_row_header(ws1.cell(row=r, column=1, value=s))
        for c in range(2, 8):
            if s == "TOTAL":
                ws1.cell(row=r, column=c).value = f"=SUM({get_column_letter(c)}7:{get_column_letter(c)}13)"
                apply_data(ws1.cell(row=r, column=c))
                ws1.cell(row=r, column=c).font = Font(name="Arial", size=10, bold=True, color=TEXT)
            else:
                apply_input(ws1.cell(row=r, column=c))
        apply_input(ws1.cell(row=r, column=8))

    section_title(ws1, 16, 1, "Use: Destination of Emissions")
    d_headers = ["Destination", "Nitrogen", "Phosphorus", "BOD", "Heavy metals", "Other", "Total"]
    for c, h in enumerate(d_headers, 1):
        apply_header(ws1.cell(row=17, column=c, value=h))
    destinations = ["To sewerage treatment", "Direct discharge to rivers",
                    "Direct discharge to sea", "Non-point source to sea", "TOTAL"]
    for r, d in enumerate(destinations, 18):
        apply_row_header(ws1.cell(row=r, column=1, value=d))
        for c in range(2, 8):
            if d == "TOTAL":
                ws1.cell(row=r, column=c).value = f"=SUM({get_column_letter(c)}18:{get_column_letter(c)}21)"
                apply_data(ws1.cell(row=r, column=c))
            else:
                apply_input(ws1.cell(row=r, column=c))

    set_col_widths(ws1, [28, 14, 14, 12, 14, 14, 12, 24])

    # Sheet 2: Solid Waste
    ws2 = wb.create_sheet("Solid Waste")
    ws2.sheet_properties.tabColor = GREEN

    ws2["A1"] = "Solid Waste Account: Waste Entering the Marine Environment"
    ws2["A1"].font = title_font
    ws2["A2"] = "Country:"; ws2["A2"].font = label_font
    ws2["B2"].fill = yellow_fill; ws2["B2"].border = bdr
    ws2["A3"] = "Year:"; ws2["A3"].font = label_font
    ws2["B3"].fill = yellow_fill; ws2["B3"].border = bdr

    section_title(ws2, 5, 1, "Solid Waste to Marine Environment (tonnes/year)")
    sw_headers = ["Waste type", "Land-based\nsources\n(tonnes/yr)", "Sea-based\nsources\n(tonnes/yr)",
                  "Total\n(tonnes/yr)", "Percentage\nof total", "Data source"]
    for c, h in enumerate(sw_headers, 1):
        apply_header(ws2.cell(row=6, column=c, value=h))

    waste_types = ["Plastics (macroplastic)", "Plastics (microplastic)",
                   "Fishing gear (ghost nets, lines)", "Ship-generated waste",
                   "Sewage sludge", "Dredging material", "Construction debris",
                   "Other solid waste", "TOTAL"]
    for r, w in enumerate(waste_types, 7):
        apply_row_header(ws2.cell(row=r, column=1, value=w))
        for c in range(2, 5):
            if w == "TOTAL":
                ws2.cell(row=r, column=c).value = f"=SUM({get_column_letter(c)}7:{get_column_letter(c)}14)"
                apply_data(ws2.cell(row=r, column=c))
            else:
                apply_input(ws2.cell(row=r, column=c))
        # Total = land + sea
        if w != "TOTAL":
            ws2.cell(row=r, column=4).value = f"=B{r}+C{r}"
            apply_data(ws2.cell(row=r, column=4))
        # Percentage
        ws2.cell(row=r, column=5).value = f'=IF(D$15=0,"",D{r}/D$15)'
        apply_data(ws2.cell(row=r, column=5))
        ws2.cell(row=r, column=5).number_format = "0.0%"
        apply_input(ws2.cell(row=r, column=6))

    section_title(ws2, 17, 1, "Land-Based Sources Estimation (Jambeck method)")
    calc_labels = [
        ("National solid waste generated (tonnes/yr)", ""),
        ("Mismanaged waste rate (%)", ""),
        ("Coastal population share (%)", ""),
        ("Marine leakage rate (%)", ""),
        ("Estimated land-based marine waste (tonnes/yr)", "=B18*B19*B20*B21"),
    ]
    for r, (lb, formula) in enumerate(calc_labels, 18):
        ws2.cell(row=r, column=1, value=lb).font = label_font
        ws2.cell(row=r, column=1).border = bdr
        if formula:
            ws2.cell(row=r, column=2, value=formula)
            apply_data(ws2.cell(row=r, column=2))
        else:
            apply_input(ws2.cell(row=r, column=2))

    set_col_widths(ws2, [32, 16, 16, 14, 12, 24])

    # Sheet 3: Air Emissions
    ws3 = wb.create_sheet("Air Emissions")
    ws3.sheet_properties.tabColor = GREEN

    ws3["A1"] = "Air Emissions Account: Ocean Industry Emissions"
    ws3["A1"].font = title_font

    ae_headers = ["Ocean sector", "Fuel consumption\n(tonnes/yr)", "CO2 emission\nfactor",
                  "CO2 emissions\n(tonnes CO2/yr)", "SO2 emissions\n(tonnes/yr)",
                  "NOx emissions\n(tonnes/yr)", "Data source"]
    for c, h in enumerate(ae_headers, 1):
        apply_header(ws3.cell(row=3, column=c, value=h))

    sectors = ["Maritime shipping (international)", "Maritime shipping (domestic)",
               "Fishing fleet", "Offshore oil and gas", "Port operations",
               "Marine construction", "Other ocean industries", "TOTAL"]
    for r, s in enumerate(sectors, 4):
        apply_row_header(ws3.cell(row=r, column=1, value=s))
        if s == "TOTAL":
            for c in range(2, 7):
                ws3.cell(row=r, column=c).value = f"=SUM({get_column_letter(c)}4:{get_column_letter(c)}10)"
                apply_data(ws3.cell(row=r, column=c))
        else:
            apply_input(ws3.cell(row=r, column=2))
            ws3.cell(row=r, column=3, value=3.2)
            apply_data(ws3.cell(row=r, column=3))
            ws3.cell(row=r, column=4).value = f"=B{r}*C{r}"
            apply_data(ws3.cell(row=r, column=4))
            for c in [5, 6]:
                apply_input(ws3.cell(row=r, column=c))
        apply_input(ws3.cell(row=r, column=7))

    set_col_widths(ws3, [32, 16, 14, 18, 14, 14, 24])

    # Glossary
    ws_g = wb.create_sheet("Glossary")
    ws_g.sheet_properties.tabColor = GREEN
    ws_g["A1"] = "Glossary"
    ws_g["A1"].font = title_font
    terms = [
        ("BOD", "Biochemical Oxygen Demand", "Amount of oxygen consumed by microorganisms to decompose organic matter in water. Higher BOD indicates more organic pollution.", "mg/L or tonnes/yr"),
        ("Nitrogen (N)", "Total nitrogen load", "Nitrogen from fertilizers, sewage, and industrial discharge that causes eutrophication (algal blooms) in coastal waters.", "tonnes N/yr"),
        ("Phosphorus (P)", "Total phosphorus load", "Phosphorus from detergents, fertilizers, and sewage that causes eutrophication.", "tonnes P/yr"),
        ("Heavy metals", "Toxic metal pollutants", "Mercury, lead, cadmium, arsenic, and other metals from industrial discharge, mining, and fuel combustion.", "tonnes/yr"),
        ("Mismanaged waste", "Waste not properly collected or disposed", "Waste that is dumped, littered, or sent to uncontrolled disposal sites. Calculated as a percentage of total waste generated.", "% or tonnes/yr"),
        ("Marine leakage", "Waste entering the ocean", "The fraction of mismanaged coastal waste that reaches the marine environment through rivers, wind, tides, or direct dumping.", "% or tonnes/yr"),
        ("Ghost gear", "Abandoned fishing equipment", "Lost, abandoned, or discarded fishing nets, lines, and traps that continue to catch marine life.", "tonnes/yr"),
        ("CO2e", "CO2 equivalent", "A standard unit for measuring greenhouse gases. Converts CH4, N2O, etc. to their warming equivalent in CO2.", "tonnes CO2e"),
        ("Emission factor", "Emissions per unit fuel", "Amount of pollutant released per unit of fuel consumed. For marine fuel: approximately 3.2 tonnes CO2 per tonne of fuel.", "tonnes CO2 per tonne fuel"),
        ("Point source", "Identifiable discharge location", "Pollution from a specific, identifiable outlet: factory pipe, sewage plant, ship discharge point.", ""),
        ("Non-point source", "Diffuse pollution", "Pollution without a single origin: agricultural runoff, urban stormwater, atmospheric deposition.", ""),
        ("MARPOL", "International Convention for the Prevention of Pollution from Ships", "IMO convention regulating ship-generated waste, oil, chemicals, sewage, and garbage discharge at sea.", ""),
    ]
    for c, h in enumerate(["Term", "Full Name", "Definition", "Unit"], 1):
        apply_header(ws_g.cell(row=3, column=c, value=h))
    for r, (term, full, defn, unit) in enumerate(terms, 4):
        ws_g.cell(row=r, column=1, value=term).font = Font(name="Arial", size=10, bold=True, color=TEAL)
        ws_g.cell(row=r, column=1).border = bdr
        ws_g.cell(row=r, column=2, value=full).font = data_font; ws_g.cell(row=r, column=2).border = bdr
        ws_g.cell(row=r, column=3, value=defn).font = data_font; ws_g.cell(row=r, column=3).border = bdr
        ws_g.cell(row=r, column=3).alignment = Alignment(wrap_text=True)
        ws_g.cell(row=r, column=4, value=unit).font = data_font; ws_g.cell(row=r, column=4).border = bdr

    set_col_widths(ws_g, [18, 30, 60, 20])

    path = Path(__file__).parent / "waste_account_template.xlsx"
    wb.save(path)
    print(f"  Created: {path}")


def generate_waste_exercise():
    wb = openpyxl.Workbook()

    # README
    ws0 = wb.active
    ws0.title = "README"
    ws0.sheet_properties.tabColor = TEAL
    ws0["A1"] = "Ocean Waste and Emissions Exercise"
    ws0["A1"].font = title_font
    ws0["A2"] = "Practice calculating pollution loads reaching the marine environment"
    ws0["A2"].font = subtitle_font
    readme = [
        "",
        "This exercise covers three types of ocean pollution:",
        "  Part 1: Nutrient loads to sea (water emissions)",
        "  Part 2: Solid waste entering the ocean (plastic and debris)",
        "  Part 3: Air emissions from ocean industries (CO2)",
        "",
        "What you will calculate:",
        "  - Nitrogen and phosphorus loads from untreated wastewater",
        "    Formula: Volume (m3) x Untreated share x Concentration (mg/L) / 1,000,000 = tonnes/yr",
        "  - Solid waste to sea using the Jambeck method",
        "    Formula: National waste x Mismanaged rate x Coastal share x Marine leakage rate",
        "  - CO2 emissions from shipping and fishing",
        "    Formula: Fuel consumption (tonnes) x Emission factor (3.2 tonnes CO2/tonne fuel)",
        "",
        "Fill yellow cells with your calculations. Check your answers on the Answers sheet.",
    ]
    for r, line in enumerate(readme, 4):
        ws0.cell(row=r, column=1, value=line).font = data_font
    ws0.column_dimensions["A"].width = 75

    # Exercise sheet
    ws = wb.create_sheet("Exercise")
    ws.sheet_properties.tabColor = GREEN
    ws["A1"] = "Ocean Waste and Emissions Exercise"
    ws["A1"].font = title_font

    section_title(ws, 3, 1, "Given Data")
    given = [
        ("Total national wastewater (substances)", "500,000 tonnes/yr"),
        ("Coastal share of wastewater", "35%"),
        ("Treatment rate (coastal sewerage)", "60%"),
        ("Untreated share (discharged to sea)", "40%"),
        ("Coastal wastewater volume", "200 million m3/yr"),
        ("Nitrogen concentration (untreated)", "15 mg/L (= 15 g/m3)"),
        ("Phosphorus concentration (untreated)", "3 mg/L (= 3 g/m3)"),
        ("", ""),
        ("National solid waste generated", "2,000,000 tonnes/yr"),
        ("Mismanaged waste rate", "25%"),
        ("Population within 50 km of coast", "40% of national"),
        ("Marine leakage rate", "15%"),
        ("", ""),
        ("Fishing fleet fuel consumption", "50,000 tonnes/yr"),
        ("Domestic shipping fuel", "200,000 tonnes/yr"),
        ("CO2 emission factor (marine fuel)", "3.2 tonnes CO2 per tonne fuel"),
    ]
    for r, (lb, val) in enumerate(given, 4):
        if lb:
            ws.cell(row=r, column=1, value=lb).font = data_font; ws.cell(row=r, column=1).border = bdr
            ws.cell(row=r, column=2, value=val).font = data_font; ws.cell(row=r, column=2).border = bdr

    row = 22
    section_title(ws, row, 1, "Part 1: Nutrient Load to Sea (Water Emissions)")
    row += 1
    for lb in [
        "Untreated wastewater volume = 200M m3 x 0.40 =",
        "Nitrogen load = untreated volume x 15 g/m3 / 1,000,000 =",
        "Phosphorus load = untreated volume x 3 g/m3 / 1,000,000 =",
    ]:
        ws.cell(row=row, column=1, value=lb).font = data_font; ws.cell(row=row, column=1).border = bdr
        apply_input(ws.cell(row=row, column=2)); row += 1

    row += 1
    section_title(ws, row, 1, "Part 2: Solid Waste to Sea (Jambeck Method)")
    row += 1
    for lb in [
        "Mismanaged waste = 2,000,000 x 0.25 =",
        "Coastal mismanaged = mismanaged x 0.40 =",
        "Marine waste = coastal mismanaged x 0.15 =",
    ]:
        ws.cell(row=row, column=1, value=lb).font = data_font; ws.cell(row=row, column=1).border = bdr
        apply_input(ws.cell(row=row, column=2)); row += 1

    row += 1
    section_title(ws, row, 1, "Part 3: Air Emissions from Ocean Industries")
    row += 1
    for lb in [
        "Total ocean fuel = fishing + shipping =",
        "CO2 emissions = total fuel x 3.2 =",
    ]:
        ws.cell(row=row, column=1, value=lb).font = data_font; ws.cell(row=row, column=1).border = bdr
        apply_input(ws.cell(row=row, column=2)); row += 1

    row += 1
    section_title(ws, row, 1, "Part 4: Summary")
    row += 1
    for c, h in enumerate(["Pollution type", "Quantity", "Unit"], 1):
        apply_header(ws.cell(row=row, column=c, value=h))
    row += 1
    for pol in ["Nitrogen to sea", "Phosphorus to sea", "Solid waste to sea", "CO2 from ocean industries"]:
        apply_row_header(ws.cell(row=row, column=1, value=pol))
        apply_input(ws.cell(row=row, column=2))
        apply_input(ws.cell(row=row, column=3))
        row += 1

    set_col_widths(ws, [52, 24, 20])

    # Answers
    ws_a = wb.create_sheet("Answers")
    ws_a.sheet_properties.tabColor = "006100"
    ws_a["A1"] = "Answer Key"
    ws_a["A1"].font = title_font

    answers = [
        ("Part 1: Water Emissions", "", ""),
        ("Untreated wastewater volume", "80,000,000 m3/yr", "(200M x 0.40)"),
        ("Nitrogen load to sea", "1,200 tonnes N/yr", "(80M m3 x 15 g/m3 / 1,000,000)"),
        ("Phosphorus load to sea", "240 tonnes P/yr", "(80M m3 x 3 g/m3 / 1,000,000)"),
        ("", "", ""),
        ("Part 2: Solid Waste", "", ""),
        ("Mismanaged waste", "500,000 tonnes/yr", "(2,000,000 x 0.25)"),
        ("Coastal mismanaged", "200,000 tonnes/yr", "(500,000 x 0.40)"),
        ("Marine solid waste", "30,000 tonnes/yr", "(200,000 x 0.15)"),
        ("", "", ""),
        ("Part 3: Air Emissions", "", ""),
        ("Total ocean fuel", "250,000 tonnes/yr", "(50,000 + 200,000)"),
        ("CO2 from ocean industries", "800,000 tonnes CO2/yr", "(250,000 x 3.2)"),
        ("", "", ""),
        ("Summary", "", ""),
        ("Nitrogen to sea", "1,200 tonnes N/yr", ""),
        ("Phosphorus to sea", "240 tonnes P/yr", ""),
        ("Solid waste to sea", "30,000 tonnes/yr", ""),
        ("CO2 from ocean industries", "800,000 tonnes CO2/yr", ""),
    ]
    for r, (lb, val, note) in enumerate(answers, 3):
        if lb and not lb.startswith("Part") and lb != "Summary":
            ws_a.cell(row=r, column=1, value=lb).font = label_font
            ws_a.cell(row=r, column=1).border = bdr
            cell = ws_a.cell(row=r, column=2, value=val)
            apply_answer(cell)
            ws_a.cell(row=r, column=3, value=note).font = Font(
                name="Arial", size=9, italic=True, color="717171")
        elif lb:
            ws_a.cell(row=r, column=1, value=lb).font = Font(
                name="Arial", size=12, bold=True, color=TEAL)

    set_col_widths(ws_a, [32, 28, 30])

    path = Path(__file__).parent / "waste_exercise.xlsx"
    wb.save(path)
    print(f"  Created: {path}")


if __name__ == "__main__":
    print("Generating waste account files...")
    generate_waste_template()
    generate_waste_exercise()
    print("Done!")
