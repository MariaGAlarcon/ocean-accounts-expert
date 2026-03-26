#!/usr/bin/env python3
"""Generate improved waste exercise workbook with step-by-step structure covering 3 sub-accounts."""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# GOAP colour palette
# ---------------------------------------------------------------------------
GREEN = "3B9C7B"
TEAL = "0A5455"
GRAY = "404040"
TEXT = "30302F"
WHITE = "FFFFFF"
YELLOW = "FFF2CC"

fill_teal = PatternFill(start_color=TEAL, end_color=TEAL, fill_type="solid")
fill_green = PatternFill(start_color=GREEN, end_color=GREEN, fill_type="solid")
fill_gray = PatternFill(start_color=GRAY, end_color=GRAY, fill_type="solid")
fill_yellow = PatternFill(start_color=YELLOW, end_color=YELLOW, fill_type="solid")
fill_white = PatternFill(start_color=WHITE, end_color=WHITE, fill_type="solid")
fill_light_green = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")

font_title = Font(name="Calibri", size=16, bold=True, color=TEAL)
font_heading = Font(name="Calibri", size=13, bold=True, color=TEAL)
font_subheading = Font(name="Calibri", size=11, bold=True, color=GRAY)
font_body = Font(name="Calibri", size=11, color=TEXT)
font_bold = Font(name="Calibri", size=11, bold=True, color=TEXT)
font_white_bold = Font(name="Calibri", size=11, bold=True, color=WHITE)
font_white_title = Font(name="Calibri", size=14, bold=True, color=WHITE)
font_small = Font(name="Calibri", size=10, color=GRAY)
font_italic = Font(name="Calibri", size=11, italic=True, color=GRAY)

align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
align_wrap = Alignment(wrap_text=True, vertical="top")

thin_border = Border(
    left=Side(style="thin", color=GRAY),
    right=Side(style="thin", color=GRAY),
    top=Side(style="thin", color=GRAY),
    bottom=Side(style="thin", color=GRAY),
)


def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def header_row(ws, row, headers, fill=fill_teal, font=font_white_bold):
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = font
        c.fill = fill
        c.alignment = align_center
        c.border = thin_border


def data_cell(ws, row, col, value, yellow=False, bold=False, number_format=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font = font_bold if bold else font_body
    c.alignment = align_center
    c.border = thin_border
    if yellow:
        c.fill = fill_yellow
    if number_format:
        c.number_format = number_format
    return c


def title_banner(ws, row, text, col_span=6):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_span)
    c = ws.cell(row=row, column=1, value=text)
    c.font = font_white_title
    c.fill = fill_teal
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 36


# ===========================================================================
wb = Workbook()

# ---- Sheet 1: README -----------------------------------------------------
ws = wb.active
ws.title = "README"
ws.sheet_properties.tabColor = TEAL
set_col_widths(ws, [3, 65, 25])

title_banner(ws, 1, "  Waste & Emissions Exercise: Building Ocean Pollution Accounts", 3)

rows = [
    (3, font_heading, "Overview"),
    (4, font_body, "This workbook walks you through building three pollution sub-accounts for the ocean economy, following the SEEA Central Framework approach."),
    (6, font_heading, "The 4 Steps"),
    (7, font_bold, "Step 1 - Water Emissions: Calculate nutrient and organic loads entering coastal waters."),
    (8, font_bold, "Step 2 - Solid Waste: Estimate marine debris using the Jambeck method."),
    (9, font_bold, "Step 3 - Air Emissions: Compute atmospheric emissions from maritime sectors."),
    (10, font_bold, "Step 4 - Build Account Tables: Compile SEEA-CF format tables for all three sub-accounts."),
    (12, font_heading, "Instructions"),
    (13, font_body, "- Yellow cells are for your answers."),
    (14, font_body, "- Work through Steps 1-3 first, then compile results in Step 4."),
    (15, font_body, "- Check your work against the 'Answers' sheet when finished."),
    (16, font_body, "- The 'Glossary' sheet defines technical terms."),
    (18, font_heading, "Learning Objectives"),
    (19, font_body, "1. Calculate pollutant loads from concentration and flow data."),
    (20, font_body, "2. Apply the Jambeck method for marine debris estimation."),
    (21, font_body, "3. Use emission factors to estimate air pollutants from fuel consumption."),
    (22, font_body, "4. Compile SEEA-CF format emission account tables."),
]
for r, fnt, val in rows:
    c = ws.cell(row=r, column=2, value=val)
    c.font = fnt
    c.alignment = align_wrap

# ---- Sheet 2: Step 1 - Water Emissions -----------------------------------
ws = wb.create_sheet("Step 1 - Water Emissions")
ws.sheet_properties.tabColor = GREEN
set_col_widths(ws, [3, 42, 22, 22, 22, 22])

title_banner(ws, 1, "  Step 1: Water Emissions to Coastal Waters", 6)

ws.cell(row=3, column=2, value="Given Data").font = font_heading
given = [
    "Total coastal wastewater volume: 200 million m\u00b3/year",
    "Treatment rate: 60% (meaning 40% is discharged untreated)",
    "Pollutant concentrations in untreated wastewater:",
    "   Nitrogen (N): 15 mg/L  =  15 g/m\u00b3",
    "   Phosphorus (P): 3 mg/L  =  3 g/m\u00b3",
    "   BOD: 45 mg/L  =  45 g/m\u00b3",
]
for i, g in enumerate(given):
    ws.cell(row=4 + i, column=2, value=g).font = font_body

ws.cell(row=11, column=2, value="Task A: Calculate untreated volume and pollutant loads").font = font_heading

header_row(ws, 13, ["", "Calculation Step", "Formula", "Your Calculation", "Result", "Unit"])
calc_rows = [
    ("Untreated volume", "200M x 0.40", "", "", "million m\u00b3/yr"),
    ("Nitrogen load", "volume x 15 g/m\u00b3 / 1,000,000", "", "", "tonnes/yr"),
    ("Phosphorus load", "volume x 3 g/m\u00b3 / 1,000,000", "", "", "tonnes/yr"),
    ("BOD load", "volume x 45 g/m\u00b3 / 1,000,000", "", "", "tonnes/yr"),
]
for i, (step, formula, calc, res, unit) in enumerate(calc_rows):
    r = 14 + i
    data_cell(ws, r, 2, step).alignment = align_left
    data_cell(ws, r, 3, formula).alignment = align_left
    data_cell(ws, r, 4, None, yellow=True)
    data_cell(ws, r, 5, None, yellow=True)
    data_cell(ws, r, 6, unit)

ws.cell(row=19, column=2, value="Note: 1 mg/L = 1 g/m\u00b3.  To convert g to tonnes: divide by 1,000,000 (10\u2076).").font = font_small
ws.merge_cells("B19:F19")

ws.cell(row=21, column=2, value="Task B: Build the supply table by source sector").font = font_heading
ws.cell(row=22, column=2, value="Assume the following source split for untreated wastewater:").font = font_italic
ws.cell(row=23, column=2, value="   Households 50%  |  Agriculture 30%  |  Industry 15%  |  Other 5%").font = font_body

header_row(ws, 25, ["", "Source Sector", "Share (%)", "N (tonnes)", "P (tonnes)", "BOD (tonnes)"])
sectors = [("Households", "50%"), ("Agriculture", "30%"), ("Industry", "15%"), ("Other", "5%")]
for i, (sec, share) in enumerate(sectors):
    r = 26 + i
    data_cell(ws, r, 2, sec).alignment = align_left
    data_cell(ws, r, 3, share)
    data_cell(ws, r, 4, None, yellow=True)
    data_cell(ws, r, 5, None, yellow=True)
    data_cell(ws, r, 6, None, yellow=True)
r_tot = 30
data_cell(ws, r_tot, 2, "TOTAL", bold=True).alignment = align_left
data_cell(ws, r_tot, 3, "100%", bold=True)
data_cell(ws, r_tot, 4, None, yellow=True)
data_cell(ws, r_tot, 5, None, yellow=True)
data_cell(ws, r_tot, 6, None, yellow=True)

# ---- Sheet 3: Step 2 - Solid Waste ---------------------------------------
ws = wb.create_sheet("Step 2 - Solid Waste")
ws.sheet_properties.tabColor = GREEN
set_col_widths(ws, [3, 42, 22, 22, 22, 22])

title_banner(ws, 1, "  Step 2: Marine Solid Waste (Jambeck Method)", 6)

ws.cell(row=3, column=2, value="Given Data").font = font_heading
given_sw = [
    "National municipal solid waste generated: 2,000,000 tonnes/year",
    "Mismanaged waste rate: 25% (not properly collected or disposed)",
    "Coastal population share: 40% of national population",
    "Marine leakage rate: 15% of mismanaged coastal waste enters the ocean",
]
for i, g in enumerate(given_sw):
    ws.cell(row=4 + i, column=2, value=g).font = font_body

ws.cell(row=9, column=2, value="Task A: Calculate marine debris using the Jambeck cascade").font = font_heading

header_row(ws, 11, ["", "Calculation Step", "Formula", "Your Calculation", "Result", "Unit"])
jambeck_rows = [
    ("1. National waste", "Given", "", "2,000,000", "tonnes/yr"),
    ("2. Mismanaged waste", "2,000,000 x 0.25", "", "", "tonnes/yr"),
    ("3. Coastal mismanaged waste", "Mismanaged x 0.40", "", "", "tonnes/yr"),
    ("4. Marine debris (leakage)", "Coastal mismanaged x 0.15", "", "", "tonnes/yr"),
]
for i, (step, formula, calc, res, unit) in enumerate(jambeck_rows):
    r = 12 + i
    data_cell(ws, r, 2, step).alignment = align_left
    data_cell(ws, r, 3, formula).alignment = align_left
    data_cell(ws, r, 4, None, yellow=True)
    if res:
        data_cell(ws, r, 5, res)
    else:
        data_cell(ws, r, 5, None, yellow=True)
    data_cell(ws, r, 6, unit)

ws.cell(row=17, column=2, value="Task B: Break down marine debris by waste type").font = font_heading
ws.cell(row=18, column=2, value="Composition: Plastics 60%  |  Fishing gear 15%  |  Other (glass, metal, paper, etc.) 25%").font = font_body
ws.merge_cells("B18:F18")

header_row(ws, 20, ["", "Waste Type", "Share (%)", "Marine Debris (tonnes)", "", ""])
waste_types = [("Plastics", "60%"), ("Fishing gear", "15%"), ("Other (glass, metal, paper, etc.)", "25%")]
for i, (wt, share) in enumerate(waste_types):
    r = 21 + i
    data_cell(ws, r, 2, wt).alignment = align_left
    data_cell(ws, r, 3, share)
    data_cell(ws, r, 4, None, yellow=True)
r_tot = 24
data_cell(ws, r_tot, 2, "TOTAL", bold=True).alignment = align_left
data_cell(ws, r_tot, 3, "100%", bold=True)
data_cell(ws, r_tot, 4, None, yellow=True)

ws.cell(row=26, column=2, value="Task C: Identify sources (land-based vs sea-based)").font = font_heading
ws.cell(row=27, column=2, value="Assume 80% of marine debris is land-based and 20% is sea-based (fishing, shipping).").font = font_body
ws.merge_cells("B27:F27")

header_row(ws, 29, ["", "Source", "Share (%)", "Debris (tonnes)", "", ""])
sources = [("Land-based (runoff, rivers, direct dumping)", "80%"), ("Sea-based (fishing, shipping)", "20%")]
for i, (src, share) in enumerate(sources):
    r = 30 + i
    data_cell(ws, r, 2, src).alignment = align_left
    data_cell(ws, r, 3, share)
    data_cell(ws, r, 4, None, yellow=True)

# ---- Sheet 4: Step 3 - Air Emissions -------------------------------------
ws = wb.create_sheet("Step 3 - Air Emissions")
ws.sheet_properties.tabColor = GREEN
set_col_widths(ws, [3, 34, 22, 20, 20, 20])

title_banner(ws, 1, "  Step 3: Air Emissions from Maritime Sectors", 6)

ws.cell(row=3, column=2, value="Given Data").font = font_heading
ws.cell(row=4, column=2, value="Annual fuel consumption by sector:").font = font_bold
fuel_data = [
    ("Fishing fleet", "50,000 tonnes/yr"),
    ("Domestic shipping", "200,000 tonnes/yr"),
    ("Offshore energy", "30,000 tonnes/yr"),
]
for i, (sec, fuel) in enumerate(fuel_data):
    ws.cell(row=5 + i, column=2, value=f"   {sec}: {fuel}").font = font_body

ws.cell(row=9, column=2, value="Emission factors (tonnes pollutant per tonne fuel burned):").font = font_bold
efs = [("CO\u2082", "3.200"), ("SO\u2082", "0.054"), ("NOx", "0.084")]
header_row(ws, 10, ["", "Pollutant", "Emission Factor", "", "", ""])
for i, (poll, ef) in enumerate(efs):
    r = 11 + i
    data_cell(ws, r, 2, poll).alignment = align_left
    data_cell(ws, r, 3, ef)

ws.cell(row=15, column=2, value="Task: Calculate emissions per sector and pollutant").font = font_heading
ws.cell(row=16, column=2, value="Formula: Emissions = Fuel consumption x Emission factor").font = font_italic

header_row(ws, 18, ["", "Sector", "Fuel (t/yr)", "CO\u2082 (tonnes)", "SO\u2082 (tonnes)", "NOx (tonnes)"])
sectors_air = [
    ("Fishing fleet", "50,000"),
    ("Domestic shipping", "200,000"),
    ("Offshore energy", "30,000"),
]
for i, (sec, fuel) in enumerate(sectors_air):
    r = 19 + i
    data_cell(ws, r, 2, sec).alignment = align_left
    data_cell(ws, r, 3, fuel)
    data_cell(ws, r, 4, None, yellow=True)
    data_cell(ws, r, 5, None, yellow=True)
    data_cell(ws, r, 6, None, yellow=True)

r_tot = 22
data_cell(ws, r_tot, 2, "TOTAL", bold=True).alignment = align_left
data_cell(ws, r_tot, 3, "280,000", bold=True)
data_cell(ws, r_tot, 4, None, yellow=True)
data_cell(ws, r_tot, 5, None, yellow=True)
data_cell(ws, r_tot, 6, None, yellow=True)

ws.cell(row=24, column=2, value="Reflection: Which sector is the largest emitter? What policy levers could reduce maritime air emissions?").font = font_italic
ws.merge_cells("B24:F24")

# ---- Sheet 5: Step 4 - Build Account Tables ------------------------------
ws = wb.create_sheet("Step 4 - Build Account Tables")
ws.sheet_properties.tabColor = GREEN
set_col_widths(ws, [3, 34, 20, 18, 18, 18, 18])

title_banner(ws, 1, "  Step 4: Compile SEEA-CF Format Account Tables", 7)

ws.cell(row=3, column=2, value="Using your results from Steps 1-3, fill in the formal account tables below.").font = font_body
ws.merge_cells("B3:G3")

# Table A: Water emissions
ws.cell(row=5, column=2, value="Table A: Water Emissions Account (tonnes/year)").font = font_heading
header_row(ws, 7, ["", "Source Sector", "Nitrogen", "Phosphorus", "BOD", "Total", ""], fill=fill_teal)
w_sectors = ["Households", "Agriculture", "Industry", "Other", "TOTAL"]
for i, sec in enumerate(w_sectors):
    r = 8 + i
    bld = (sec == "TOTAL")
    data_cell(ws, r, 2, sec, bold=bld).alignment = align_left
    for col in range(3, 7):
        data_cell(ws, r, col, None, yellow=True)

# Table B: Solid waste
ws.cell(row=14, column=2, value="Table B: Solid Waste to Ocean Account (tonnes/year)").font = font_heading
header_row(ws, 16, ["", "Source", "Plastics", "Fishing Gear", "Other", "Total", ""], fill=fill_teal)
sw_sources = ["Land-based", "Sea-based", "TOTAL"]
for i, src in enumerate(sw_sources):
    r = 17 + i
    bld = (src == "TOTAL")
    data_cell(ws, r, 2, src, bold=bld).alignment = align_left
    for col in range(3, 7):
        data_cell(ws, r, col, None, yellow=True)

# Table C: Air emissions
ws.cell(row=21, column=2, value="Table C: Air Emissions Account (tonnes/year)").font = font_heading
header_row(ws, 23, ["", "Sector", "CO\u2082", "SO\u2082", "NOx", "Total", ""], fill=fill_teal)
air_sectors = ["Fishing fleet", "Domestic shipping", "Offshore energy", "TOTAL"]
for i, sec in enumerate(air_sectors):
    r = 24 + i
    bld = (sec == "TOTAL")
    data_cell(ws, r, 2, sec, bold=bld).alignment = align_left
    for col in range(3, 7):
        data_cell(ws, r, col, None, yellow=True)

# Summary
ws.cell(row=29, column=2, value="Summary: Total Pollution Load to Ocean").font = font_heading
header_row(ws, 31, ["", "Pollution Type", "Total (tonnes/yr)", "Primary Source", "", "", ""], fill=fill_teal)
summary_rows = ["Water emissions (N + P + BOD)", "Solid waste (marine debris)", "Air emissions (CO\u2082 + SO\u2082 + NOx)", "GRAND TOTAL"]
for i, sr in enumerate(summary_rows):
    r = 32 + i
    bld = (sr == "GRAND TOTAL")
    data_cell(ws, r, 2, sr, bold=bld).alignment = align_left
    data_cell(ws, r, 3, None, yellow=True)
    data_cell(ws, r, 4, None, yellow=True).alignment = align_left

# ---- Sheet 6: Answers ----------------------------------------------------
ws = wb.create_sheet("Answers")
ws.sheet_properties.tabColor = GRAY
set_col_widths(ws, [3, 42, 22, 22, 22, 22])

title_banner(ws, 1, "  Answers: Full Calculations", 6)

# ---- Water emissions answers ----
ws.cell(row=3, column=2, value="Step 1 Answers - Water Emissions").font = font_heading

header_row(ws, 5, ["", "Step", "Calculation", "Result", "Unit", ""], fill=fill_gray)
water_ans = [
    ("Untreated volume", "200M x 0.40", "80,000,000", "m\u00b3/yr"),
    ("  = in million m\u00b3", "", "80", "million m\u00b3/yr"),
    ("Nitrogen load", "80,000,000 x 15 / 1,000,000", "1,200", "tonnes/yr"),
    ("Phosphorus load", "80,000,000 x 3 / 1,000,000", "240", "tonnes/yr"),
    ("BOD load", "80,000,000 x 45 / 1,000,000", "3,600", "tonnes/yr"),
]
for i, (step, calc, res, unit) in enumerate(water_ans):
    r = 6 + i
    data_cell(ws, r, 2, step).alignment = align_left
    data_cell(ws, r, 3, calc).alignment = align_left
    data_cell(ws, r, 4, res)
    data_cell(ws, r, 5, unit)

ws.cell(row=12, column=2, value="Supply table by sector:").font = font_bold
header_row(ws, 13, ["", "Sector", "Share", "N (t)", "P (t)", "BOD (t)"], fill=fill_gray)
sector_ans = [
    ("Households", "50%", "600", "120", "1,800"),
    ("Agriculture", "30%", "360", "72", "1,080"),
    ("Industry", "15%", "180", "36", "540"),
    ("Other", "5%", "60", "12", "180"),
    ("TOTAL", "100%", "1,200", "240", "3,600"),
]
for i, (sec, share, n, p, bod) in enumerate(sector_ans):
    r = 14 + i
    bld = (sec == "TOTAL")
    data_cell(ws, r, 2, sec, bold=bld).alignment = align_left
    data_cell(ws, r, 3, share, bold=bld)
    data_cell(ws, r, 4, n, bold=bld)
    data_cell(ws, r, 5, p, bold=bld)
    data_cell(ws, r, 6, bod, bold=bld)

# ---- Solid waste answers ----
ws.cell(row=20, column=2, value="Step 2 Answers - Solid Waste").font = font_heading

header_row(ws, 22, ["", "Step", "Calculation", "Result", "Unit", ""], fill=fill_gray)
sw_ans = [
    ("National waste", "Given", "2,000,000", "tonnes/yr"),
    ("Mismanaged waste", "2,000,000 x 0.25", "500,000", "tonnes/yr"),
    ("Coastal mismanaged", "500,000 x 0.40", "200,000", "tonnes/yr"),
    ("Marine debris", "200,000 x 0.15", "30,000", "tonnes/yr"),
]
for i, (step, calc, res, unit) in enumerate(sw_ans):
    r = 23 + i
    data_cell(ws, r, 2, step).alignment = align_left
    data_cell(ws, r, 3, calc).alignment = align_left
    data_cell(ws, r, 4, res)
    data_cell(ws, r, 5, unit)

ws.cell(row=28, column=2, value="By waste type:").font = font_bold
header_row(ws, 29, ["", "Type", "Share", "Tonnes", "", ""], fill=fill_gray)
type_ans = [("Plastics", "60%", "18,000"), ("Fishing gear", "15%", "4,500"), ("Other", "25%", "7,500"), ("TOTAL", "100%", "30,000")]
for i, (t, s, v) in enumerate(type_ans):
    r = 30 + i
    bld = (t == "TOTAL")
    data_cell(ws, r, 2, t, bold=bld).alignment = align_left
    data_cell(ws, r, 3, s, bold=bld)
    data_cell(ws, r, 4, v, bold=bld)

ws.cell(row=35, column=2, value="By source:").font = font_bold
header_row(ws, 36, ["", "Source", "Share", "Tonnes", "", ""], fill=fill_gray)
src_ans = [("Land-based", "80%", "24,000"), ("Sea-based", "20%", "6,000"), ("TOTAL", "100%", "30,000")]
for i, (s, sh, v) in enumerate(src_ans):
    r = 37 + i
    bld = (s == "TOTAL")
    data_cell(ws, r, 2, s, bold=bld).alignment = align_left
    data_cell(ws, r, 3, sh, bold=bld)
    data_cell(ws, r, 4, v, bold=bld)

# ---- Air emissions answers ----
ws.cell(row=41, column=2, value="Step 3 Answers - Air Emissions").font = font_heading

header_row(ws, 43, ["", "Sector", "Fuel (t)", "CO\u2082 (t)", "SO\u2082 (t)", "NOx (t)"], fill=fill_gray)
# CO2=3.2, SO2=0.054, NOx=0.084
air_ans = [
    ("Fishing fleet", "50,000", "160,000", "2,700", "4,200"),
    ("Domestic shipping", "200,000", "640,000", "10,800", "16,800"),
    ("Offshore energy", "30,000", "96,000", "1,620", "2,520"),
    ("TOTAL", "280,000", "896,000", "15,120", "23,520"),
]
for i, (sec, fuel, co2, so2, nox) in enumerate(air_ans):
    r = 44 + i
    bld = (sec == "TOTAL")
    data_cell(ws, r, 2, sec, bold=bld).alignment = align_left
    data_cell(ws, r, 3, fuel, bold=bld)
    data_cell(ws, r, 4, co2, bold=bld)
    data_cell(ws, r, 5, so2, bold=bld)
    data_cell(ws, r, 6, nox, bold=bld)

# ---- Completed account tables ----
ws.cell(row=49, column=2, value="Completed Account Tables").font = font_heading

ws.cell(row=50, column=2, value="Table A: Water Emissions (tonnes/yr)").font = font_bold
header_row(ws, 51, ["", "Sector", "N", "P", "BOD", "Total"], fill=fill_gray)
table_a = [
    ("Households", "600", "120", "1,800", "2,520"),
    ("Agriculture", "360", "72", "1,080", "1,512"),
    ("Industry", "180", "36", "540", "756"),
    ("Other", "60", "12", "180", "252"),
    ("TOTAL", "1,200", "240", "3,600", "5,040"),
]
for i, row_data in enumerate(table_a):
    r = 52 + i
    bld = (row_data[0] == "TOTAL")
    for j, val in enumerate(row_data):
        data_cell(ws, r, 2 + j, val, bold=bld).alignment = align_left if j == 0 else align_center

ws.cell(row=58, column=2, value="Table B: Solid Waste to Ocean (tonnes/yr)").font = font_bold
header_row(ws, 59, ["", "Source", "Plastics", "Fishing Gear", "Other", "Total"], fill=fill_gray)
# Land 80%, Sea 20% applied to type breakdown: P=18000, FG=4500, O=7500
# For simplicity, land-based is primarily plastics/other, sea-based is primarily fishing gear
# But following simple proportional: land 80% of each, sea 20% of each
table_b = [
    ("Land-based", "14,400", "3,600", "6,000", "24,000"),
    ("Sea-based", "3,600", "900", "1,500", "6,000"),
    ("TOTAL", "18,000", "4,500", "7,500", "30,000"),
]
for i, row_data in enumerate(table_b):
    r = 60 + i
    bld = (row_data[0] == "TOTAL")
    for j, val in enumerate(row_data):
        data_cell(ws, r, 2 + j, val, bold=bld).alignment = align_left if j == 0 else align_center

ws.cell(row=64, column=2, value="Table C: Air Emissions (tonnes/yr)").font = font_bold
header_row(ws, 65, ["", "Sector", "CO\u2082", "SO\u2082", "NOx", "Total"], fill=fill_gray)
table_c = [
    ("Fishing fleet", "160,000", "2,700", "4,200", "166,900"),
    ("Domestic shipping", "640,000", "10,800", "16,800", "667,600"),
    ("Offshore energy", "96,000", "1,620", "2,520", "100,140"),
    ("TOTAL", "896,000", "15,120", "23,520", "934,640"),
]
for i, row_data in enumerate(table_c):
    r = 66 + i
    bld = (row_data[0] == "TOTAL")
    for j, val in enumerate(row_data):
        data_cell(ws, r, 2 + j, val, bold=bld).alignment = align_left if j == 0 else align_center

ws.cell(row=71, column=2, value="Summary:").font = font_bold
ws.cell(row=72, column=2, value="Total water emissions: 5,040 tonnes/yr (N + P + BOD)").font = font_body
ws.cell(row=73, column=2, value="Total marine debris: 30,000 tonnes/yr").font = font_body
ws.cell(row=74, column=2, value="Total air emissions: 934,640 tonnes/yr (CO\u2082 + SO\u2082 + NOx)").font = font_body

# ---- Sheet 7: Glossary ---------------------------------------------------
ws = wb.create_sheet("Glossary")
ws.sheet_properties.tabColor = TEAL
set_col_widths(ws, [3, 28, 65])

title_banner(ws, 1, "  Glossary of Key Terms", 3)

glossary = [
    ("BOD", "Biochemical Oxygen Demand - the amount of dissolved oxygen needed by aerobic organisms to break down organic matter. A key indicator of organic water pollution."),
    ("Nitrogen (N)", "A nutrient that, in excess, causes eutrophication (algal blooms) in coastal waters. Sources include sewage, agricultural runoff, and industrial discharge."),
    ("Phosphorus (P)", "A nutrient that, like nitrogen, drives eutrophication. Often the limiting nutrient in freshwater systems. Key sources: detergents, fertilizers, sewage."),
    ("Heavy metals", "Toxic metallic elements (mercury, lead, cadmium, etc.) that accumulate in marine food chains. Sources: industrial discharge, mining, anti-fouling paints."),
    ("Jambeck method", "An estimation approach (Jambeck et al., 2015) that calculates marine debris as: National waste x Mismanaged rate x Coastal population share x Marine leakage rate."),
    ("MARPOL", "International Convention for the Prevention of Pollution from Ships. Sets rules for discharge of pollutants from vessels (oil, chemicals, sewage, garbage, air emissions)."),
    ("Emission factor", "A coefficient relating the quantity of a pollutant released to a unit of activity (e.g., tonnes CO\u2082 per tonne of fuel burned)."),
    ("SEEA-CF", "System of Environmental-Economic Accounting - Central Framework. The international statistical standard for environmental-economic accounts."),
    ("Eutrophication", "Excessive nutrient enrichment of water bodies leading to algal blooms, oxygen depletion, and ecosystem degradation."),
    ("Mismanaged waste", "Waste that is not properly collected, treated, or disposed of - including open dumps, uncontrolled landfills, and littering."),
    ("Marine leakage rate", "The proportion of mismanaged coastal waste that enters the ocean through rivers, wind, tides, or direct dumping."),
    ("Supply table", "In SEEA, a table showing the sources (sectors/industries) that generate a particular type of emission or waste flow."),
    ("CO\u2082", "Carbon dioxide - the primary greenhouse gas from fossil fuel combustion. Maritime CO\u2082 comes from ship engines and offshore operations."),
    ("SO\u2082", "Sulphur dioxide - an air pollutant causing acid rain and respiratory harm. Ships burning high-sulphur fuel are a major source."),
    ("NOx", "Nitrogen oxides - air pollutants contributing to smog and acid rain. Produced by high-temperature combustion in ship engines."),
]
header_row(ws, 3, ["", "Term", "Definition"], fill=fill_teal)
for i, (term, defn) in enumerate(glossary):
    r = 4 + i
    data_cell(ws, r, 2, term, bold=True).alignment = align_left
    data_cell(ws, r, 3, defn).alignment = align_wrap
    ws.row_dimensions[r].height = 42

# ---------------------------------------------------------------------------
# Save
# ---------------------------------------------------------------------------
out_dir = os.path.dirname(os.path.abspath(__file__))
out_path = os.path.join(out_dir, "waste_exercise.xlsx")
wb.save(out_path)
print(f"Saved: {out_path}")
