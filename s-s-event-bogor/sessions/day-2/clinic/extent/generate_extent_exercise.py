#!/usr/bin/env python3
"""
Generate a GOAP-styled Excel workbook for the Extent Account exercise.
"""

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, NamedStyle
)
from openpyxl.utils import get_column_letter
import os

OUTPUT = os.path.join(os.path.dirname(os.path.abspath(__file__)), "extent_exercise.xlsx")

# Colours
GREEN = "3B9C7B"
TEAL = "0A5455"
YELLOW_INPUT = "FFF2CC"
ANSWER_GREEN = "C6EFCE"
BORDER_CLR = "404040"
TEXT_DARK = "30302F"
WHITE = "FFFFFF"

# Reusable styles
thin_border = Border(
    left=Side(style="thin", color=BORDER_CLR),
    right=Side(style="thin", color=BORDER_CLR),
    top=Side(style="thin", color=BORDER_CLR),
    bottom=Side(style="thin", color=BORDER_CLR),
)
font_normal = Font(name="Arial", size=10, color=TEXT_DARK)
font_bold = Font(name="Arial", size=10, color=TEXT_DARK, bold=True)
font_header = Font(name="Arial", size=10, color=WHITE, bold=True)
font_title = Font(name="Arial", size=14, color=TEAL, bold=True)
font_subtitle = Font(name="Arial", size=11, color=TEAL, bold=True)

fill_green = PatternFill(start_color=GREEN, end_color=GREEN, fill_type="solid")
fill_teal = PatternFill(start_color=TEAL, end_color=TEAL, fill_type="solid")
fill_yellow = PatternFill(start_color=YELLOW_INPUT, end_color=YELLOW_INPUT, fill_type="solid")
fill_answer = PatternFill(start_color=ANSWER_GREEN, end_color=ANSWER_GREEN, fill_type="solid")

align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)


def header_style(cell, text):
    cell.value = text
    cell.font = font_header
    cell.fill = fill_green
    cell.border = thin_border
    cell.alignment = align_center


def row_header(cell, text):
    cell.value = text
    cell.font = Font(name="Arial", size=10, color=WHITE, bold=True)
    cell.fill = fill_teal
    cell.border = thin_border
    cell.alignment = align_left


def data(cell, value, align="center"):
    cell.value = value
    cell.font = font_normal
    cell.border = thin_border
    cell.alignment = align_center if align == "center" else align_left


def yellow_input(cell, value=None):
    cell.value = value
    cell.font = font_normal
    cell.fill = fill_yellow
    cell.border = thin_border
    cell.alignment = align_center


def answer_cell(cell, value):
    cell.value = value
    cell.font = font_bold
    cell.fill = fill_answer
    cell.border = thin_border
    cell.alignment = align_center


def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ── Sheet 1: Exercise ────────────────────────────────────────────────────────

def build_exercise(wb):
    ws = wb.active
    ws.title = "Exercise"
    set_col_widths(ws, [22, 18, 18, 18, 18, 18, 18])

    # Title
    ws.merge_cells("A1:G1")
    c = ws["A1"]
    c.value = "Extent Account Exercise"
    c.font = font_title
    c.alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:G2")
    c = ws["A2"]
    c.value = "Hypothetical coastal ecosystem accounting area  |  2018 - 2023"
    c.font = Font(name="Arial", size=10, color=TEXT_DARK, italic=True)
    c.alignment = Alignment(horizontal="center")

    # ── Given data: Opening extent ──
    row = 4
    ws.merge_cells(f"A{row}:C{row}")
    c = ws.cell(row=row, column=1)
    c.value = "Given Data: Opening Extent (2018, ha)"
    c.font = font_subtitle

    row = 5
    for ci, h in enumerate(["Ecosystem type", "Opening extent (ha)"], 1):
        header_style(ws.cell(row=row, column=ci), h)
    given_opening = [("Coral reef", 1240), ("Seagrass beds", 860),
                     ("Mangroves", 315), ("Total", 2415)]
    for i, (eco, val) in enumerate(given_opening):
        r = row + 1 + i
        if eco == "Total":
            header_style(ws.cell(row=r, column=1), eco)
            header_style(ws.cell(row=r, column=2), val)
        else:
            row_header(ws.cell(row=r, column=1), eco)
            data(ws.cell(row=r, column=2), val)

    # ── Given data: Closing extent ──
    row = 11
    ws.merge_cells(f"A{row}:C{row}")
    c = ws.cell(row=row, column=1)
    c.value = "Given Data: Closing Extent (2023, ha)"
    c.font = font_subtitle

    row = 12
    for ci, h in enumerate(["Ecosystem type", "Closing extent (ha)"], 1):
        header_style(ws.cell(row=row, column=ci), h)
    given_closing = [("Coral reef", 1180), ("Seagrass beds", 825),
                     ("Mangroves", 340), ("Total", 2345)]
    for i, (eco, val) in enumerate(given_closing):
        r = row + 1 + i
        if eco == "Total":
            header_style(ws.cell(row=r, column=1), eco)
            header_style(ws.cell(row=r, column=2), val)
        else:
            row_header(ws.cell(row=r, column=1), eco)
            data(ws.cell(row=r, column=2), val)

    # ── Given data: Change matrix ──
    row = 18
    ws.merge_cells(f"A{row}:F{row}")
    c = ws.cell(row=row, column=1)
    c.value = "Given Data: Ecosystem Extent Change Matrix (ha, 2018-2023)"
    c.font = font_subtitle

    row = 19
    cm_headers = ["From \\ To", "Coral reef", "Seagrass", "Mangrove", "Other",
                  "Total reductions"]
    for ci, h in enumerate(cm_headers, 1):
        header_style(ws.cell(row=row, column=ci), h)

    cm_data = [
        ["Coral reef",     "---", 15,  5,   80,  100],
        ["Seagrass beds",  10,    "---", 10, 50,  70],
        ["Mangroves",      0,     5,   "---", 10, 15],
        ["Other marine",   30,    15,  25,  "---", 70],
        ["Total additions", 40,   35,  40,  140, "---"],
    ]
    for i, r_data in enumerate(cm_data):
        r = row + 1 + i
        for ci, val in enumerate(r_data):
            cell = ws.cell(row=r, column=ci + 1)
            if ci == 0:
                if "Total" in str(val):
                    header_style(cell, val)
                else:
                    row_header(cell, val)
            elif val == "---":
                cell.value = "---"
                cell.font = font_normal
                cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9",
                                        fill_type="solid")
                cell.border = thin_border
                cell.alignment = align_center
            elif "Total" in str(cm_data[i][0]):
                header_style(cell, val)
            else:
                data(cell, val)

    # ── Part 1: Fill in the extent account ──
    row = 27
    ws.merge_cells(f"A{row}:G{row}")
    c = ws.cell(row=row, column=1)
    c.value = "Part 1: Complete the Extent Account table below (fill yellow cells)"
    c.font = font_subtitle

    row = 28
    ea_headers = ["Ecosystem type", "Opening\nextent (ha)", "Additions\n(ha)",
                  "Reductions\n(ha)", "Net change\n(ha)", "Closing\nextent (ha)"]
    for ci, h in enumerate(ea_headers, 1):
        header_style(ws.cell(row=row, column=ci), h)

    ecosystems = ["Coral reef", "Seagrass beds", "Mangroves"]
    for i, eco in enumerate(ecosystems):
        r = row + 1 + i
        row_header(ws.cell(row=r, column=1), eco)
        # Opening extent is given
        data(ws.cell(row=r, column=2), [1240, 860, 315][i])
        # Additions, Reductions, Net change, Closing = yellow input
        for col in range(3, 7):
            yellow_input(ws.cell(row=r, column=col))

    # Total row
    r_total = row + 4
    header_style(ws.cell(row=r_total, column=1), "Total")
    for col in range(2, 7):
        yellow_input(ws.cell(row=r_total, column=col))

    # ── Part 2: Percentage change ──
    row = 34
    ws.merge_cells(f"A{row}:D{row}")
    c = ws.cell(row=row, column=1)
    c.value = "Part 2: Calculate the percentage change for each ecosystem"
    c.font = font_subtitle

    row = 35
    pct_headers = ["Ecosystem type", "Opening (ha)", "Closing (ha)", "% change"]
    for ci, h in enumerate(pct_headers, 1):
        header_style(ws.cell(row=row, column=ci), h)
    for i, eco in enumerate(ecosystems):
        r = row + 1 + i
        row_header(ws.cell(row=r, column=1), eco)
        data(ws.cell(row=r, column=2), [1240, 860, 315][i])
        data(ws.cell(row=r, column=3), [1180, 825, 340][i])
        yellow_input(ws.cell(row=r, column=4))

    return ws


# ── Sheet 2: Answers ─────────────────────────────────────────────────────────

def build_answers(wb):
    ws = wb.create_sheet("Answers")
    set_col_widths(ws, [22, 18, 18, 18, 18, 18, 18])

    ws.merge_cells("A1:G1")
    c = ws["A1"]
    c.value = "Extent Account Exercise -- ANSWERS"
    c.font = font_title
    c.alignment = Alignment(horizontal="center")

    # Extent account answer
    row = 3
    ws.merge_cells(f"A{row}:G{row}")
    c = ws.cell(row=row, column=1)
    c.value = "Part 1: Completed Extent Account"
    c.font = font_subtitle

    row = 4
    ea_headers = ["Ecosystem type", "Opening\nextent (ha)", "Additions\n(ha)",
                  "Reductions\n(ha)", "Net change\n(ha)", "Closing\nextent (ha)"]
    for ci, h in enumerate(ea_headers, 1):
        header_style(ws.cell(row=row, column=ci), h)

    answers = [
        ["Coral reef",    1240, 40,  100, -60,  1180],
        ["Seagrass beds", 860,  35,  70,  -35,  825],
        ["Mangroves",     315,  40,  15,   25,  340],
        ["Total",         2415, 115, 185, -70,  2345],
    ]
    for i, ans in enumerate(answers):
        r = row + 1 + i
        if ans[0] == "Total":
            header_style(ws.cell(row=r, column=1), ans[0])
            for ci in range(1, 6):
                header_style(ws.cell(row=r, column=ci + 1), ans[ci])
        else:
            row_header(ws.cell(row=r, column=1), ans[0])
            # Opening is given data
            data(ws.cell(row=r, column=2), ans[1])
            # Answer cells in green
            for ci in range(2, 6):
                answer_cell(ws.cell(row=r, column=ci + 1), ans[ci])

    # Percentage change answer
    row = 11
    ws.merge_cells(f"A{row}:D{row}")
    c = ws.cell(row=row, column=1)
    c.value = "Part 2: Percentage Change"
    c.font = font_subtitle

    row = 12
    pct_headers = ["Ecosystem type", "Opening (ha)", "Closing (ha)", "% change"]
    for ci, h in enumerate(pct_headers, 1):
        header_style(ws.cell(row=row, column=ci), h)

    pct_answers = [
        ["Coral reef",    1240, 1180, "-4.8%"],
        ["Seagrass beds", 860,  825,  "-4.1%"],
        ["Mangroves",     315,  340,  "+7.9%"],
    ]
    for i, ans in enumerate(pct_answers):
        r = row + 1 + i
        row_header(ws.cell(row=r, column=1), ans[0])
        data(ws.cell(row=r, column=2), ans[1])
        data(ws.cell(row=r, column=3), ans[2])
        answer_cell(ws.cell(row=r, column=4), ans[3])

    # Explanation
    row = 17
    ws.merge_cells(f"A{row}:G{row}")
    c = ws.cell(row=row, column=1)
    c.value = ("Note: Additions and reductions come from the 'Total additions' row "
               "and 'Total reductions' column of the change matrix. "
               "Net change = Additions - Reductions. "
               "Closing = Opening + Net change. "
               "% change = (Net change / Opening) x 100.")
    c.font = Font(name="Arial", size=10, color=TEXT_DARK, italic=True)
    c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[row].height = 40


def main():
    wb = openpyxl.Workbook()
    build_exercise(wb)
    build_answers(wb)
    wb.save(OUTPUT)
    print(f"Saved: {OUTPUT}")


if __name__ == "__main__":
    main()
