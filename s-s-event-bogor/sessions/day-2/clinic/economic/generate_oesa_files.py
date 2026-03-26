#!/usr/bin/env python3
"""
Generate OESA Excel template and exercise files.
Produces:
  - oesa_template.xlsx  (6 sheets)
  - oesa_exercise.xlsx  (2 sheets)
"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# GOAP colour palette
# ---------------------------------------------------------------------------
GREEN = "3B9C7B"
TEAL = "0A5455"
GRAY = "404040"
TEXT = "30302F"
WHITE = "FFFFFF"
YELLOW = "FFF2CC"

# Reusable styles
font_header = Font(name="Arial", size=10, bold=True, color=WHITE)
font_row_header = Font(name="Arial", size=10, bold=True, color=WHITE)
font_normal = Font(name="Arial", size=10, color=TEXT)
font_bold = Font(name="Arial", size=10, bold=True, color=TEXT)
font_title = Font(name="Arial", size=14, bold=True, color=TEAL)

fill_green = PatternFill(start_color=GREEN, end_color=GREEN, fill_type="solid")
fill_teal = PatternFill(start_color=TEAL, end_color=TEAL, fill_type="solid")
fill_yellow = PatternFill(start_color=YELLOW, end_color=YELLOW, fill_type="solid")
fill_white = PatternFill(start_color=WHITE, end_color=WHITE, fill_type="solid")

thin_border = Border(
    left=Side(style="thin", color=GRAY),
    right=Side(style="thin", color=GRAY),
    top=Side(style="thin", color=GRAY),
    bottom=Side(style="thin", color=GRAY),
)

align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
align_right = Alignment(horizontal="right", vertical="center", wrap_text=True)

OUT_DIR = os.path.dirname(os.path.abspath(__file__))


# ---------------------------------------------------------------------------
# Helper utilities
# ---------------------------------------------------------------------------
def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_title(ws, row, col, text, merge_end_col=None):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = font_title
    cell.alignment = align_left
    if merge_end_col:
        ws.merge_cells(
            start_row=row, start_column=col, end_row=row, end_column=merge_end_col
        )


def write_header_row(ws, row, values, fill=None):
    f = fill or fill_green
    for col_idx, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col_idx, value=val)
        cell.font = font_header
        cell.fill = f
        cell.alignment = align_center
        cell.border = thin_border


def write_data_row(ws, row, values, styles=None):
    """styles is a list of dicts per cell, keys: fill, font, align, fmt."""
    for col_idx, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col_idx, value=val)
        cell.font = font_normal
        cell.border = thin_border
        cell.alignment = align_left
        if styles and col_idx <= len(styles) and styles[col_idx - 1]:
            s = styles[col_idx - 1]
            if "fill" in s:
                cell.fill = s["fill"]
            if "font" in s:
                cell.font = s["font"]
            if "align" in s:
                cell.alignment = s["align"]
            if "fmt" in s:
                cell.number_format = s["fmt"]


def yellow_style(align=None, fmt=None):
    d = {"fill": fill_yellow}
    if align:
        d["align"] = align
    if fmt:
        d["fmt"] = fmt
    return d


def formula_style(fmt=None):
    d = {"font": font_bold}
    if fmt:
        d["fmt"] = fmt
    return d


# ===================================================================
# FILE A: oesa_template.xlsx
# ===================================================================
def build_template():
    wb = Workbook()

    # ------------------------------------------------------------------
    # Sheet 1: Scope Definition
    # ------------------------------------------------------------------
    ws = wb.active
    ws.title = "Scope Definition"
    set_col_widths(ws, [30, 30, 14, 14, 30])

    write_title(ws, 1, 1, "Step 1-2: Define Scope and Classify Activities", 5)
    ws.cell(row=2, column=1, value="Fill in your national industries and mark whether they are in scope.").font = font_normal

    headers = ["Ocean Economy Group", "National Industry", "ISIC Code", "In Scope (Y/N)", "Notes"]
    write_header_row(ws, 4, headers)

    groups = {
        "Living Resources": [
            ("Marine fishing", "0311"),
            ("Marine aquaculture", "0321"),
            ("Fish processing", "1020"),
        ],
        "Marine Minerals": [
            ("Offshore oil & gas extraction", "0610"),
            ("Marine mining", "0810"),
            ("Salt extraction", "0893"),
        ],
        "Marine Construction": [
            ("Port & harbour construction", "4290"),
            ("Coastal protection works", "4290"),
            ("Submarine cable laying", "4220"),
        ],
        "Ship and Boat Building": [
            ("Shipbuilding", "3011"),
            ("Boat building", "3012"),
            ("Ship repair & maintenance", "3315"),
        ],
        "Marine Transport": [
            ("Ocean freight transport", "5012"),
            ("Coastal passenger transport", "5011"),
            ("Port & terminal operations", "5222"),
            ("Marine cargo handling", "5224"),
        ],
        "Coastal Tourism": [
            ("Accommodation (coastal)", "5510"),
            ("Food & beverage (coastal)", "5610"),
            ("Recreation & sports (marine)", "9319"),
            ("Travel agencies (marine tours)", "7911"),
        ],
    }

    row = 5
    teal_s = {"fill": fill_teal, "font": font_row_header}
    for group, industries in groups.items():
        # Group header row
        write_data_row(ws, row, [group, "", "", "", ""], [teal_s, teal_s, teal_s, teal_s, teal_s])
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        row += 1
        for name, isic in industries:
            write_data_row(
                ws, row,
                [group, name, isic, "", ""],
                [None, None, {"align": align_center}, yellow_style(align_center), yellow_style()],
            )
            row += 1
        # Extra blank sub-row for user
        write_data_row(
            ws, row,
            [group, "", "", "", ""],
            [None, yellow_style(), yellow_style(align_center), yellow_style(align_center), yellow_style()],
        )
        row += 1

    # ------------------------------------------------------------------
    # Sheet 2: Ocean Partials
    # ------------------------------------------------------------------
    ws2 = wb.create_sheet("Ocean Partials")
    set_col_widths(ws2, [30, 14, 20, 28, 18])

    write_title(ws2, 1, 1, "Step 4: Ocean Economy Partials", 5)
    ws2.cell(row=2, column=1, value="Enter the ocean partial for each industry. Use South Africa examples as a starting point.").font = font_normal

    headers2 = ["Industry", "ISIC Code", "Ocean Partial (%)", "Data Source", "Confidence (RAG)"]
    write_header_row(ws2, 4, headers2)

    partials_data = [
        ("Marine fishing", "0311", 100, "By definition", "Green"),
        ("Marine aquaculture", "0321", 100, "By definition", "Green"),
        ("Fish processing", "1020", 85, "Industry survey", "Amber"),
        ("Offshore oil & gas", "0610", 100, "By definition", "Green"),
        ("Marine mining", "0810", 30, "Expert estimate", "Red"),
        ("Salt extraction", "0893", 50, "Expert estimate", "Amber"),
        ("Port construction", "4290", 60, "Budget records", "Amber"),
        ("Coastal protection", "4290", 100, "By definition", "Green"),
        ("Shipbuilding", "3011", 100, "By definition", "Green"),
        ("Boat building", "3012", 100, "By definition", "Green"),
        ("Ship repair", "3315", 80, "Industry survey", "Amber"),
        ("Ocean freight", "5012", 100, "By definition", "Green"),
        ("Coastal passenger", "5011", 100, "By definition", "Green"),
        ("Port operations", "5222", 100, "By definition", "Green"),
        ("Accommodation (coastal)", "5510", 35, "Visitor survey", "Amber"),
        ("Food & beverage (coastal)", "5610", 20, "Geographic proxy", "Red"),
        ("Marine recreation", "9319", 45, "Activity survey", "Amber"),
        ("Travel agencies (marine)", "7911", 25, "Expert estimate", "Red"),
    ]

    for i, (name, isic, partial, source, rag) in enumerate(partials_data):
        r = 5 + i
        write_data_row(
            ws2, r,
            [name, isic, partial / 100.0, source, rag],
            [
                None,
                {"align": align_center},
                yellow_style(align_center, "0%"),
                yellow_style(),
                yellow_style(align_center),
            ],
        )

    # ------------------------------------------------------------------
    # Sheet 3: Output Table
    # ------------------------------------------------------------------
    ws3 = wb.create_sheet("Output Table")
    set_col_widths(ws3, [30, 22, 20, 22])

    write_title(ws3, 1, 1, "OESA: Gross Output by Ocean Economy Group", 4)
    ws3.cell(row=2, column=1, value="Enter National Output and Ocean Partial. Ocean Output is calculated automatically.").font = font_normal

    headers3 = ["Ocean Economy Group", "National Output (M)", "Ocean Partial", "Ocean Output (M)"]
    write_header_row(ws3, 4, headers3)

    output_groups = [
        "Living Resources",
        "Marine Minerals",
        "Marine Construction",
        "Ship and Boat Building",
        "Marine Transport",
        "Coastal Tourism",
    ]

    for i, g in enumerate(output_groups):
        r = 5 + i
        b_col = get_column_letter(2)
        c_col = get_column_letter(3)
        formula = f"={b_col}{r}*{c_col}{r}"
        write_data_row(
            ws3, r,
            [g, None, None, formula],
            [
                {"font": font_bold},
                yellow_style(align_right, '#,##0.0'),
                yellow_style(align_center, '0%'),
                formula_style('#,##0.0'),
            ],
        )

    # Total row
    total_row = 5 + len(output_groups)
    d_col = get_column_letter(4)
    sum_formula = f"=SUM({d_col}5:{d_col}{total_row - 1})"
    write_data_row(
        ws3, total_row,
        ["TOTAL", None, None, sum_formula],
        [
            {"font": font_bold, "fill": fill_green, "font": font_header},
            {"fill": fill_green, "font": font_header, "fmt": '#,##0.0'},
            {"fill": fill_green, "font": font_header},
            {"fill": fill_green, "font": font_header, "fmt": '#,##0.0'},
        ],
    )

    # ------------------------------------------------------------------
    # Sheet 4: GVA Table
    # ------------------------------------------------------------------
    ws4 = wb.create_sheet("GVA Table")
    set_col_widths(ws4, [30, 22, 22, 22, 18])

    write_title(ws4, 1, 1, "OESA: Gross Value Added", 5)
    ws4.cell(row=2, column=1, value="Ocean Output links from Output Table. Enter Ocean IC. GVA = Output - IC.").font = font_normal

    headers4 = ["Ocean Economy Group", "Ocean Output (M)", "Ocean IC (M)", "Ocean GVA (M)", "Employment"]
    write_header_row(ws4, 4, headers4)

    for i, g in enumerate(output_groups):
        r = 5 + i
        out_ref = f"='Output Table'!D{r}"
        gva_formula = f"=B{r}-C{r}"
        write_data_row(
            ws4, r,
            [g, out_ref, None, gva_formula, None],
            [
                {"font": font_bold},
                formula_style('#,##0.0'),
                yellow_style(align_right, '#,##0.0'),
                formula_style('#,##0.0'),
                yellow_style(align_right, '#,##0'),
            ],
        )

    total_r = 5 + len(output_groups)
    write_data_row(
        ws4, total_r,
        [
            "TOTAL",
            f"=SUM(B5:B{total_r - 1})",
            f"=SUM(C5:C{total_r - 1})",
            f"=SUM(D5:D{total_r - 1})",
            f"=SUM(E5:E{total_r - 1})",
        ],
        [
            {"font": font_header, "fill": fill_green},
            {"font": font_header, "fill": fill_green, "fmt": '#,##0.0'},
            {"font": font_header, "fill": fill_green, "fmt": '#,##0.0'},
            {"font": font_header, "fill": fill_green, "fmt": '#,##0.0'},
            {"font": font_header, "fill": fill_green, "fmt": '#,##0'},
        ],
    )

    # ------------------------------------------------------------------
    # Sheet 5: Summary
    # ------------------------------------------------------------------
    ws5 = wb.create_sheet("Summary")
    set_col_widths(ws5, [30, 22, 22])

    write_title(ws5, 1, 1, "OESA Summary", 3)

    # Key metrics
    r = 3
    write_header_row(ws5, r, ["Key Metric", "Value", ""])

    ws5.cell(row=4, column=1, value="Total Ocean GVA (M)").font = font_bold
    ws5.cell(row=4, column=1).border = thin_border
    c = ws5.cell(row=4, column=2, value=f"='GVA Table'!D{total_r}")
    c.font = font_bold
    c.border = thin_border
    c.number_format = '#,##0.0'

    ws5.cell(row=5, column=1, value="National GDP (M)").font = font_bold
    ws5.cell(row=5, column=1).border = thin_border
    c = ws5.cell(row=5, column=2)
    c.fill = fill_yellow
    c.border = thin_border
    c.number_format = '#,##0.0'

    ws5.cell(row=6, column=1, value="Ocean % of GDP").font = font_bold
    ws5.cell(row=6, column=1).border = thin_border
    c = ws5.cell(row=6, column=2, value="=IF(B5>0,B4/B5,0)")
    c.font = font_bold
    c.border = thin_border
    c.number_format = '0.0%'

    # Breakdown by group
    r = 8
    write_header_row(ws5, r, ["Ocean Economy Group", "Ocean GVA (M)", "Employment"])
    for i, g in enumerate(output_groups):
        rr = r + 1 + i
        ws5.cell(row=rr, column=1, value=g).font = font_bold
        ws5.cell(row=rr, column=1).border = thin_border

        c = ws5.cell(row=rr, column=2, value=f"='GVA Table'!D{5 + i}")
        c.font = font_normal
        c.border = thin_border
        c.number_format = '#,##0.0'

        c = ws5.cell(row=rr, column=3, value=f"='GVA Table'!E{5 + i}")
        c.font = font_normal
        c.border = thin_border
        c.number_format = '#,##0'

    tr = r + 1 + len(output_groups)
    write_data_row(
        ws5, tr,
        ["TOTAL", f"=SUM(B{r+1}:B{tr-1})", f"=SUM(C{r+1}:C{tr-1})"],
        [
            {"font": font_header, "fill": fill_green},
            {"font": font_header, "fill": fill_green, "fmt": '#,##0.0'},
            {"font": font_header, "fill": fill_green, "fmt": '#,##0'},
        ],
    )

    # ------------------------------------------------------------------
    # Sheet 6: Instructions
    # ------------------------------------------------------------------
    ws6 = wb.create_sheet("Instructions")
    set_col_widths(ws6, [80])

    write_title(ws6, 1, 1, "How to Use This Template")
    instructions = [
        "Step 1: Go to the 'Scope Definition' sheet. List your country's ocean-related industries and classify them into groups.",
        "Step 2: Mark each industry as in scope (Y) or out of scope (N).",
        "Step 3: Go to the 'Ocean Partials' sheet. For each in-scope industry, enter the ocean partial (the share of that industry attributable to the ocean).",
        "Step 4: Go to the 'Output Table' sheet. Enter national gross output for each ocean economy group. The ocean output will be calculated automatically.",
        "Step 5: Go to the 'GVA Table' sheet. Enter intermediate consumption (IC) for each group. GVA = Output - IC will be calculated automatically.",
        "Step 6: Go to the 'Summary' sheet. Enter your national GDP. The ocean share of GDP will be calculated automatically.",
        "",
        "Tips:",
        "- Yellow cells are for your input.",
        "- White cells with formulas will calculate automatically.",
        "- Use the RAG column in Ocean Partials to indicate your confidence: Green (high), Amber (medium), Red (low).",
        "- Refer to GOAP Technical Guidance for detailed methodology.",
    ]
    for i, line in enumerate(instructions):
        c = ws6.cell(row=3 + i, column=1, value=line)
        c.font = font_normal
        c.alignment = Alignment(wrap_text=True)

    # Save
    path = os.path.join(OUT_DIR, "oesa_template.xlsx")
    wb.save(path)
    print(f"Created: {path}")


# ===================================================================
# FILE B: oesa_exercise.xlsx
# ===================================================================
def build_exercise():
    wb = Workbook()

    # ------------------------------------------------------------------
    # Sheet 1: Exercise
    # ------------------------------------------------------------------
    ws = wb.active
    ws.title = "Exercise"
    set_col_widths(ws, [26, 20, 18, 16, 16, 20, 18, 18])

    write_title(ws, 1, 1, "OESA Exercise: Small Island Developing State", 8)
    ws.cell(row=2, column=1, value="Use the given data to calculate ocean output, ocean IC, and ocean GVA for each sector.").font = font_normal

    # Given data
    ws.cell(row=4, column=1, value="National GDP (USD M):").font = font_bold
    ws.cell(row=4, column=1).border = thin_border
    c = ws.cell(row=4, column=2, value=2000)
    c.font = font_bold
    c.border = thin_border
    c.number_format = '#,##0'

    headers = [
        "Ocean Economy Group",
        "National Output (M)",
        "National IC (M)",
        "Ocean Partial",
        "Employment",
        "Ocean Output (M)",
        "Ocean IC (M)",
        "Ocean GVA (M)",
    ]
    write_header_row(ws, 6, headers)

    # Given data rows: (group, output, IC, partial, employment)
    given_data = [
        ("Fisheries (Living Resources)", 190, 100, 0.90, 5000),
        ("Marine Mining", 450, 280, 0.15, 2000),
        ("Marine Construction", 500, 315, 0.30, 8000),
        ("Water Transport", 150, 88, 0.80, 3500),
        ("Accommodation (Coastal Tourism)", 600, 315, 0.40, 12000),
        ("Boat Building", 50, 29, 0.70, 1500),
    ]

    for i, (group, output, ic, partial, emp) in enumerate(given_data):
        r = 7 + i
        write_data_row(
            ws, r,
            [group, output, ic, partial, emp, None, None, None],
            [
                {"font": font_bold},
                {"align": align_right, "fmt": '#,##0'},
                {"align": align_right, "fmt": '#,##0'},
                {"align": align_center, "fmt": '0%'},
                {"align": align_right, "fmt": '#,##0'},
                yellow_style(align_right, '#,##0.0'),
                yellow_style(align_right, '#,##0.0'),
                yellow_style(align_right, '#,##0.0'),
            ],
        )

    # Total row
    total_r = 7 + len(given_data)
    write_data_row(
        ws, total_r,
        ["TOTAL", None, None, None, None, None, None, None],
        [
            {"font": font_header, "fill": fill_green},
            {"font": font_header, "fill": fill_green},
            {"font": font_header, "fill": fill_green},
            {"font": font_header, "fill": fill_green},
            {"font": font_header, "fill": fill_green},
            yellow_style(align_right, '#,##0.0'),
            yellow_style(align_right, '#,##0.0'),
            yellow_style(align_right, '#,##0.0'),
        ],
    )

    # Summary section
    sr = total_r + 2
    write_title(ws, sr, 1, "Summary", 3)
    sr += 1
    write_header_row(ws, sr, ["Metric", "Value", ""], fill_green)

    metrics = [
        ("Total Ocean GVA (M)", None),
        ("National GDP (M)", 2000),
        ("Ocean % of GDP", None),
    ]
    for i, (metric, val) in enumerate(metrics):
        r = sr + 1 + i
        ws.cell(row=r, column=1, value=metric).font = font_bold
        ws.cell(row=r, column=1).border = thin_border
        c = ws.cell(row=r, column=2, value=val)
        c.border = thin_border
        if val is None:
            c.fill = fill_yellow
        c.font = font_bold
        if i == 2:
            c.number_format = '0.0%'
        else:
            c.number_format = '#,##0.0'

    # Hints
    hr = sr + 5
    ws.cell(row=hr, column=1, value="Hints:").font = font_bold
    ws.cell(row=hr + 1, column=1, value="Ocean Output = National Output x Ocean Partial").font = font_normal
    ws.cell(row=hr + 2, column=1, value="Ocean IC = National IC x Ocean Partial").font = font_normal
    ws.cell(row=hr + 3, column=1, value="Ocean GVA = Ocean Output - Ocean IC").font = font_normal
    ws.cell(row=hr + 4, column=1, value="Ocean % of GDP = Total Ocean GVA / National GDP").font = font_normal

    # ------------------------------------------------------------------
    # Sheet 2: Answers
    # ------------------------------------------------------------------
    ws2 = wb.create_sheet("Answers")
    set_col_widths(ws2, [26, 20, 18, 16, 20, 18, 18])

    fill_answer = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    write_title(ws2, 1, 1, "OESA Exercise: Answer Key", 7)

    headers2 = [
        "Ocean Economy Group",
        "National Output (M)",
        "Ocean Partial",
        "Employment",
        "Ocean Output (M)",
        "Ocean IC (M)",
        "Ocean GVA (M)",
    ]
    write_header_row(ws2, 3, headers2)

    # Answers: (group, natl_output, partial, emp, ocean_output, ocean_ic, ocean_gva)
    answers = [
        ("Fisheries (Living Resources)", 190, 0.90, 5000, 171, 90.25, 80.75),
        ("Marine Mining", 450, 0.15, 2000, 67.5, 42, 25.5),
        ("Marine Construction", 500, 0.30, 8000, 150, 95, 55),  # 500*0.3=150, 315*0.3=94.5 ~ 95
        ("Water Transport", 150, 0.80, 3500, 120, 70, 50),  # 150*0.8=120, 88*0.8=70.4 ~ 70
        ("Accommodation (Coastal Tourism)", 600, 0.40, 12000, 240, 126, 114),
        ("Boat Building", 50, 0.70, 1500, 35, 20, 15),  # 50*0.7=35, 29*0.7=20.3 ~ 20
    ]

    ans_style = {"fill": fill_answer, "align": align_right, "fmt": '#,##0.0'}

    for i, (group, output, partial, emp, o_out, o_ic, o_gva) in enumerate(answers):
        r = 4 + i
        write_data_row(
            ws2, r,
            [group, output, partial, emp, o_out, o_ic, o_gva],
            [
                {"font": font_bold},
                {"align": align_right, "fmt": '#,##0'},
                {"align": align_center, "fmt": '0%'},
                {"align": align_right, "fmt": '#,##0'},
                ans_style,
                ans_style,
                ans_style,
            ],
        )

    total_r2 = 4 + len(answers)
    total_gva = sum(a[6] for a in answers)
    total_output = sum(a[4] for a in answers)
    total_ic = sum(a[5] for a in answers)
    write_data_row(
        ws2, total_r2,
        ["TOTAL", None, None, None, total_output, total_ic, total_gva],
        [
            {"font": font_header, "fill": fill_green},
            {"font": font_header, "fill": fill_green},
            {"font": font_header, "fill": fill_green},
            {"font": font_header, "fill": fill_green},
            {"font": font_header, "fill": fill_green, "fmt": '#,##0.0'},
            {"font": font_header, "fill": fill_green, "fmt": '#,##0.0'},
            {"font": font_header, "fill": fill_green, "fmt": '#,##0.0'},
        ],
    )

    # Summary
    sr2 = total_r2 + 2
    write_header_row(ws2, sr2, ["Metric", "Value", "", "", "", "", ""])
    summary_answers = [
        ("Total Ocean GVA (M)", total_gva, '#,##0.00'),
        ("National GDP (M)", 2000, '#,##0'),
        ("Ocean % of GDP", total_gva / 2000, '0.0%'),
    ]
    for i, (metric, val, fmt) in enumerate(summary_answers):
        r = sr2 + 1 + i
        ws2.cell(row=r, column=1, value=metric).font = font_bold
        ws2.cell(row=r, column=1).border = thin_border
        c = ws2.cell(row=r, column=2, value=val)
        c.font = font_bold
        c.fill = fill_answer
        c.border = thin_border
        c.number_format = fmt

    path = os.path.join(OUT_DIR, "oesa_exercise.xlsx")
    wb.save(path)
    print(f"Created: {path}")


# ===================================================================
if __name__ == "__main__":
    build_template()
    build_exercise()
    print("Done.")
