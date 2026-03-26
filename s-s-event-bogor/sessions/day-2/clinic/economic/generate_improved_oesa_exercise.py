#!/usr/bin/env python3
"""
Generate an improved OESA (Ocean Economy Satellite Account) exercise workbook.
Walks fellows through the 7-step OESA compilation process with step-by-step guidance.
"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# -- GOAP colour palette ---------------------------------------------------
GREEN  = "3B9C7B"
TEAL   = "0A5455"
GRAY   = "404040"
TEXT   = "30302F"
WHITE  = "FFFFFF"
YELLOW = "FFF2CC"

# -- Reusable styles -------------------------------------------------------
font_title   = Font(name="Arial", size=14, bold=True, color=TEAL)
font_section = Font(name="Arial", size=12, bold=True, color=TEAL)
font_header  = Font(name="Arial", size=10, bold=True, color=WHITE)
font_row_hdr = Font(name="Arial", size=10, bold=True, color=WHITE)
font_normal  = Font(name="Arial", size=10, color=TEXT)
font_bold    = Font(name="Arial", size=10, bold=True, color=TEXT)
font_italic  = Font(name="Arial", size=10, italic=True, color=GRAY)
font_answer  = Font(name="Arial", size=10, color=TEAL)

fill_green  = PatternFill("solid", fgColor=GREEN)
fill_teal   = PatternFill("solid", fgColor=TEAL)
fill_yellow = PatternFill("solid", fgColor=YELLOW)
fill_white  = PatternFill("solid", fgColor=WHITE)
fill_light  = PatternFill("solid", fgColor="E8F5F0")
fill_answer = PatternFill("solid", fgColor="D5F5E3")

align_wrap   = Alignment(wrap_text=True, vertical="top")
align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
align_left   = Alignment(horizontal="left", vertical="top", wrap_text=True)

thin_side  = Side(style="thin", color=GRAY)
border_all = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)


# -- Helper functions -------------------------------------------------------
def style_cell(cell, font=font_normal, fill=None, alignment=align_wrap,
               border=None, number_format=None):
    cell.font = font
    if fill:
        cell.fill = fill
    cell.alignment = alignment
    if border:
        cell.border = border
    if number_format:
        cell.number_format = number_format


def write_title(ws, row, col, text):
    c = ws.cell(row=row, column=col, value=text)
    style_cell(c, font=font_title)
    return row + 1


def write_section(ws, row, col, text):
    c = ws.cell(row=row, column=col, value=text)
    style_cell(c, font=font_section)
    return row + 1


def write_text(ws, row, col, text, font_=font_normal, fill_=None):
    c = ws.cell(row=row, column=col, value=text)
    style_cell(c, font=font_, fill=fill_)
    return row + 1


def header_row(ws, row, headers, col_start=1):
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=col_start + i, value=h)
        style_cell(c, font=font_header, fill=fill_green, alignment=align_center,
                   border=border_all)


def teal_label(ws, row, col, text):
    c = ws.cell(row=row, column=col, value=text)
    style_cell(c, font=font_row_hdr, fill=fill_teal, alignment=align_wrap,
               border=border_all)


def data_cell(ws, row, col, value=None, is_input=False, font_=font_normal,
              fmt=None):
    c = ws.cell(row=row, column=col, value=value)
    fill_ = fill_yellow if is_input else fill_white
    style_cell(c, font=font_, fill=fill_, alignment=align_center, border=border_all,
               number_format=fmt)
    return c


def answer_cell(ws, row, col, value=None, fmt=None):
    c = ws.cell(row=row, column=col, value=value)
    style_cell(c, font=font_answer, fill=fill_answer, alignment=align_center,
               border=border_all, number_format=fmt)
    return c


def set_col_widths(ws, widths):
    for letter, w in widths.items():
        ws.column_dimensions[letter].width = w


# ===================================================================
# SHEET 1: README
# ===================================================================
def build_readme(wb):
    ws = wb.active
    ws.title = "README"
    ws.sheet_properties.tabColor = TEAL

    set_col_widths(ws, {"A": 95})

    r = 1
    r = write_title(ws, r, 1, "Ocean Economy Satellite Account (OESA) Exercise")
    r += 1
    r = write_text(ws, r, 1,
        "This workbook walks you through the 7-step OESA compilation process. "
        "By the end you will have built a complete satellite account for a "
        "hypothetical Small Island Developing State (SIDS).",
        font_=font_bold)
    r += 1

    r = write_section(ws, r, 1, "What you will do")
    r += 1

    steps = [
        ("Step 1 -- Define Scope",
         "Determine which industries belong to the ocean economy by applying "
         "three pillars: on-water production, coastal-dependent activity, and "
         "for-ocean-use manufacturing. This scoping step is essential before "
         "any data work begins."),
        ("Step 2 -- Classify Activities",
         "Map each in-scope industry to one of the 6 standard ocean economy "
         "groups (Living resources, Marine minerals, Ship/boat building, "
         "Marine construction, Marine transport, Coastal tourism). This "
         "creates your classification framework."),
        ("Step 3 -- Estimate Partials",
         "Determine what share of each national industry is ocean-related. "
         "Many industries are only partly ocean -- e.g., only 30% of hotels "
         "may be coastal. You will estimate these partials and rate your "
         "confidence using a RAG (Red/Amber/Green) system."),
        ("Step 4 -- Calculate GVA",
         "Apply your partials to national Supply-Use Table (SUT) data to "
         "calculate ocean output, intermediate consumption, and GVA for "
         "each group. Key formulas:\n"
         "  Ocean Output = National Output x Partial\n"
         "  Ocean IC = National IC x Partial\n"
         "  Ocean GVA = Ocean Output - Ocean IC\n"
         "  Ocean GDP % = Total Ocean GVA / National GDP"),
        ("Step 5 -- Build OESA Tables",
         "Transfer your calculations into formal OESA output tables "
         "(Gross Output, GVA by group, Summary) and interpret the results. "
         "This is the final analytical product."),
    ]

    for title, desc in steps:
        r = write_text(ws, r, 1, title, font_=font_bold)
        r = write_text(ws, r, 1, desc)
        r += 1

    r = write_section(ws, r, 1, "What each step teaches")
    r += 1
    lessons = [
        "Step 1: How to decide what is 'ocean economy' -- the boundary question",
        "Step 2: The standard classification used internationally for comparability",
        "Step 3: The hardest analytical judgment -- estimating the ocean share",
        "Step 4: Core SNA arithmetic -- from output through IC to GVA",
        "Step 5: How to present results for policy audiences",
    ]
    for lesson in lessons:
        r = write_text(ws, r, 1, lesson)

    r += 1
    r = write_text(ws, r, 1,
        "The Answers sheet contains a full answer key so you can check your work "
        "at every step. The Glossary sheet defines all technical terms used.",
        font_=font_italic)

    return ws


# ===================================================================
# SHEET 2: Step 1 - Define Scope
# ===================================================================
def build_step1(wb):
    ws = wb.create_sheet("Step 1 - Define Scope")
    ws.sheet_properties.tabColor = GREEN

    set_col_widths(ws, {"A": 30, "B": 12, "C": 16, "D": 20, "E": 18, "F": 14})

    r = 1
    r = write_title(ws, r, 1,
        "Step 1: Define the Scope of the Ocean Economy")
    r += 1

    r = write_text(ws, r, 1,
        "Scenario: You are compiling an OESA for a Small Island Developing State.",
        font_=font_bold)
    r += 1

    r = write_section(ws, r, 1, "Geographic scope")
    r = write_text(ws, r, 1,
        "The ocean economy covers: the Exclusive Economic Zone (EEZ), "
        "territorial waters, and coastal municipalities.")
    r += 1

    r = write_section(ws, r, 1, "Three pillars of inclusion")
    r += 1

    pillars = [
        ("1. On-water production",
         "Activities that take place on or under the ocean surface "
         "(e.g., fishing, shipping, offshore mining)."),
        ("2. Coastal-dependent",
         "Activities located on the coast whose location is determined "
         "by proximity to the ocean (e.g., coastal hotels, port services)."),
        ("3. For-ocean-use",
         "Manufacturing or services that produce goods specifically for "
         "ocean use (e.g., ship building, marine equipment)."),
    ]
    for title, desc in pillars:
        r = write_text(ws, r, 1, title, font_=font_bold)
        r = write_text(ws, r, 1, desc)
        r += 1

    r = write_text(ws, r, 1,
        "Task: For each industry below, decide whether it is in scope (Y or N). "
        "Consider which pillar(s) apply.",
        font_=font_italic)
    r += 1

    headers = ["Industry", "ISIC code", "On-water?", "Coastal-\ndependent?",
               "For-ocean-use?", "In scope?"]
    header_row(ws, r, headers)
    r += 1

    industries = [
        ("Marine capture fisheries", "031", "Y", "", ""),
        ("Aquaculture", "032", "", "Y", ""),
        ("Offshore oil and gas", "061", "Y", "", ""),
        ("Coastal hotels", "551", "", "Y", ""),
        ("Ship building", "301", "", "", "Y"),
        ("Water transport", "501", "Y", "", ""),
        ("Inland farming", "011", "", "", ""),
        ("Software development", "620", "", "", ""),
        ("Coastal restaurants", "561", "", "Y", ""),
        ("Port services", "522", "", "Y", "Y"),
    ]

    for name, isic, on_water, coastal, for_ocean in industries:
        teal_label(ws, r, 1, name)
        data_cell(ws, r, 2, isic)
        data_cell(ws, r, 3, on_water, font_=font_normal)
        data_cell(ws, r, 4, coastal, font_=font_normal)
        data_cell(ws, r, 5, for_ocean, font_=font_normal)
        data_cell(ws, r, 6, is_input=True)  # yellow -- fellows decide
        r += 1

    r += 1
    r = write_text(ws, r, 1,
        "Hint: An industry is in scope if at least one pillar applies (Y in any "
        "of the three pillar columns). Inland farming and software development "
        "have no ocean connection.",
        font_=font_italic)

    return ws


# ===================================================================
# SHEET 3: Step 2 - Classify Activities
# ===================================================================
def build_step2(wb):
    ws = wb.create_sheet("Step 2 - Classify Activities")
    ws.sheet_properties.tabColor = GREEN

    set_col_widths(ws, {"A": 30, "B": 12, "C": 32})

    r = 1
    r = write_title(ws, r, 1,
        "Step 2: Map Industries to Ocean Economy Groups")
    r += 1
    r = write_text(ws, r, 1,
        "Now that you have identified the in-scope industries from Step 1, "
        "assign each to one of the 6 standard ocean economy groups.",
        font_=font_italic)
    r += 1

    # Reference table of 6 groups
    r = write_section(ws, r, 1, "Reference: 6 Standard Ocean Economy Groups")
    r += 1
    groups_ref = [
        ("1. Living resources", "Fisheries, aquaculture, seaweed harvesting"),
        ("2. Marine minerals", "Offshore oil/gas extraction, seabed mining, salt"),
        ("3. Ship/boat building", "Construction and repair of ships and boats"),
        ("4. Marine construction", "Coastal infrastructure, port construction, seawalls"),
        ("5. Marine transport", "Shipping, ferries, water freight, port services"),
        ("6. Coastal tourism", "Coastal accommodation, restaurants, recreation"),
    ]
    header_row(ws, r, ["Ocean economy group", "Description", "Examples"])
    r += 1
    for gname, examples in groups_ref:
        teal_label(ws, r, 1, gname)
        data_cell(ws, r, 2, examples, font_=font_normal)
        # col C not used here
        r += 1

    r += 2

    # Task table
    r = write_section(ws, r, 1, "Task: Classify the in-scope industries")
    r = write_text(ws, r, 1,
        "For each in-scope industry, type the name of the ocean economy group "
        "it belongs to in the yellow cell.",
        font_=font_italic)
    r += 1

    headers = ["In-scope industry", "ISIC", "Ocean economy group\n(choose from above)"]
    header_row(ws, r, headers)
    r += 1

    in_scope = [
        ("Marine capture fisheries", "031"),
        ("Aquaculture", "032"),
        ("Offshore oil and gas", "061"),
        ("Coastal hotels", "551"),
        ("Ship building", "301"),
        ("Water transport", "501"),
        ("Coastal restaurants", "561"),
        ("Port services", "522"),
    ]

    for name, isic in in_scope:
        teal_label(ws, r, 1, name)
        data_cell(ws, r, 2, isic)
        data_cell(ws, r, 3, is_input=True)  # yellow -- fellows assign group
        r += 1

    return ws


# ===================================================================
# SHEET 4: Step 3 - Estimate Partials
# ===================================================================
def build_step3(wb):
    ws = wb.create_sheet("Step 3 - Estimate Partials")
    ws.sheet_properties.tabColor = GREEN

    set_col_widths(ws, {"A": 22, "B": 22, "C": 22, "D": 34, "E": 20, "F": 16})

    r = 1
    r = write_title(ws, r, 1,
        "Step 3: Estimate Ocean Economy Partials")
    r += 1
    r = write_text(ws, r, 1,
        "Most industries are not 100% ocean-related. A 'partial' is the share "
        "(0 to 1) of national industry output attributable to the ocean economy.",
        font_=font_bold)
    r += 1
    r = write_text(ws, r, 1,
        "Use the hints provided to estimate each partial. Then rate your "
        "confidence: G = Green (strong data), A = Amber (some uncertainty), "
        "R = Red (rough estimate).",
        font_=font_italic)
    r += 1

    headers = ["Ocean group", "Industry", "National output\n(USD million)",
               "Hint for partial estimation", "Your partial\nestimate (0-1)",
               "Confidence\n(G/A/R)"]
    header_row(ws, r, headers)
    r += 1

    rows = [
        ("Living resources", "Fisheries", 180,
         "95% of catch is marine,\n5% is freshwater"),
        ("Marine minerals", "Mining", 450,
         "Only offshore gas = 15%\nof total mining output"),
        ("Ship/boat building", "Boat building", 35,
         "All output is marine\nvessel construction"),
        ("Marine construction", "Construction", 600,
         "25% of construction is\ncoastal infrastructure"),
        ("Marine transport", "Water transport", 120,
         "All water transport is\nmarine (no inland waterways)"),
        ("Coastal tourism", "Accommodation", 500,
         "30% of accommodation\nis in coastal areas"),
        ("Coastal tourism", "Food services", 300,
         "30% of food services\nare in coastal areas"),
    ]

    for group, industry, output, hint in rows:
        teal_label(ws, r, 1, group)
        data_cell(ws, r, 2, industry, font_=font_normal)
        data_cell(ws, r, 3, output, fmt="#,##0")
        data_cell(ws, r, 4, hint, font_=font_normal)
        data_cell(ws, r, 5, is_input=True)  # yellow -- partial estimate
        data_cell(ws, r, 6, is_input=True)  # yellow -- RAG confidence
        r += 1

    r += 1
    r = write_section(ws, r, 1, "Guidance on estimating partials")
    r += 1
    guidance = [
        "If an industry is entirely ocean (e.g., all water transport is marine), the partial = 1.0",
        "If you have survey data on the coastal share, use it directly (e.g., 30% coastal hotels = 0.30)",
        "If the hint says '15% of total', the partial = 0.15",
        "Green confidence: based on administrative or survey data",
        "Amber confidence: based on informed estimates or proxy data",
        "Red confidence: rough assumption, needs validation",
    ]
    for g in guidance:
        r = write_text(ws, r, 1, g)

    return ws


# ===================================================================
# SHEET 5: Step 4 - Calculate GVA
# ===================================================================
def build_step4(wb):
    ws = wb.create_sheet("Step 4 - Calculate GVA")
    ws.sheet_properties.tabColor = GREEN

    set_col_widths(ws, {"A": 22, "B": 18, "C": 16, "D": 18, "E": 16, "F": 16,
                        "G": 16, "H": 16})

    r = 1
    r = write_title(ws, r, 1,
        "Step 4: Calculate Ocean Output, IC, and GVA")
    r += 1
    r = write_text(ws, r, 1,
        "Apply your partials from Step 3 to the national data below. "
        "Calculate the ocean economy values for each group.",
        font_=font_bold)
    r += 1
    r = write_text(ws, r, 1,
        "Formulas:\n"
        "  Ocean Output = National Output x Partial\n"
        "  Ocean IC = National IC x Partial\n"
        "  Ocean GVA = Ocean Output - Ocean IC\n"
        "  Employment: use the same partial as a rough estimate",
        font_=font_italic)
    r += 1

    headers = ["Ocean economy\ngroup", "National\noutput\n(USD m)",
               "Partial\n(from Step 3)", "Ocean\noutput\n(USD m)",
               "National IC\n(USD m)", "Ocean IC\n(USD m)",
               "Ocean GVA\n(USD m)", "Employment\n(persons)"]
    header_row(ws, r, headers)
    r += 1

    # National data: (group, nat_output, nat_ic, nat_employment)
    groups = [
        ("Living resources", 180, 95, 4500),
        ("Marine minerals", 450, 280, 2000),
        ("Ship/boat building", 35, 20, 800),
        ("Marine construction", 600, 380, 12000),
        ("Marine transport", 120, 70, 3000),
        ("Coastal tourism", 800, 420, 18000),
    ]

    data_rows_start = r
    for group, nat_out, nat_ic, nat_emp in groups:
        teal_label(ws, r, 1, group)
        data_cell(ws, r, 2, nat_out, fmt="#,##0")
        data_cell(ws, r, 3, is_input=True)       # partial -- yellow
        data_cell(ws, r, 4, is_input=True, fmt="#,##0.0")  # ocean output -- yellow
        data_cell(ws, r, 5, nat_ic, fmt="#,##0")
        data_cell(ws, r, 6, is_input=True, fmt="#,##0.0")  # ocean IC -- yellow
        data_cell(ws, r, 7, is_input=True, fmt="#,##0.0")  # ocean GVA -- yellow
        data_cell(ws, r, 8, is_input=True, fmt="#,##0")    # employment -- yellow
        r += 1

    # TOTAL row
    teal_label(ws, r, 1, "TOTAL")
    data_cell(ws, r, 2, "")
    data_cell(ws, r, 3, "")
    data_cell(ws, r, 4, is_input=True, fmt="#,##0.0")  # total ocean output
    data_cell(ws, r, 5, "")
    data_cell(ws, r, 6, is_input=True, fmt="#,##0.0")  # total ocean IC
    data_cell(ws, r, 7, is_input=True, fmt="#,##0.0")  # total ocean GVA
    data_cell(ws, r, 8, is_input=True, fmt="#,##0")    # total employment
    r += 2

    # National GDP and ocean share
    r = write_section(ws, r, 1, "Ocean economy as share of national GDP")
    r += 1
    r = write_text(ws, r, 1, "National GDP: USD 2,000 million", font_=font_bold)
    r += 1

    ws.cell(row=r, column=1, value="Ocean economy % of GDP:").font = font_bold
    data_cell(ws, r, 2, is_input=True, fmt="0.0%")  # yellow
    r += 1
    r = write_text(ws, r, 1,
        "Formula: Ocean GVA (total from above) / National GDP (2,000)",
        font_=font_italic)

    return ws


# ===================================================================
# SHEET 6: Step 5 - OESA Tables
# ===================================================================
def build_step5(wb):
    ws = wb.create_sheet("Step 5 - OESA Tables")
    ws.sheet_properties.tabColor = GREEN

    set_col_widths(ws, {"A": 26, "B": 18, "C": 18, "D": 18, "E": 18})

    r = 1
    r = write_title(ws, r, 1,
        "Step 5: Build the Formal OESA Output Tables")
    r += 1
    r = write_text(ws, r, 1,
        "Transfer your Step 4 results into these formal tables. "
        "These are the tables you would present to policy makers.",
        font_=font_italic)
    r += 1

    # ── Table A: Gross Output ──────────────────────────────────────
    r = write_section(ws, r, 1, "Table A: Gross Output by Ocean Economy Group")
    r += 1
    header_row(ws, r, ["Ocean economy group", "Gross output\n(USD million)"])
    r += 1

    table_a_groups = [
        "Living resources", "Marine minerals", "Ship/boat building",
        "Marine construction", "Marine transport", "Coastal tourism", "TOTAL"
    ]
    for g in table_a_groups:
        teal_label(ws, r, 1, g)
        data_cell(ws, r, 2, is_input=True, fmt="#,##0.0")
        r += 1

    r += 2

    # ── Table B: GVA by Group ─────────────────────────────────────
    r = write_section(ws, r, 1, "Table B: Gross Value Added by Ocean Economy Group")
    r += 1
    header_row(ws, r, ["Ocean economy group", "GVA\n(USD million)", "% of ocean\nGVA"])
    r += 1

    for g in table_a_groups:
        teal_label(ws, r, 1, g)
        data_cell(ws, r, 2, is_input=True, fmt="#,##0.0")
        data_cell(ws, r, 3, is_input=True, fmt="0.0%")
        r += 1

    r += 2

    # ── Table C: Summary ──────────────────────────────────────────
    r = write_section(ws, r, 1, "Table C: Ocean Economy Summary")
    r += 1
    header_row(ws, r, ["Indicator", "Value"])
    r += 1

    summary_items = [
        "Total ocean gross output (USD m)",
        "Total ocean GVA (USD m)",
        "National GDP (USD m)",
        "Ocean economy % of GDP",
        "Total ocean employment (persons)",
    ]
    for item in summary_items:
        teal_label(ws, r, 1, item)
        data_cell(ws, r, 2, is_input=True)
        r += 1

    r += 2

    # ── Interpretation questions ──────────────────────────────────
    r = write_section(ws, r, 1, "Interpretation Questions")
    r += 1
    questions = [
        "1. Which sector dominates ocean GVA? Why might this be the case for a SIDS?",
        "2. What percentage of national GDP comes from the ocean economy?",
        "3. What are the implications for ocean policy and investment priorities?",
        "4. Which sectors have the highest employment relative to GVA? What does this suggest?",
        "5. If you were advising the government, which ocean sector would you prioritise for growth?",
    ]
    for q in questions:
        r = write_text(ws, r, 1, q, font_=font_bold)
        data_cell(ws, r, 1, is_input=True)  # yellow cell for answer
        ws.cell(row=r, column=1).alignment = align_left
        r += 2

    return ws


# ===================================================================
# SHEET 7: Answers
# ===================================================================
def build_answers(wb):
    ws = wb.create_sheet("Answers")
    ws.sheet_properties.tabColor = TEAL

    set_col_widths(ws, {"A": 30, "B": 16, "C": 20, "D": 20, "E": 20, "F": 18,
                        "G": 18, "H": 18})

    r = 1
    r = write_title(ws, r, 1, "Answer Key -- All Steps")
    r += 1

    # ── Step 1 Answers ────────────────────────────────────────────
    r = write_section(ws, r, 1, "Step 1: Which industries are in scope?")
    r += 1
    r = write_text(ws, r, 1,
        "8 industries are in scope; 2 are not (Inland farming and Software development).",
        font_=font_bold)
    r += 1

    header_row(ws, r, ["Industry", "ISIC", "In scope?", "Reason"])
    r += 1

    step1_answers = [
        ("Marine capture fisheries", "031", "Y", "On-water production"),
        ("Aquaculture", "032", "Y", "Coastal-dependent"),
        ("Offshore oil and gas", "061", "Y", "On-water production"),
        ("Coastal hotels", "551", "Y", "Coastal-dependent"),
        ("Ship building", "301", "Y", "For-ocean-use"),
        ("Water transport", "501", "Y", "On-water production"),
        ("Inland farming", "011", "N", "No ocean connection"),
        ("Software development", "620", "N", "No ocean connection"),
        ("Coastal restaurants", "561", "Y", "Coastal-dependent"),
        ("Port services", "522", "Y", "Coastal-dependent + For-ocean-use"),
    ]
    for name, isic, scope, reason in step1_answers:
        teal_label(ws, r, 1, name)
        answer_cell(ws, r, 2, isic)
        answer_cell(ws, r, 3, scope)
        answer_cell(ws, r, 4, reason)
        r += 1

    r += 2

    # ── Step 2 Answers ────────────────────────────────────────────
    r = write_section(ws, r, 1, "Step 2: Correct group assignments")
    r += 1

    header_row(ws, r, ["In-scope industry", "ISIC", "Ocean economy group"])
    r += 1

    step2_answers = [
        ("Marine capture fisheries", "031", "Living resources"),
        ("Aquaculture", "032", "Living resources"),
        ("Offshore oil and gas", "061", "Marine minerals"),
        ("Coastal hotels", "551", "Coastal tourism"),
        ("Ship building", "301", "Ship/boat building"),
        ("Water transport", "501", "Marine transport"),
        ("Coastal restaurants", "561", "Coastal tourism"),
        ("Port services", "522", "Marine transport"),
    ]
    for name, isic, group in step2_answers:
        teal_label(ws, r, 1, name)
        answer_cell(ws, r, 2, isic)
        answer_cell(ws, r, 3, group)
        r += 1

    r += 2

    # ── Step 3 Answers ────────────────────────────────────────────
    r = write_section(ws, r, 1, "Step 3: Correct partials")
    r += 1

    header_row(ws, r, ["Ocean group", "Industry", "Partial", "Confidence"])
    r += 1

    step3_answers = [
        ("Living resources", "Fisheries", 0.95, "G"),
        ("Marine minerals", "Mining", 0.15, "A"),
        ("Ship/boat building", "Boat building", 1.00, "G"),
        ("Marine construction", "Construction", 0.25, "A"),
        ("Marine transport", "Water transport", 1.00, "G"),
        ("Coastal tourism", "Accommodation", 0.30, "A"),
        ("Coastal tourism", "Food services", 0.30, "A"),
    ]
    for group, industry, partial, conf in step3_answers:
        teal_label(ws, r, 1, group)
        answer_cell(ws, r, 2, industry)
        answer_cell(ws, r, 3, partial, fmt="0.00")
        answer_cell(ws, r, 4, conf)
        r += 1

    r += 2

    # ── Step 4 Answers ────────────────────────────────────────────
    r = write_section(ws, r, 1, "Step 4: Full GVA calculations")
    r += 1

    header_row(ws, r, ["Ocean group", "Nat. output", "Partial",
                       "Ocean output", "Nat. IC", "Ocean IC",
                       "Ocean GVA", "Employment"])
    r += 1

    step4_data = [
        ("Living resources", 180, 0.95, 171.0, 95, 90.25, 80.75, 4275),
        ("Marine minerals", 450, 0.15, 67.5, 280, 42.0, 25.5, 300),
        ("Ship/boat building", 35, 1.00, 35.0, 20, 20.0, 15.0, 800),
        ("Marine construction", 600, 0.25, 150.0, 380, 95.0, 55.0, 3000),
        ("Marine transport", 120, 1.00, 120.0, 70, 70.0, 50.0, 3000),
        ("Coastal tourism", 800, 0.30, 240.0, 420, 126.0, 114.0, 5400),
    ]

    for group, nat_out, partial, oc_out, nat_ic, oc_ic, oc_gva, emp in step4_data:
        teal_label(ws, r, 1, group)
        answer_cell(ws, r, 2, nat_out, fmt="#,##0")
        answer_cell(ws, r, 3, partial, fmt="0.00")
        answer_cell(ws, r, 4, oc_out, fmt="#,##0.0")
        answer_cell(ws, r, 5, nat_ic, fmt="#,##0")
        answer_cell(ws, r, 6, oc_ic, fmt="#,##0.00")
        answer_cell(ws, r, 7, oc_gva, fmt="#,##0.00")
        answer_cell(ws, r, 8, emp, fmt="#,##0")
        r += 1

    # TOTAL row
    teal_label(ws, r, 1, "TOTAL")
    answer_cell(ws, r, 2, "")
    answer_cell(ws, r, 3, "")
    answer_cell(ws, r, 4, 783.5, fmt="#,##0.0")
    answer_cell(ws, r, 5, "")
    answer_cell(ws, r, 6, 443.25, fmt="#,##0.00")
    answer_cell(ws, r, 7, 340.25, fmt="#,##0.00")
    answer_cell(ws, r, 8, 16775, fmt="#,##0")
    r += 2

    r = write_text(ws, r, 1, "National GDP: USD 2,000 million", font_=font_bold)
    r = write_text(ws, r, 1,
        "Ocean economy % of GDP: 340.25 / 2,000 = 17.0%", font_=font_bold)
    r += 1

    # ── Step 5 Answers: Completed tables ──────────────────────────
    r = write_section(ws, r, 1, "Step 5: Completed OESA tables")
    r += 1

    # Table A
    r = write_text(ws, r, 1, "Table A: Gross Output", font_=font_bold)
    r += 1
    header_row(ws, r, ["Ocean economy group", "Gross output (USD m)"])
    r += 1
    table_a = [
        ("Living resources", 171.0),
        ("Marine minerals", 67.5),
        ("Ship/boat building", 35.0),
        ("Marine construction", 150.0),
        ("Marine transport", 120.0),
        ("Coastal tourism", 240.0),
        ("TOTAL", 783.5),
    ]
    for g, val in table_a:
        teal_label(ws, r, 1, g)
        answer_cell(ws, r, 2, val, fmt="#,##0.0")
        r += 1

    r += 1

    # Table B
    r = write_text(ws, r, 1, "Table B: GVA by Group", font_=font_bold)
    r += 1
    header_row(ws, r, ["Ocean economy group", "GVA (USD m)", "% of ocean GVA"])
    r += 1
    table_b = [
        ("Living resources", 80.75, 80.75 / 340.25),
        ("Marine minerals", 25.5, 25.5 / 340.25),
        ("Ship/boat building", 15.0, 15.0 / 340.25),
        ("Marine construction", 55.0, 55.0 / 340.25),
        ("Marine transport", 50.0, 50.0 / 340.25),
        ("Coastal tourism", 114.0, 114.0 / 340.25),
        ("TOTAL", 340.25, 1.0),
    ]
    for g, gva, pct in table_b:
        teal_label(ws, r, 1, g)
        answer_cell(ws, r, 2, gva, fmt="#,##0.00")
        answer_cell(ws, r, 3, pct, fmt="0.0%")
        r += 1

    r += 1

    # Table C
    r = write_text(ws, r, 1, "Table C: Summary", font_=font_bold)
    r += 1
    header_row(ws, r, ["Indicator", "Value"])
    r += 1
    summary = [
        ("Total ocean gross output (USD m)", 783.5),
        ("Total ocean GVA (USD m)", 340.25),
        ("National GDP (USD m)", 2000),
        ("Ocean economy % of GDP", "17.0%"),
        ("Total ocean employment (persons)", 16775),
    ]
    for label, val in summary:
        teal_label(ws, r, 1, label)
        answer_cell(ws, r, 2, val)
        r += 1

    r += 2

    # ── Interpretation ────────────────────────────────────────────
    r = write_section(ws, r, 1, "Interpretation")
    r += 1

    interpretations = [
        "Coastal tourism dominates ocean GVA at 33.5% (USD 114m of 340.25m), "
        "reflecting the typical economic structure of SIDS where tourism is "
        "the primary driver of economic activity.",
        "Living resources (fisheries + aquaculture) is the second-largest "
        "contributor at 23.7% (USD 80.75m), highlighting the importance of "
        "sustainable fisheries management.",
        "The ocean economy accounts for 17.0% of national GDP -- a significant "
        "share that justifies dedicated ocean economic policy.",
        "Marine transport contributes USD 50m (14.7%) reflecting the island's "
        "dependence on shipping for trade.",
        "Marine construction (16.2%) includes coastal infrastructure essential "
        "for climate adaptation.",
        "Ship building is the smallest sector (4.4%) but may have strategic "
        "importance for self-sufficiency.",
        "Policy implication: Investment in coastal tourism and fisheries "
        "sustainability would have the highest economic impact. Climate "
        "adaptation spending on marine construction is also justified given "
        "the ocean economy's size.",
    ]
    for interp in interpretations:
        r = write_text(ws, r, 1, interp)
        r += 1

    return ws


# ===================================================================
# SHEET 8: Glossary
# ===================================================================
def build_glossary(wb):
    ws = wb.create_sheet("Glossary")
    ws.sheet_properties.tabColor = TEAL

    set_col_widths(ws, {"A": 18, "B": 80})

    r = 1
    r = write_title(ws, r, 1, "Glossary of Key Terms")
    r += 1

    header_row(ws, r, ["Term", "Definition"])
    r += 1

    terms = [
        ("GVA",
         "Gross Value Added. The value of output minus intermediate consumption. "
         "Measures the contribution of an industry to GDP."),
        ("GDP",
         "Gross Domestic Product. The total value of all final goods and services "
         "produced in a country. GDP = sum of all GVA + taxes on products - subsidies."),
        ("IC",
         "Intermediate Consumption. The value of goods and services consumed as "
         "inputs in production (e.g., fuel, raw materials, purchased services)."),
        ("Output",
         "Gross Output. The total value of goods and services produced by an "
         "industry, before deducting the cost of inputs."),
        ("SUT",
         "Supply and Use Tables. A matrix framework in the SNA showing what "
         "products are produced (supply) and how they are used (use) by industry."),
        ("SNA",
         "System of National Accounts. The international statistical standard "
         "for measuring economic activity (UN SNA 2008)."),
        ("OESA",
         "Ocean Economy Satellite Account. A satellite account that measures "
         "the economic contribution of ocean-related industries using SNA methods."),
        ("OTSA",
         "Ocean Tourism Satellite Account. A satellite account specifically "
         "measuring the economic contribution of ocean-related tourism."),
        ("Partial",
         "The share (0 to 1) of a national industry's output that is "
         "attributable to the ocean economy. Used to split mixed industries."),
        ("ISIC",
         "International Standard Industrial Classification of All Economic "
         "Activities. The UN system for classifying industries (Rev. 4)."),
        ("RAG",
         "Red-Amber-Green rating system. Used to indicate confidence levels: "
         "Green = strong data; Amber = some uncertainty; Red = rough estimate."),
        ("EEZ",
         "Exclusive Economic Zone. The area of ocean extending 200 nautical "
         "miles from a country's coastline, within which it has sovereign rights "
         "over natural resources."),
        ("SEEA",
         "System of Environmental-Economic Accounting. The international standard "
         "for environmental accounts, complementing the SNA. The OESA draws on "
         "both SNA and SEEA frameworks."),
        ("SIDS",
         "Small Island Developing States. A group of developing countries that "
         "share common vulnerabilities including small size, remoteness, and "
         "dependence on ocean resources."),
        ("NCP",
         "Nature's Contributions to People. The contributions of ecosystems "
         "to human wellbeing, also known as ecosystem services."),
    ]

    for term, defn in terms:
        teal_label(ws, r, 1, term)
        c = ws.cell(row=r, column=2, value=defn)
        style_cell(c, font=font_normal, fill=fill_white, alignment=align_left,
                   border=border_all)
        r += 1

    return ws


# ===================================================================
# MAIN
# ===================================================================
def main():
    wb = Workbook()

    build_readme(wb)
    build_step1(wb)
    build_step2(wb)
    build_step3(wb)
    build_step4(wb)
    build_step5(wb)
    build_answers(wb)
    build_glossary(wb)

    out_dir = os.path.dirname(os.path.abspath(__file__))
    out_path = os.path.join(out_dir, "oesa_exercise.xlsx")
    wb.save(out_path)
    print(f"Saved: {out_path}")
    print(f"Sheets: {wb.sheetnames}")


if __name__ == "__main__":
    main()
