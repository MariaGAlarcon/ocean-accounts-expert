#!/usr/bin/env python3
"""Generate GOAP-styled example account tables: filled Excel + HTML slides."""

from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---- GOAP Styles ----
GREEN = "3B9C7B"
TEAL = "0A5455"
GRAY = "404040"
TEXT = "30302F"
WHITE = "FFFFFF"
YELLOW = "FFF2CC"
MINT = "D4EEE5"
LIGHT_GRAY = "F2F2F2"

green_fill = PatternFill(start_color=GREEN, end_color=GREEN, fill_type="solid")
teal_fill = PatternFill(start_color=TEAL, end_color=TEAL, fill_type="solid")
mint_fill = PatternFill(start_color=MINT, end_color=MINT, fill_type="solid")
yellow_fill = PatternFill(start_color=YELLOW, end_color=YELLOW, fill_type="solid")
bdr = Border(left=Side("thin", GRAY), right=Side("thin", GRAY),
             top=Side("thin", GRAY), bottom=Side("thin", GRAY))
hdr_font = Font(name="Arial", size=10, bold=True, color=WHITE)
data_font = Font(name="Arial", size=10, color=TEXT)
bold_font = Font(name="Arial", size=10, bold=True, color=TEXT)
label_font = Font(name="Arial", size=10, bold=True, color=TEAL)
title_font = Font(name="Arial", size=14, bold=True, color=TEAL)
subtitle_font = Font(name="Arial", size=11, color=GREEN)
note_font = Font(name="Arial", size=9, italic=True, color="717171")


def apply_header(cell):
    cell.font = hdr_font; cell.fill = green_fill; cell.border = bdr
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def apply_row_header(cell):
    cell.font = Font(name="Arial", size=10, bold=True, color=WHITE)
    cell.fill = teal_fill; cell.border = bdr
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

def apply_data(cell, align="right", bold=False):
    cell.font = bold_font if bold else data_font
    cell.border = bdr
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.number_format = "#,##0.00" if isinstance(cell.value, float) else "#,##0"

def apply_data_pct(cell):
    cell.font = data_font; cell.border = bdr
    cell.alignment = Alignment(horizontal="right", vertical="center")
    cell.number_format = "0.0%"

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def section_title(ws, row, col, text):
    ws.cell(row=row, column=col, value=text).font = Font(
        name="Arial", size=12, bold=True, color=TEAL)


def generate_example_workbook(outdir):
    wb = openpyxl.Workbook()

    # ================================================================
    # SHEET 1: Extent Account
    # ================================================================
    ws = wb.active
    ws.title = "Extent Account"
    ws.sheet_properties.tabColor = GREEN

    ws["A1"] = "Table 1: Ecosystem Extent Account"
    ws["A1"].font = title_font
    ws["A2"] = "Coastal District Example, 2018 to 2023"
    ws["A2"].font = subtitle_font

    headers = ["", "Coral reefs\n(M1.3)", "Seagrass\n(M1.1)", "Mangroves\n(MFT1.2)",
               "Other\n(sand, rubble,\ndeep water)", "Total\naccounting\narea"]
    for c, h in enumerate(headers, 1):
        apply_header(ws.cell(row=4, column=c, value=h))

    data = [
        ("Opening extent (ha)", [1240, 860, 315, 7585, 10000]),
        ("Additions: natural expansion", [30, 20, 35, 85, 170]),
        ("Additions: managed restoration", [10, 15, 5, 0, 30]),
        ("Additions: reclassification", [0, 0, 0, 70, 70]),
        ("Total additions", [40, 35, 40, 155, 270]),
        ("Reductions: natural reduction", [50, 40, 10, 0, 100]),
        ("Reductions: managed conversion", [15, 10, 0, 0, 25]),
        ("Reductions: reclassification", [35, 20, 5, 85, 145]),
        ("Total reductions", [100, 70, 15, 85, 270]),
        ("Net change", [-60, -35, 25, 70, 0]),
        ("Closing extent (ha)", [1180, 825, 340, 7655, 10000]),
    ]
    for r, (label, values) in enumerate(data, 5):
        cell = ws.cell(row=r, column=1, value=label)
        apply_row_header(cell)
        for c, v in enumerate(values, 2):
            cell = ws.cell(row=r, column=c, value=v)
            is_total = label.startswith("Total") or label.startswith("Net") or label.startswith("Opening") or label.startswith("Closing")
            apply_data(cell, bold=is_total)

    ws["A17"] = "Source: Satellite classification (Sentinel-2), 2018 and 2023 imagery"
    ws["A17"].font = note_font

    set_col_widths(ws, [30, 16, 16, 16, 18, 16])

    # ================================================================
    # SHEET 1b: Extent Time Series (multi-period)
    # ================================================================
    ws1b = wb.create_sheet("Extent Time Series")
    ws1b.sheet_properties.tabColor = GREEN

    ws1b["A1"] = "Table 1b: Ecosystem Extent Time Series"
    ws1b["A1"].font = title_font
    ws1b["A2"] = "Coastal District Example, 2010 to 2023 (multiple accounting periods)"
    ws1b["A2"].font = subtitle_font
    ws1b["A3"] = "Each period's closing extent becomes the next period's opening extent."
    ws1b["A3"].font = note_font

    ts_headers = ["Ecosystem type", "2010\n(ha)", "2014\n(ha)", "2018\n(ha)", "2023\n(ha)",
                  "Total change\n2010-2023\n(ha)", "Total change\n(%)"]
    for c, h in enumerate(ts_headers, 1):
        apply_header(ws1b.cell(row=5, column=c, value=h))

    ts_data = [
        ("Coral reefs (M1.3)",   [1420, 1340, 1240, 1180, -240, -16.9]),
        ("Seagrass (M1.1)",      [950, 910, 860, 825, -125, -13.2]),
        ("Mangroves (MFT1.2)",   [280, 290, 315, 340, 60, 21.4]),
        ("Other",                [7350, 7460, 7585, 7655, 305, 4.1]),
        ("Total accounting area", [10000, 10000, 10000, 10000, 0, 0.0]),
    ]
    for r, (eco, vals) in enumerate(ts_data, 6):
        apply_row_header(ws1b.cell(row=r, column=1, value=eco))
        for c, v in enumerate(vals, 2):
            cell = ws1b.cell(row=r, column=c, value=v)
            is_total = (eco == "Total accounting area")
            apply_data(cell, bold=is_total)
            if c <= 5:
                cell.number_format = "#,##0"
            elif c == 6:
                cell.number_format = "#,##0"
            else:
                cell.number_format = "+0.0%;-0.0%" if not is_total else "0.0%"
                cell.value = v / 100

    # Net change per period rows
    ws1b.cell(row=12, column=1).value = ""
    section_title(ws1b, 13, 1, "Net change per accounting period")
    nc_headers = ["Ecosystem type", "2010-2014\n(ha)", "2014-2018\n(ha)", "2018-2023\n(ha)", "Trend"]
    for c, h in enumerate(nc_headers, 1):
        apply_header(ws1b.cell(row=14, column=c, value=h))

    nc_data = [
        ("Coral reefs",  [-80, -100, -60, "Declining (slowing)"]),
        ("Seagrass",     [-40, -50, -35, "Declining (slowing)"]),
        ("Mangroves",    [10, 25, 25, "Increasing (stable rate)"]),
    ]
    for r, (eco, vals) in enumerate(nc_data, 15):
        apply_row_header(ws1b.cell(row=r, column=1, value=eco))
        for c, v in enumerate(vals, 2):
            cell = ws1b.cell(row=r, column=c, value=v)
            if isinstance(v, str):
                apply_data(cell, align="left")
            else:
                apply_data(cell)
                cell.number_format = "+#,##0;-#,##0"

    ws1b["A19"] = "Trend interpretation: compare magnitude of net change across periods."
    ws1b["A19"].font = note_font
    ws1b["A20"] = "Coral and seagrass losses are slowing, suggesting either reduced pressures or approaching a floor."
    ws1b["A20"].font = note_font
    ws1b["A21"] = "Mangrove gains are consistent, likely reflecting sustained restoration programmes."
    ws1b["A21"].font = note_font

    set_col_widths(ws1b, [26, 14, 14, 14, 14, 18, 16])

    # ================================================================
    # SHEET 2: Change Matrix
    # ================================================================
    ws2 = wb.create_sheet("Change Matrix")
    ws2.sheet_properties.tabColor = TEAL

    ws2["A1"] = "Table 2: Ecosystem Type Change Matrix (hectares)"
    ws2["A1"].font = title_font
    ws2["A2"] = "Coastal District Example, 2018 to 2023"
    ws2["A2"].font = subtitle_font

    labels = ["Coral reefs", "Seagrass", "Mangroves", "Other"]
    ch = ["Opening \\ Closing"] + labels + ["Opening\ntotal"]
    for c, h in enumerate(ch, 1):
        apply_header(ws2.cell(row=4, column=c, value=h))

    matrix = [
        ("Coral reefs",  [1140, 15, 0, 85, 1240]),
        ("Seagrass",     [10, 790, 5, 55, 860]),
        ("Mangroves",    [0, 0, 300, 15, 315]),
        ("Other",        [30, 20, 35, 7500, 7585]),
        ("Closing total", [1180, 825, 340, 7655, 10000]),
    ]
    for r, (label, values) in enumerate(matrix, 5):
        apply_row_header(ws2.cell(row=r, column=1, value=label))
        for c, v in enumerate(values, 2):
            cell = ws2.cell(row=r, column=c, value=v)
            apply_data(cell, bold=(label == "Closing total"))
            # Highlight diagonal (stable area)
            if r - 5 == c - 2 and r < 9:
                cell.fill = mint_fill

    ws2["A11"] = "Green cells = stable area (no change in ecosystem type)"
    ws2["A11"].font = note_font

    set_col_widths(ws2, [20, 16, 16, 16, 16, 16])

    # ================================================================
    # SHEET 3: Condition Account
    # ================================================================
    ws3 = wb.create_sheet("Condition Account")
    ws3.sheet_properties.tabColor = GREEN

    ws3["A1"] = "Table 3: Ecosystem Condition Account"
    ws3["A1"].font = title_font
    ws3["A2"] = "Coral Reefs (M1.3), Marine Reserve Example, 2019 to 2024"
    ws3["A2"].font = subtitle_font

    c_headers = ["Indicator", "Unit", "Direction", "Reference\nlevel",
                 "Opening\nvalue\n(2019)", "Opening\nCI", "Closing\nvalue\n(2024)",
                 "Closing\nCI", "Change\nin CI", "Interpretation"]
    for c, h in enumerate(c_headers, 1):
        apply_header(ws3.cell(row=4, column=c, value=h))

    indicators = [
        ("Live coral cover", "%", "Higher is better", 50, 31.13, 0.62, 26.63, 0.53, -0.09, "Declining"),
        ("Fleshy macroalgae cover", "%", "Higher is worse", 60, 22.63, 0.62, 27.38, 0.54, -0.08, "Declining"),
        ("Fish biomass", "kg/ha", "Higher is better", 500, 213.75, 0.43, 193.75, 0.39, -0.04, "Declining"),
        ("Coral recruit density", "per m2", "Higher is better", 15, 8.50, 0.57, 6.20, 0.41, -0.15, "Declining"),
        ("Reef relief", "cm", "Higher is better", 150, 82.50, 0.55, 78.00, 0.52, -0.03, "Stable"),
        ("Bleaching prevalence", "% colonies", "Higher is worse", 100, 8.50, 0.92, 18.30, 0.82, -0.10, "Declining"),
    ]
    for r, (ind, unit, dirn, ref, ov, oci, cv, cci, ch, interp) in enumerate(indicators, 5):
        apply_row_header(ws3.cell(row=r, column=1, value=ind))
        for c, v in enumerate([unit, dirn, ref, ov, oci, cv, cci, ch, interp], 2):
            cell = ws3.cell(row=r, column=c, value=v)
            if isinstance(v, float) and c in (6, 8):
                apply_data(cell)
                cell.number_format = "0.00"
            elif isinstance(v, float) and c == 9:
                apply_data(cell)
                cell.number_format = "+0.00;-0.00;0.00"
            elif isinstance(v, (int, float)):
                apply_data(cell)
            else:
                apply_data(cell, align="left")

    # Composite row
    r_comp = 11
    apply_row_header(ws3.cell(row=r_comp, column=1, value="Composite Condition Index"))
    ws3.cell(row=r_comp, column=2, value="average").font = data_font
    ws3.cell(row=r_comp, column=2).border = bdr
    for c in range(3, 6):
        ws3.cell(row=r_comp, column=c).border = bdr
    cell = ws3.cell(row=r_comp, column=6, value=0.62)
    apply_data(cell, bold=True); cell.number_format = "0.00"
    ws3.cell(row=r_comp, column=7).border = bdr
    cell = ws3.cell(row=r_comp, column=8, value=0.53)
    apply_data(cell, bold=True); cell.number_format = "0.00"
    cell = ws3.cell(row=r_comp, column=9, value=-0.08)
    apply_data(cell, bold=True); cell.number_format = "+0.00;-0.00;0.00"
    cell = ws3.cell(row=r_comp, column=10, value="Declining")
    apply_data(cell, align="left", bold=True)

    ws3["A13"] = "CI scale: 0 = fully degraded, 1 = reference condition. Composite = mean of all indicator CIs."
    ws3["A13"].font = note_font
    ws3["A14"] = "Reference levels from regional literature (Indo-Pacific benchmarks) and McClanahan et al. 2011."
    ws3["A14"].font = note_font

    set_col_widths(ws3, [26, 12, 16, 12, 12, 10, 12, 10, 10, 14])

    # ================================================================
    # SHEET 4: Services -- Part A Physical Supply
    # ================================================================
    ws4 = wb.create_sheet("Services - Part A Physical")
    ws4.sheet_properties.tabColor = GREEN

    ws4["A1"] = "Table 4a: Ecosystem Service Physical Supply Account (Part A)"
    ws4["A1"].font = title_font
    ws4["A2"] = "Coastal District Example, 2023"
    ws4["A2"].font = subtitle_font

    s_headers = ["Service", "CICES\ncategory", "Unit", "Coral\nreefs", "Seagrass", "Mangroves", "Total"]
    for c, h in enumerate(s_headers, 1):
        apply_header(ws4.cell(row=4, column=c, value=h))

    phys = [
        ("Wild fish provisioning", "Provisioning", "kg/yr", 120000, 45000, 15000, 180000),
        ("Wood provisioning", "Provisioning", "tonnes/yr", None, None, 85, 85),
        ("Carbon sequestration", "Regulating", "Mg CO2/yr", None, 1040, 1295, 2335),
        ("Coastal protection", "Regulating", "m coastline", 12000, None, 3500, 15500),
        ("Nursery habitat", "Regulating", "kg biomass/yr", 4500, 2800, None, 7300),
        ("Sediment retention", "Regulating", "m3 CaCO3/yr", 8400, None, None, 8400),
        ("Recreation (reef diving)", "Cultural", "visitors/yr", 15000, None, None, 15000),
        ("Recreation (mangrove tours)", "Cultural", "trips/yr", None, None, 2400, 2400),
        ("Gleaning", "Cultural", "hours/yr", None, 18000, None, 18000),
        ("Gleaning harvest", "Cultural", "kg/yr", None, 4200, None, 4200),
    ]
    for r, (svc, cat, unit, cr, sg, mg, total) in enumerate(phys, 5):
        apply_row_header(ws4.cell(row=r, column=1, value=svc))
        apply_data(ws4.cell(row=r, column=2, value=cat), align="left")
        apply_data(ws4.cell(row=r, column=3, value=unit), align="left")
        for c, v in enumerate([cr, sg, mg, total], 4):
            cell = ws4.cell(row=r, column=c, value=v if v else "")
            apply_data(cell, bold=(c == 7))
            if v:
                cell.number_format = "#,##0"

    ws4["A16"] = "Part A is a complete account. Monetary valuation (Part B) is optional and builds on these quantities."
    ws4["A16"].font = note_font

    set_col_widths(ws4, [28, 14, 16, 14, 14, 14, 14])

    # ================================================================
    # SHEET 5: Services -- Part B Monetary Supply
    # ================================================================
    ws5 = wb.create_sheet("Services - Part B Monetary")
    ws5.sheet_properties.tabColor = TEAL

    ws5["A1"] = "Table 4b: Ecosystem Service Monetary Supply Account (Part B)"
    ws5["A1"].font = title_font
    ws5["A2"] = "Coastal District Example, 2023 (USD/yr)"
    ws5["A2"].font = subtitle_font

    m_headers = ["Service", "Valuation\nmethod", "Value\ntype", "Coral\nreefs", "Seagrass", "Mangroves", "Total"]
    for c, h in enumerate(m_headers, 1):
        apply_header(ws5.cell(row=4, column=c, value=h))

    monetary = [
        ("Wild fish provisioning", "Resource rent", "Market", 140000, 52500, 17500, 210000),
        ("Wood provisioning", "Substitute cost", "Non-market", None, None, 12750, 12750),
        ("Carbon sequestration", "Social cost of carbon", "Non-market", None, 53040, 66045, 119085),
        ("Coastal protection", "Replacement cost", "Non-market", 1200000, None, 175000, 1375000),
        ("Nursery habitat", "Productivity change", "Market (indirect)", 78750, 49000, None, 127750),
        ("Sediment retention", "Replacement cost", "Non-market", 420000, None, None, 420000),
        ("Recreation (reef)", "Direct expenditure", "Market", 1275000, None, None, 1275000),
        ("Recreation (mangrove)", "Direct expenditure", "Market", None, None, 84000, 84000),
        ("Gleaning", "Equivalent wage", "Mixed", None, 63000, None, 63000),
    ]
    for r, (svc, method, vtype, cr, sg, mg, total) in enumerate(monetary, 5):
        apply_row_header(ws5.cell(row=r, column=1, value=svc))
        apply_data(ws5.cell(row=r, column=2, value=method), align="left")
        apply_data(ws5.cell(row=r, column=3, value=vtype), align="left")
        for c, v in enumerate([cr, sg, mg, total], 4):
            cell = ws5.cell(row=r, column=c, value=v if v else "")
            apply_data(cell, bold=(c == 7))
            if v:
                cell.number_format = "#,##0"

    # Total row
    r_tot = 14
    apply_row_header(ws5.cell(row=r_tot, column=1, value="TOTAL"))
    ws5.cell(row=r_tot, column=2).border = bdr
    ws5.cell(row=r_tot, column=3).border = bdr
    totals = [
        sum(v for v in [140000, 1200000, 78750, 420000, 1275000] if v),  # coral
        sum(v for v in [52500, 53040, 49000, 63000] if v),  # seagrass
        sum(v for v in [17500, 12750, 66045, 175000, 84000] if v),  # mangrove
        3686585,  # total
    ]
    for c, v in enumerate(totals, 4):
        cell = ws5.cell(row=r_tot, column=c, value=v)
        apply_data(cell, bold=True)
        cell.number_format = "#,##0"

    ws5["A16"] = "Non-market values are estimated economic significance, not observed market prices."
    ws5["A16"].font = note_font
    ws5["A17"] = "Resource rent = gross revenue minus labour, fuel, gear, and capital costs."
    ws5["A17"].font = note_font
    ws5["A18"] = "SCC = USD 51/Mg CO2 (US EPA 2020 central estimate)."
    ws5["A18"].font = note_font

    set_col_widths(ws5, [28, 20, 16, 14, 14, 14, 14])

    # ================================================================
    # SHEET 6: Summary -- All Accounts
    # ================================================================
    ws6 = wb.create_sheet("Summary")
    ws6.sheet_properties.tabColor = GREEN

    ws6["A1"] = "Ocean Accounts Summary"
    ws6["A1"].font = title_font
    ws6["A2"] = "Coastal District Example"
    ws6["A2"].font = subtitle_font

    section_title(ws6, 4, 1, "Extent (Table 1)")
    for c, h in enumerate(["Ecosystem", "Opening (ha)", "Closing (ha)", "Change", "Change (%)"], 1):
        apply_header(ws6.cell(row=5, column=c, value=h))
    extent_sum = [
        ("Coral reefs", 1240, 1180, -60, -4.8),
        ("Seagrass", 860, 825, -35, -4.1),
        ("Mangroves", 315, 340, 25, 7.9),
    ]
    for r, (eco, o, cl, ch, pct) in enumerate(extent_sum, 6):
        apply_row_header(ws6.cell(row=r, column=1, value=eco))
        for c, v in enumerate([o, cl, ch], 2):
            cell = ws6.cell(row=r, column=c, value=v)
            apply_data(cell)
            cell.number_format = "#,##0"
        cell = ws6.cell(row=r, column=5, value=pct/100)
        apply_data(cell)
        cell.number_format = "+0.0%;-0.0%"

    section_title(ws6, 10, 1, "Condition (Table 3)")
    for c, h in enumerate(["Indicator", "Opening CI", "Closing CI", "Change", "Status"], 1):
        apply_header(ws6.cell(row=11, column=c, value=h))
    cond_sum = [
        ("Live coral cover", 0.62, 0.53, -0.09, "Declining"),
        ("Macroalgae cover", 0.62, 0.54, -0.08, "Declining"),
        ("Fish biomass", 0.43, 0.39, -0.04, "Declining"),
        ("Composite", 0.62, 0.53, -0.08, "Declining"),
    ]
    for r, (ind, o, cl, ch, st) in enumerate(cond_sum, 12):
        apply_row_header(ws6.cell(row=r, column=1, value=ind))
        for c, v in enumerate([o, cl], 2):
            cell = ws6.cell(row=r, column=c, value=v)
            apply_data(cell); cell.number_format = "0.00"
        cell = ws6.cell(row=r, column=4, value=ch)
        apply_data(cell); cell.number_format = "+0.00;-0.00"
        apply_data(ws6.cell(row=r, column=5, value=st), align="left")

    section_title(ws6, 18, 1, "Ecosystem Services (Tables 4a + 4b)")
    for c, h in enumerate(["Service", "Physical supply", "Monetary value\n(USD/yr)", "Method"], 1):
        apply_header(ws6.cell(row=19, column=c, value=h))
    svc_sum = [
        ("Fish provisioning", "180,000 kg/yr", 210000, "Resource rent"),
        ("Carbon sequestration", "2,335 Mg CO2/yr", 119085, "SCC"),
        ("Coastal protection", "15,500 m coastline", 1375000, "Replacement cost"),
        ("Nursery habitat", "7,300 kg biomass/yr", 127750, "Productivity change"),
        ("Sediment retention", "8,400 m3/yr", 420000, "Replacement cost"),
        ("Recreation", "17,400 visitors/yr", 1359000, "Direct expenditure"),
        ("Gleaning", "18,000 hours/yr", 63000, "Equivalent wage"),
        ("Wood provisioning", "85 tonnes/yr", 12750, "Substitute cost"),
        ("TOTAL", "", 3686585, ""),
    ]
    for r, (svc, phys, val, meth) in enumerate(svc_sum, 20):
        apply_row_header(ws6.cell(row=r, column=1, value=svc))
        apply_data(ws6.cell(row=r, column=2, value=phys), align="left")
        cell = ws6.cell(row=r, column=3, value=val)
        apply_data(cell, bold=(svc == "TOTAL"))
        cell.number_format = "#,##0"
        apply_data(ws6.cell(row=r, column=4, value=meth), align="left")

    set_col_widths(ws6, [26, 22, 18, 22, 14])

    # ================================================================
    # SHEET 7: Data Sources Reference
    # ================================================================
    ws7 = wb.create_sheet("Data Sources")
    ws7.sheet_properties.tabColor = GREEN

    ws7["A1"] = "Data Sources for Ecosystem Service Calculations"
    ws7["A1"].font = title_font
    ws7["A2"] = "Where to find the data for each service (Part A physical and Part B monetary)"
    ws7["A2"].font = subtitle_font

    sources = [
        ("SERVICE", "DATA NEEDED", "PART", "SOURCE", "URL / REFERENCE"),
        ("Fish provisioning", "Total catch (kg/yr)", "A", "National fisheries agency landing records", "FAO: fao.org/fishery"),
        ("Fish provisioning", "Catch by species", "A", "Landing site surveys, fisher cooperatives", "FishBase: fishbase.org"),
        ("Fish provisioning", "Catch by country (Tier 1)", "A", "FAO FishStatJ or Sea Around Us", "seaaroundus.org"),
        ("Fish provisioning", "Market prices (USD/kg)", "B", "Local fish market surveys, national stats", ""),
        ("Fish provisioning", "Fishing costs", "B", "Fisher household surveys", "GOAP survey templates"),
        ("", "", "", "", ""),
        ("Carbon sequestration", "Ecosystem extent (ha)", "A", "Extent account, or global datasets", "globalmangrovewatch.org, allencoralatlas.org"),
        ("Carbon sequestration", "NCP rates (Mg CO2/ha/yr)", "A", "Published literature", "Alongi (2014), Fourqurean et al. (2012)"),
        ("Carbon sequestration", "IPCC default rates", "A", "IPCC Wetlands Supplement (2013)", "ipcc.ch"),
        ("Carbon sequestration", "Primary carbon data (Tier 2-3)", "A", "Field: soil cores, allometric equations", "National forest inventories"),
        ("Carbon sequestration", "Social cost of carbon", "B", "US EPA Interagency Working Group", "USD 51/Mg CO2 (2021 central)"),
        ("Carbon sequestration", "Voluntary market price", "B", "Ecosystem Marketplace", "ecosystemmarketplace.com"),
        ("Carbon sequestration", "VCS blue carbon credits", "B", "Verra project registry", "verra.org"),
        ("", "", "", "", ""),
        ("Coastal protection", "Ecosystem extent along coast", "A", "GIS overlay of extent map + coastline", "From extent account"),
        ("Coastal protection", "Buildings behind buffer", "A", "OpenStreetMap, national cadastres", "openstreetmap.org"),
        ("Coastal protection", "Population behind buffer", "A", "Census data, WorldPop", "worldpop.org"),
        ("Coastal protection", "Seawall cost (USD/m)", "B", "Engineering firms, govt infrastructure budgets", "Range: USD 2,000-15,000/m"),
        ("Coastal protection", "Wave attenuation data (Tier 3)", "B", "Wave models, buoy data", "NOAA WaveWatch III"),
        ("", "", "", "", ""),
        ("Nursery habitat", "Fish density (nursery vs non-nursery)", "A", "Field surveys (UVC, seine nets)", "National monitoring programs"),
        ("Nursery habitat", "Log Response Ratio (LRR)", "A", "Literature meta-analyses", "Coral: 31%, Seagrass: 13%"),
        ("Nursery habitat", "Total fish biomass", "A", "From fish provisioning account", ""),
        ("Nursery habitat", "Market price of fish", "B", "Same as fish provisioning", ""),
        ("", "", "", "", ""),
        ("Recreation", "Visitor arrivals", "A", "National tourism authority, UNWTO", "unwto.org"),
        ("Recreation", "Dive/snorkel activity", "A", "Dive centres, tour operators", "PADI statistics, local associations"),
        ("Recreation", "MPA visitor counts", "A", "Park management, entry permits", ""),
        ("Recreation", "Activity fees", "B", "Operator records, online pricing", "Direct observation"),
        ("Recreation", "Visitor expenditure surveys", "B", "Tourist surveys, accommodation records", ""),
        ("", "", "", "", ""),
        ("Gleaning", "Gleaner counts and effort", "A", "Household surveys, community monitoring", "Participatory methods"),
        ("Gleaning", "Harvest by species (kg)", "A", "Gleaner interviews, landing measurements", ""),
        ("Gleaning", "Local wage rate", "B", "National statistics, minimum wage", ""),
        ("Gleaning", "Market prices for harvest", "B", "Local market surveys", ""),
        ("", "", "", "", ""),
        ("Sediment retention", "CaCO3 production rates", "A", "Literature", "Perry et al. (2012)"),
        ("Sediment retention", "Beach nourishment cost", "B", "Coastal management agencies", "Range: USD 5-20/m3"),
    ]

    for c, h in enumerate(sources[0], 1):
        apply_header(ws7.cell(row=4, column=c, value=h))

    for r, row in enumerate(sources[1:], 5):
        svc, data, part, source, url = row
        if svc == "":
            continue
        ws7.cell(row=r, column=1, value=svc).font = bold_font
        ws7.cell(row=r, column=1).border = bdr
        ws7.cell(row=r, column=2, value=data).font = data_font
        ws7.cell(row=r, column=2).border = bdr
        ws7.cell(row=r, column=3, value=part).font = data_font
        ws7.cell(row=r, column=3).border = bdr
        ws7.cell(row=r, column=3).alignment = Alignment(horizontal="center")
        ws7.cell(row=r, column=4, value=source).font = data_font
        ws7.cell(row=r, column=4).border = bdr
        ws7.cell(row=r, column=5, value=url).font = Font(name="Arial", size=9, color="5EB593")
        ws7.cell(row=r, column=5).border = bdr

    # Global datasets section
    r_global = 5 + len(sources)
    section_title(ws7, r_global, 1, "Global Open Datasets (free, get started quickly)")
    r_global += 1
    for c, h in enumerate(["Dataset", "URL", "What it provides"], 1):
        apply_header(ws7.cell(row=r_global, column=c, value=h))
    globals_data = [
        ("Allen Coral Atlas", "allencoralatlas.org", "Coral reef and seagrass extent, benthic cover"),
        ("Global Mangrove Watch", "globalmangrovewatch.org", "Mangrove extent 1996-2020, canopy height, biomass"),
        ("Global Fishing Watch", "globalfishingwatch.org", "Fishing effort, vessel tracking"),
        ("Sea Around Us", "seaaroundus.org", "Reconstructed catch by country, EEZ, species"),
        ("FAO FishStatJ", "fao.org/fishery/statistics", "National fisheries production and trade"),
        ("UNWTO", "unwto.org", "International tourism arrivals and receipts"),
        ("WorldPop", "worldpop.org", "Gridded population estimates"),
        ("OpenStreetMap", "openstreetmap.org", "Building footprints, infrastructure"),
        ("NOAA Coral Reef Watch", "coralreefwatch.noaa.gov", "SST, DHW, bleaching alerts"),
        ("ESVD", "esvd.info", "Published valuations for value transfer"),
        ("Ecosystem Marketplace", "ecosystemmarketplace.com", "Carbon credit prices and market trends"),
    ]
    for r, (name, url, desc) in enumerate(globals_data, r_global + 1):
        apply_row_header(ws7.cell(row=r, column=1, value=name))
        ws7.cell(row=r, column=2, value=url).font = Font(name="Arial", size=9, color="5EB593")
        ws7.cell(row=r, column=2).border = bdr
        ws7.cell(row=r, column=3, value=desc).font = data_font
        ws7.cell(row=r, column=3).border = bdr

    set_col_widths(ws7, [22, 30, 6, 40, 36])

    path = outdir / "examples" / "example_completed_accounts.xlsx"
    path.parent.mkdir(exist_ok=True)
    wb.save(path)
    print(f"  Created: {path}")


def generate_example_html(outdir):
    """Generate HTML slide showing all example account tables."""
    html = """<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<title>Example Ocean Account Tables</title>
<style>
* { margin: 0; padding: 0; box-sizing: border-box; }
body { font-family: Arial, sans-serif; background: #f5f5f5; color: #30302F; }
.slide { width: 100%; max-width: 1400px; margin: 2rem auto; background: #fff;
         padding: 2rem 2.5rem; border-radius: 8px; box-shadow: 0 2px 12px rgba(0,0,0,0.08); }
h1 { font-size: 32px; color: #0A5455; border-bottom: 4px solid #0A5455;
     padding-bottom: 0.5rem; margin-bottom: 1.5rem; }
h2 { font-size: 20px; color: #0A5455; margin: 2rem 0 0.8rem 0; }
h3 { font-size: 16px; color: #3B9C7B; margin: 1.5rem 0 0.5rem 0; }
table { border-collapse: collapse; width: 100%; margin-bottom: 1rem; font-size: 13px; }
th { background: #3B9C7B; color: #fff; font-weight: bold; padding: 8px 10px;
     border: 1px solid #404040; text-align: center; }
td { padding: 6px 10px; border: 1px solid #404040; }
td:first-child { background: #0A5455; color: #fff; font-weight: bold; }
td.num { text-align: right; }
td.bold { font-weight: bold; }
td.stable { background: #D4EEE5; }
.note { font-size: 12px; color: #717171; font-style: italic; margin: 0.5rem 0 1rem 0; }
.highlight { background: linear-gradient(to right, #E8F3F5, #F9F5F0);
             border: 2px solid #66A3B5; border-radius: 6px; padding: 1rem; margin: 1rem 0; }
.highlight strong { color: #0A5455; }
.page-break { page-break-before: always; }
</style>
</head>
<body>

<div class="slide">
<h1>Example Ocean Account Tables</h1>
<p style="font-size:14px; color:#3B9C7B;">Coastal District Example -- Illustrative data for training purposes</p>

<h2>Table 1: Ecosystem Extent Account (2018 to 2023)</h2>
<table>
<tr><th></th><th>Coral reefs<br>(M1.3)</th><th>Seagrass<br>(M1.1)</th><th>Mangroves<br>(MFT1.2)</th><th>Other</th><th>Total</th></tr>
<tr><td>Opening extent (ha)</td><td class="num bold">1,240</td><td class="num bold">860</td><td class="num bold">315</td><td class="num bold">7,585</td><td class="num bold">10,000</td></tr>
<tr><td>Additions: natural expansion</td><td class="num">30</td><td class="num">20</td><td class="num">35</td><td class="num">85</td><td class="num">170</td></tr>
<tr><td>Additions: managed restoration</td><td class="num">10</td><td class="num">15</td><td class="num">5</td><td class="num">0</td><td class="num">30</td></tr>
<tr><td>Additions: reclassification</td><td class="num">0</td><td class="num">0</td><td class="num">0</td><td class="num">70</td><td class="num">70</td></tr>
<tr><td>Total additions</td><td class="num bold">40</td><td class="num bold">35</td><td class="num bold">40</td><td class="num bold">155</td><td class="num bold">270</td></tr>
<tr><td>Total reductions</td><td class="num bold">100</td><td class="num bold">70</td><td class="num bold">15</td><td class="num bold">85</td><td class="num bold">270</td></tr>
<tr><td>Net change</td><td class="num bold">-60</td><td class="num bold">-35</td><td class="num bold">+25</td><td class="num bold">+70</td><td class="num bold">0</td></tr>
<tr><td>Closing extent (ha)</td><td class="num bold">1,180</td><td class="num bold">825</td><td class="num bold">340</td><td class="num bold">7,655</td><td class="num bold">10,000</td></tr>
</table>
<p class="note">Total accounting area is conserved (10,000 ha). Net change across all types sums to zero.</p>

<h2>Table 2: Ecosystem Type Change Matrix (hectares)</h2>
<table>
<tr><th>Opening \\ Closing</th><th>Coral</th><th>Seagrass</th><th>Mangrove</th><th>Other</th><th>Opening total</th></tr>
<tr><td>Coral reefs</td><td class="num stable">1,140</td><td class="num">15</td><td class="num">0</td><td class="num">85</td><td class="num bold">1,240</td></tr>
<tr><td>Seagrass</td><td class="num">10</td><td class="num stable">790</td><td class="num">5</td><td class="num">55</td><td class="num bold">860</td></tr>
<tr><td>Mangroves</td><td class="num">0</td><td class="num">0</td><td class="num stable">300</td><td class="num">15</td><td class="num bold">315</td></tr>
<tr><td>Other</td><td class="num">30</td><td class="num">20</td><td class="num">35</td><td class="num stable">7,500</td><td class="num bold">7,585</td></tr>
<tr><td>Closing total</td><td class="num bold">1,180</td><td class="num bold">825</td><td class="num bold">340</td><td class="num bold">7,655</td><td class="num bold">10,000</td></tr>
</table>
<p class="note">Green cells = stable area (no change). Off-diagonal = transitions between ecosystem types.</p>

<h2>Table 1b: Ecosystem Extent Time Series (multiple accounting periods)</h2>
<p style="font-size:13px; color:#545353; margin-bottom:0.8rem;">Each period's closing becomes the next period's opening. This table builds up over time.</p>
<table>
<tr><th>Ecosystem type</th><th>2010 (ha)</th><th>2014 (ha)</th><th>2018 (ha)</th><th>2023 (ha)</th><th>Total change</th><th>Change (%)</th></tr>
<tr><td>Coral reefs (M1.3)</td><td class="num">1,420</td><td class="num">1,340</td><td class="num">1,240</td><td class="num">1,180</td><td class="num bold">-240</td><td class="num bold">-16.9%</td></tr>
<tr><td>Seagrass (M1.1)</td><td class="num">950</td><td class="num">910</td><td class="num">860</td><td class="num">825</td><td class="num bold">-125</td><td class="num bold">-13.2%</td></tr>
<tr><td>Mangroves (MFT1.2)</td><td class="num">280</td><td class="num">290</td><td class="num">315</td><td class="num">340</td><td class="num bold">+60</td><td class="num bold">+21.4%</td></tr>
<tr><td>Other</td><td class="num">7,350</td><td class="num">7,460</td><td class="num">7,585</td><td class="num">7,655</td><td class="num bold">+305</td><td class="num bold">+4.1%</td></tr>
<tr><td>Total</td><td class="num bold">10,000</td><td class="num bold">10,000</td><td class="num bold">10,000</td><td class="num bold">10,000</td><td class="num bold">0</td><td class="num bold">0%</td></tr>
</table>

<h3>Net change per accounting period</h3>
<table>
<tr><th>Ecosystem</th><th>2010-2014</th><th>2014-2018</th><th>2018-2023</th><th>Trend</th></tr>
<tr><td>Coral reefs</td><td class="num">-80</td><td class="num">-100</td><td class="num">-60</td><td>Declining (slowing)</td></tr>
<tr><td>Seagrass</td><td class="num">-40</td><td class="num">-50</td><td class="num">-35</td><td>Declining (slowing)</td></tr>
<tr><td>Mangroves</td><td class="num">+10</td><td class="num">+25</td><td class="num">+25</td><td>Increasing (stable rate)</td></tr>
</table>

<div class="highlight">
<strong>Reading the time series:</strong> Coral and seagrass losses are slowing across periods, suggesting either reduced pressures or approaching a degradation floor. Mangrove gains are consistent, likely reflecting sustained restoration. The time series is where the value of repeated accounting becomes visible: a single snapshot shows status, but the series shows trajectory.
</div>
</div>

<div class="slide page-break">
<h1>Example Ocean Account Tables (continued)</h1>

<h2>Table 3: Ecosystem Condition Account -- Coral Reefs (M1.3)</h2>
<p style="font-size:13px; color:#545353; margin-bottom:0.8rem;">Marine Reserve Example, 2019 to 2024. CI scale: 0 = degraded, 1 = reference condition.</p>
<table>
<tr><th>Indicator</th><th>Unit</th><th>Direction</th><th>Reference</th><th>Opening<br>value</th><th>Opening<br>CI</th><th>Closing<br>value</th><th>Closing<br>CI</th><th>Change</th><th>Status</th></tr>
<tr><td>Live coral cover</td><td class="num">%</td><td>Higher better</td><td class="num">50%</td><td class="num">31.1</td><td class="num">0.62</td><td class="num">26.6</td><td class="num">0.53</td><td class="num">-0.09</td><td>Declining</td></tr>
<tr><td>Macroalgae cover</td><td class="num">%</td><td>Higher worse</td><td class="num">60%</td><td class="num">22.6</td><td class="num">0.62</td><td class="num">27.4</td><td class="num">0.54</td><td class="num">-0.08</td><td>Declining</td></tr>
<tr><td>Fish biomass</td><td class="num">kg/ha</td><td>Higher better</td><td class="num">500</td><td class="num">213.8</td><td class="num">0.43</td><td class="num">193.8</td><td class="num">0.39</td><td class="num">-0.04</td><td>Declining</td></tr>
<tr><td>Recruit density</td><td class="num">per m2</td><td>Higher better</td><td class="num">15</td><td class="num">8.5</td><td class="num">0.57</td><td class="num">6.2</td><td class="num">0.41</td><td class="num">-0.15</td><td>Declining</td></tr>
<tr><td>Reef relief</td><td class="num">cm</td><td>Higher better</td><td class="num">150</td><td class="num">82.5</td><td class="num">0.55</td><td class="num">78.0</td><td class="num">0.52</td><td class="num">-0.03</td><td>Stable</td></tr>
<tr><td>Bleaching prevalence</td><td class="num">% colonies</td><td>Higher worse</td><td class="num">100%</td><td class="num">8.5</td><td class="num">0.92</td><td class="num">18.3</td><td class="num">0.82</td><td class="num">-0.10</td><td>Declining</td></tr>
<tr><td style="background:#19383B">Composite CI</td><td></td><td></td><td></td><td></td><td class="num bold">0.62</td><td></td><td class="num bold">0.53</td><td class="num bold">-0.08</td><td class="bold">Declining</td></tr>
</table>

<div class="highlight">
<strong>Reading this table:</strong> The composite condition index declined from 0.62 to 0.53 over 5 years. The reef is at roughly half its reference condition, with coral cover, fish biomass, and bleaching showing the largest declines. Reef relief (physical structure) remained relatively stable.
</div>
</div>

<div class="slide page-break">
<h1>Example Ocean Account Tables (continued)</h1>

<h2>Table 4a: Ecosystem Service Physical Supply (Part A)</h2>
<p style="font-size:13px; color:#545353; margin-bottom:0.8rem;">Part A is a complete account on its own. No monetary data needed.</p>
<table>
<tr><th>Service</th><th>Category</th><th>Unit</th><th>Coral reefs</th><th>Seagrass</th><th>Mangroves</th><th>Total</th></tr>
<tr><td>Fish provisioning</td><td>Provisioning</td><td>kg/yr</td><td class="num">120,000</td><td class="num">45,000</td><td class="num">15,000</td><td class="num bold">180,000</td></tr>
<tr><td>Wood provisioning</td><td>Provisioning</td><td>tonnes/yr</td><td class="num"></td><td class="num"></td><td class="num">85</td><td class="num bold">85</td></tr>
<tr><td>Carbon sequestration</td><td>Regulating</td><td>Mg CO2/yr</td><td class="num"></td><td class="num">1,040</td><td class="num">1,295</td><td class="num bold">2,335</td></tr>
<tr><td>Coastal protection</td><td>Regulating</td><td>m coastline</td><td class="num">12,000</td><td class="num"></td><td class="num">3,500</td><td class="num bold">15,500</td></tr>
<tr><td>Nursery habitat</td><td>Regulating</td><td>kg biomass/yr</td><td class="num">4,500</td><td class="num">2,800</td><td class="num"></td><td class="num bold">7,300</td></tr>
<tr><td>Sediment retention</td><td>Regulating</td><td>m3 CaCO3/yr</td><td class="num">8,400</td><td class="num"></td><td class="num"></td><td class="num bold">8,400</td></tr>
<tr><td>Recreation (reef)</td><td>Cultural</td><td>visitors/yr</td><td class="num">15,000</td><td class="num"></td><td class="num"></td><td class="num bold">15,000</td></tr>
<tr><td>Recreation (mangrove)</td><td>Cultural</td><td>trips/yr</td><td class="num"></td><td class="num"></td><td class="num">2,400</td><td class="num bold">2,400</td></tr>
<tr><td>Gleaning</td><td>Cultural</td><td>hours/yr</td><td class="num"></td><td class="num">18,000</td><td class="num"></td><td class="num bold">18,000</td></tr>
</table>

<h2>Table 4b: Ecosystem Service Monetary Supply (Part B) -- USD/yr</h2>
<table>
<tr><th>Service</th><th>Method</th><th>Type</th><th>Coral reefs</th><th>Seagrass</th><th>Mangroves</th><th>Total</th></tr>
<tr><td>Fish provisioning</td><td>Resource rent</td><td>Market</td><td class="num">140,000</td><td class="num">52,500</td><td class="num">17,500</td><td class="num bold">210,000</td></tr>
<tr><td>Wood provisioning</td><td>Substitute cost</td><td>Non-market</td><td class="num"></td><td class="num"></td><td class="num">12,750</td><td class="num bold">12,750</td></tr>
<tr><td>Carbon sequestration</td><td>SCC ($51/Mg)</td><td>Non-market</td><td class="num"></td><td class="num">53,040</td><td class="num">66,045</td><td class="num bold">119,085</td></tr>
<tr><td>Coastal protection</td><td>Replacement cost</td><td>Non-market</td><td class="num">1,200,000</td><td class="num"></td><td class="num">175,000</td><td class="num bold">1,375,000</td></tr>
<tr><td>Nursery habitat</td><td>Productivity change</td><td>Market</td><td class="num">78,750</td><td class="num">49,000</td><td class="num"></td><td class="num bold">127,750</td></tr>
<tr><td>Sediment retention</td><td>Replacement cost</td><td>Non-market</td><td class="num">420,000</td><td class="num"></td><td class="num"></td><td class="num bold">420,000</td></tr>
<tr><td>Recreation</td><td>Direct expenditure</td><td>Market</td><td class="num">1,275,000</td><td class="num"></td><td class="num">84,000</td><td class="num bold">1,359,000</td></tr>
<tr><td>Gleaning</td><td>Equiv. wage + market</td><td>Mixed</td><td class="num"></td><td class="num">63,000</td><td class="num"></td><td class="num bold">63,000</td></tr>
<tr><td style="background:#19383B">TOTAL</td><td></td><td></td><td class="num bold">3,113,750</td><td class="num bold">217,540</td><td class="num bold">355,295</td><td class="num bold">3,686,585</td></tr>
</table>

<div class="highlight">
<strong>Key insight:</strong> Regulating and cultural services (coastal protection $1.4M, recreation $1.4M) account for 75% of total ecosystem service value. Fish provisioning, often assumed to be the dominant service, represents only 6% of the total. This pattern is common in tropical coastal systems and has direct implications for policy: protecting reef structure and water quality protects the highest-value services.
</div>
</div>

</body>
</html>"""

    path = outdir / "examples" / "slide_example_account_tables.html"
    path.parent.mkdir(exist_ok=True)
    with open(path, "w") as f:
        f.write(html)
    print(f"  Created: {path}")


if __name__ == "__main__":
    outdir = Path(__file__).parent
    print("Generating example account tables...")
    generate_example_workbook(outdir)
    generate_example_html(outdir)
    print("Done!")
