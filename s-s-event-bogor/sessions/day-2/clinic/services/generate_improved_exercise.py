#!/usr/bin/env python3
"""
Generate an improved ecosystem services exercise workbook.
Walks fellows through the full process of building a services account.
"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── GOAP colour palette ──────────────────────────────────────────────
GREEN  = "3B9C7B"
TEAL   = "0A5455"
GRAY   = "404040"
TEXT   = "30302F"
WHITE  = "FFFFFF"
YELLOW = "FFF2CC"

# ── Reusable styles ──────────────────────────────────────────────────
font_title   = Font(name="Arial", size=14, bold=True, color=TEAL)
font_section = Font(name="Arial", size=12, bold=True, color=TEAL)
font_header  = Font(name="Arial", size=10, bold=True, color=WHITE)
font_row_hdr = Font(name="Arial", size=10, bold=True, color=WHITE)
font_normal  = Font(name="Arial", size=10, color=TEXT)
font_bold    = Font(name="Arial", size=10, bold=True, color=TEXT)
font_italic  = Font(name="Arial", size=10, italic=True, color=GRAY)
font_answer  = Font(name="Arial", size=10, color=TEAL)

fill_green  = PatternFill("solid", fgColor=GREEN)
fill_teal   = PatternFill("solid", fgColor=TEAL)
fill_yellow = PatternFill("solid", fgColor=YELLOW)
fill_white  = PatternFill("solid", fgColor=WHITE)
fill_light  = PatternFill("solid", fgColor="E8F5F0")  # light green tint
fill_answer = PatternFill("solid", fgColor="D5F5E3")

align_wrap  = Alignment(wrap_text=True, vertical="top")
align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
align_left  = Alignment(horizontal="left", vertical="top", wrap_text=True)

thin_side = Side(style="thin", color=GRAY)
border_all = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)


# ── Helper functions ─────────────────────────────────────────────────
def style_cell(cell, font=font_normal, fill=None, alignment=align_wrap,
               border=None, number_format=None):
    cell.font = font
    if fill:
        cell.fill = fill
    cell.alignment = alignment
    if border:
        cell.border = border
    if number_format:
        cell.number_format = number_format


def write_title(ws, row, col, text):
    c = ws.cell(row=row, column=col, value=text)
    style_cell(c, font=font_title)
    return row + 1


def write_section(ws, row, col, text):
    c = ws.cell(row=row, column=col, value=text)
    style_cell(c, font=font_section)
    return row + 1


def write_text(ws, row, col, text, font_=font_normal, fill_=None):
    c = ws.cell(row=row, column=col, value=text)
    style_cell(c, font=font_, fill=fill_)
    return row + 1


def header_row(ws, row, headers, col_start=1):
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=col_start + i, value=h)
        style_cell(c, font=font_header, fill=fill_green, alignment=align_center,
                    border=border_all)


def teal_label(ws, row, col, text):
    c = ws.cell(row=row, column=col, value=text)
    style_cell(c, font=font_row_hdr, fill=fill_teal, alignment=align_wrap,
                border=border_all)


def data_cell(ws, row, col, value=None, is_input=False, is_formula=False,
              fmt=None, font_=font_normal):
    c = ws.cell(row=row, column=col, value=value)
    fill_ = fill_yellow if is_input else fill_white
    style_cell(c, font=font_, fill=fill_, alignment=align_center, border=border_all,
               number_format=fmt)
    return c


def set_col_widths(ws, widths):
    """widths is dict {col_letter: width}"""
    for letter, w in widths.items():
        ws.column_dimensions[letter].width = w


# =====================================================================
# SHEET 1: README
# =====================================================================
def build_readme(wb):
    ws = wb.active
    ws.title = "README"
    ws.sheet_properties.tabColor = TEAL

    set_col_widths(ws, {"A": 90})

    r = 1
    r = write_title(ws, r, 1, "Ecosystem Services Account Exercise")
    r += 1
    r = write_text(ws, r, 1,
        "This workbook walks you through the full process of building "
        "an ecosystem services supply account, following the SEEA EA framework.",
        font_=font_bold)
    r += 1

    steps = [
        ("Step 1 -- Identify your ecosystems and services",
         "Determine which services are supplied by which ecosystem types. "
         "This is the scoping step: you need to understand the relationship "
         "between ecosystem types and the services they provide before "
         "you can quantify anything."),
        ("Step 2 -- Build the Physical Supply Table (Part A)",
         "Quantify each service in physical units (kg, Mg CO2, metres, visitors). "
         "You will use extent data and rates (e.g., carbon sequestration = "
         "Extent (ha) x NCP rate (Mg CO2/ha/yr))."),
        ("Step 3 -- Select valuation methods for each service",
         "Choose the appropriate monetary valuation method for each service. "
         "Different services need different approaches: resource rent for "
         "marketed goods, social cost of carbon for climate regulation, "
         "replacement cost for protective services, etc."),
        ("Step 4 -- Calculate monetary values (Part B)",
         "Apply your chosen methods step by step to convert physical quantities "
         "into monetary values. Key formulas:\n"
         "  - Resource rent = Gross revenue - Total costs\n"
         "  - Carbon value = Physical supply (Mg CO2) x Social cost of carbon ($51/Mg)\n"
         "  - Replacement cost = Total replacement cost / Asset lifespan\n"
         "  - Direct expenditure = Visitors x Spend per visitor"),
        ("Step 5 -- Build the complete supply tables and interpret",
         "Assemble your results into formal SEEA EA supply tables (Part A "
         "and Part B) and interpret what they mean for management priorities."),
    ]

    for title, desc in steps:
        r = write_text(ws, r, 1, title, font_=font_bold)
        r = write_text(ws, r, 1, desc)
        r += 1

    r = write_text(ws, r, 1,
        "The Answers sheet contains a full answer key so you can check your work.",
        font_=font_italic)

    return ws


# =====================================================================
# SHEET 2: Step 1 - Identify Services
# =====================================================================
def build_step1(wb):
    ws = wb.create_sheet("Step 1 - Identify Services")
    ws.sheet_properties.tabColor = GREEN

    set_col_widths(ws, {"A": 28, "B": 14, "C": 16, "D": 16, "E": 16})

    r = 1
    r = write_title(ws, r, 1,
        "Step 1: Which services come from which ecosystems?")
    r += 1

    r = write_text(ws, r, 1,
        "Scenario: A coastal district has three ecosystem types:", font_=font_bold)
    r += 1

    eco_info = [
        ("Coral reefs (M1.3)", "1,200 ha"),
        ("Seagrass meadows (M1.1)", "800 ha"),
        ("Mangroves (MFT1.2)", "350 ha"),
    ]
    for name, area in eco_info:
        ws.cell(row=r, column=1, value=name).font = font_bold
        ws.cell(row=r, column=2, value=area).font = font_normal
        r += 1
    r += 1

    r = write_text(ws, r, 1,
        "Task: For each ecosystem-service combination below, enter Y if the "
        "ecosystem supplies that service, or N if it does not.", font_=font_italic)
    r += 1

    headers = ["Service", "SEEA EA ref", "Coral reefs\n(M1.3)",
               "Seagrass\n(M1.1)", "Mangroves\n(MFT1.2)"]
    header_row(ws, r, headers)
    r += 1

    services = [
        ("Fish provisioning", "6.1"),
        ("Carbon sequestration", "6.3"),
        ("Coastal protection", "6.4"),
        ("Nursery habitat", "6.2"),
        ("Recreation", "6.6"),
        ("Gleaning", "6.6"),
    ]

    for svc, ref in services:
        teal_label(ws, r, 1, svc)
        data_cell(ws, r, 2, ref)
        data_cell(ws, r, 3, is_input=True)  # coral
        data_cell(ws, r, 4, is_input=True)  # seagrass
        data_cell(ws, r, 5, is_input=True)  # mangroves
        r += 1

    return ws


# =====================================================================
# SHEET 3: Step 2 - Physical Supply
# =====================================================================
def build_step2(wb):
    ws = wb.create_sheet("Step 2 - Physical Supply")
    ws.sheet_properties.tabColor = GREEN

    set_col_widths(ws, {"A": 32, "B": 14, "C": 18, "D": 18, "E": 18, "F": 18})

    r = 1
    r = write_title(ws, r, 1,
        "Step 2: Build the Physical Supply Table (Part A)")
    r += 1
    r = write_text(ws, r, 1,
        "For each service, use the given data to calculate the physical supply "
        "by ecosystem type.", font_=font_italic)
    r += 1

    # ── Fish provisioning ────────────────────────────────────────────
    r = write_section(ws, r, 1, "A. Fish provisioning")
    r = write_text(ws, r, 1,
        "Given data: Reef fish catch = 120,000 kg/yr; "
        "Seagrass-associated catch = 45,000 kg/yr; "
        "Mangrove-associated catch = 15,000 kg/yr")
    r += 1
    headers = ["Ecosystem", "Catch (kg/yr)"]
    header_row(ws, r, headers)
    r += 1
    for eco, val in [("Coral reefs", None), ("Seagrass", None),
                      ("Mangroves", None), ("TOTAL", None)]:
        teal_label(ws, r, 1, eco)
        data_cell(ws, r, 2, is_input=True)
        r += 1
    r += 1

    # ── Carbon sequestration ─────────────────────────────────────────
    r = write_section(ws, r, 1, "B. Carbon sequestration")
    r = write_text(ws, r, 1,
        "Given: Mangrove extent = 350 ha, NCP rate = 3.7 Mg CO2/ha/yr. "
        "Seagrass extent = 800 ha, NCP rate = 1.3 Mg CO2/ha/yr. "
        "Coral reefs: negligible net sequestration.")
    r += 1
    r = write_text(ws, r, 1,
        "Formula: Physical supply = Extent (ha) x NCP rate (Mg CO2/ha/yr)",
        font_=font_italic)
    r += 1

    headers = ["Ecosystem", "Extent (ha)", "NCP rate\n(Mg CO2/ha/yr)",
               "Physical supply\n(Mg CO2/yr)"]
    header_row(ws, r, headers)
    r += 1

    # Mangroves row
    teal_label(ws, r, 1, "Mangroves")
    data_cell(ws, r, 2, 350)
    data_cell(ws, r, 3, 3.7)
    # formula: =B{r}*C{r}
    data_cell(ws, r, 4, is_input=True)
    man_carbon_row = r
    r += 1

    # Seagrass row
    teal_label(ws, r, 1, "Seagrass")
    data_cell(ws, r, 2, 800)
    data_cell(ws, r, 3, 1.3)
    data_cell(ws, r, 4, is_input=True)
    sea_carbon_row = r
    r += 1

    # Total
    teal_label(ws, r, 1, "TOTAL")
    data_cell(ws, r, 2, "")
    data_cell(ws, r, 3, "")
    data_cell(ws, r, 4, is_input=True)
    r += 2

    # ── Coastal protection ───────────────────────────────────────────
    r = write_section(ws, r, 1, "C. Coastal protection")
    r = write_text(ws, r, 1,
        "Given: 12,000 m of coastline fronted by coral reefs; "
        "3,500 m fronted by mangroves.")
    r += 1

    headers = ["Ecosystem", "Coastline\nprotected (m)"]
    header_row(ws, r, headers)
    r += 1
    for eco in ["Coral reefs", "Mangroves", "TOTAL"]:
        teal_label(ws, r, 1, eco)
        data_cell(ws, r, 2, is_input=True)
        r += 1
    r += 1

    # ── Recreation ───────────────────────────────────────────────────
    r = write_section(ws, r, 1, "D. Recreation")
    r = write_text(ws, r, 1,
        "Given: 15,000 reef visitors/yr (diving/snorkelling); "
        "2,400 mangrove tour trips/yr.")
    r += 1

    headers = ["Ecosystem", "Visitors/trips\nper year"]
    header_row(ws, r, headers)
    r += 1
    for eco in ["Coral reefs", "Mangroves", "TOTAL"]:
        teal_label(ws, r, 1, eco)
        data_cell(ws, r, 2, is_input=True)
        r += 1
    r += 2

    # ── Complete Part A table ────────────────────────────────────────
    r = write_section(ws, r, 1,
        "Complete Physical Supply Table (Part A)")
    r = write_text(ws, r, 1,
        "Transfer all your answers above into this summary table.",
        font_=font_italic)
    r += 1

    headers = ["Service", "Unit", "Coral reefs", "Seagrass", "Mangroves", "Total"]
    header_row(ws, r, headers)
    r += 1

    summary_rows = [
        ("Fish provisioning", "kg/yr"),
        ("Carbon sequestration", "Mg CO2/yr"),
        ("Coastal protection", "m coastline"),
        ("Recreation", "visitors/yr"),
    ]
    for svc, unit in summary_rows:
        teal_label(ws, r, 1, svc)
        data_cell(ws, r, 2, unit, font_=font_normal)
        data_cell(ws, r, 3, is_input=True)
        data_cell(ws, r, 4, is_input=True)
        data_cell(ws, r, 5, is_input=True)
        data_cell(ws, r, 6, is_input=True)
        r += 1

    return ws


# =====================================================================
# SHEET 4: Step 3 - Valuation Methods
# =====================================================================
def build_step3(wb):
    ws = wb.create_sheet("Step 3 - Valuation Methods")
    ws.sheet_properties.tabColor = GREEN

    set_col_widths(ws, {"A": 26, "B": 22, "C": 24, "D": 26, "E": 16})

    r = 1
    r = write_title(ws, r, 1,
        "Step 3: Select the valuation method for each service")
    r += 1
    r = write_text(ws, r, 1,
        "Use the reference table below to choose an appropriate valuation "
        "method for each service. Fill in the yellow cells.",
        font_=font_italic)
    r += 1

    headers = ["Service", "Physical quantity\n(from Step 2)",
               "Valuation method", "Data needed", "Available?\n(Y/N)"]
    header_row(ws, r, headers)
    r += 1

    task_rows = [
        ("Fish provisioning", "180,000 kg/yr"),
        ("Carbon sequestration", "2,335 Mg CO2/yr"),
        ("Coastal protection", "15,500 m coastline"),
        ("Recreation", "17,400 visitors/yr"),
    ]
    for svc, qty in task_rows:
        teal_label(ws, r, 1, svc)
        data_cell(ws, r, 2, qty, font_=font_normal)
        data_cell(ws, r, 3, is_input=True)
        data_cell(ws, r, 4, is_input=True)
        data_cell(ws, r, 5, is_input=True)
        r += 1

    r += 2

    # ── Reference table ──────────────────────────────────────────────
    r = write_section(ws, r, 1, "Reference: Valuation Methods")
    r += 1

    ref_headers = ["Method", "How it works", "When to use", "Example"]
    header_row(ws, r, ref_headers)
    r += 1

    ref_data = [
        ("Resource rent",
         "Revenue minus all costs\n(labour, fuel, gear, capital)",
         "Marketed goods with\ncost data available",
         "Fish: price x catch - costs"),
        ("Social cost of carbon\n(SCC)",
         "Physical quantity x SCC\nprice per Mg CO2",
         "Climate regulation\nservices",
         "Carbon: Mg CO2 x $51"),
        ("Replacement cost",
         "Cost to replace the service\nwith built infrastructure",
         "Protective services\n(coastal, flood)",
         "Seawall cost per metre\nannualized over lifespan"),
        ("Direct expenditure",
         "Total visitor spending\nattributable to the ecosystem",
         "Tourism and recreation",
         "Visitors x spend per visit"),
        ("Equivalent wage",
         "Hours of activity x local\nwage rate",
         "Unpaid / informal\nactivities (gleaning)",
         "Gleaning hours x\nlocal daily wage"),
        ("Value transfer",
         "Published $/ha/yr values\nfrom comparable studies",
         "When no local data is\navailable",
         "ESVD database values\napplied to local extent"),
    ]
    for method, how, when, example in ref_data:
        teal_label(ws, r, 1, method)
        data_cell(ws, r, 2, how, font_=font_normal)
        data_cell(ws, r, 3, when, font_=font_normal)
        data_cell(ws, r, 4, example, font_=font_normal)
        r += 1

    return ws


# =====================================================================
# SHEET 5: Step 4 - Monetary Values
# =====================================================================
def build_step4(wb):
    ws = wb.create_sheet("Step 4 - Monetary Values")
    ws.sheet_properties.tabColor = GREEN

    set_col_widths(ws, {
        "A": 8, "B": 32, "C": 18, "D": 20, "E": 20, "F": 18, "G": 18
    })

    r = 1
    r = write_title(ws, r, 1,
        "Step 4: Calculate monetary values (Part B)")
    r += 1
    r = write_text(ws, r, 1,
        "Work through each service calculation step by step. "
        "Yellow cells are for you to fill in (calculate).",
        font_=font_italic)
    r += 1

    # ── A. Fish provisioning ─────────────────────────────────────────
    r = write_section(ws, r, 1, "A. Fish provisioning -- Resource Rent method")
    r += 1

    fish_headers = ["Step", "Item", "Value (USD)"]
    header_row(ws, r, fish_headers, col_start=1)
    r += 1

    fish_steps = [
        ("1", "Total catch (kg/yr) -- from Step 2", 180000, False),
        ("2", "Average price (USD/kg)", 3.50, False),
        ("3", "Gross revenue = catch x price", None, True),
        ("4", "Labour costs", 180000, False),
        ("5", "Fuel costs", 120000, False),
        ("6", "Gear and maintenance", 80000, False),
        ("7", "Capital depreciation", 40000, False),
        ("8", "Total costs = sum of steps 4-7", None, True),
        ("9", "Resource rent = revenue - costs", None, True),
    ]
    for step, item, val, is_inp in fish_steps:
        c1 = ws.cell(row=r, column=1, value=step)
        style_cell(c1, font=font_bold, alignment=align_center, border=border_all)
        c2 = ws.cell(row=r, column=2, value=item)
        style_cell(c2, font=font_normal, alignment=align_wrap, border=border_all)
        if is_inp:
            data_cell(ws, r, 3, val, is_input=True, fmt="#,##0")
        else:
            data_cell(ws, r, 3, val, fmt="#,##0.00" if isinstance(val, float) else "#,##0")
        r += 1
    r += 2

    # ── B. Carbon sequestration ──────────────────────────────────────
    r = write_section(ws, r, 1, "B. Carbon sequestration -- Social Cost of Carbon")
    r = write_text(ws, r, 1, "SCC = $51 per Mg CO2 (US EPA estimate)", font_=font_italic)
    r += 1

    carb_headers = ["Step", "Ecosystem", "Extent (ha)",
                    "NCP rate\n(Mg CO2/ha/yr)",
                    "Physical supply\n(Mg CO2/yr)", "SCC ($/Mg)", "Value (USD)"]
    header_row(ws, r, carb_headers)
    r += 1

    carb_rows = [
        ("1", "Mangroves", 350, 3.7),
        ("2", "Seagrass", 800, 1.3),
    ]
    for step, eco, ext, rate in carb_rows:
        ws.cell(row=r, column=1, value=step)
        style_cell(ws.cell(row=r, column=1), font=font_bold, alignment=align_center,
                    border=border_all)
        teal_label(ws, r, 2, eco)
        data_cell(ws, r, 3, ext, fmt="#,##0")
        data_cell(ws, r, 4, rate, fmt="0.0")
        data_cell(ws, r, 5, is_input=True, fmt="#,##0.0")   # phys supply
        data_cell(ws, r, 6, 51, fmt="#,##0")
        data_cell(ws, r, 7, is_input=True, fmt="#,##0")     # value
        r += 1

    # total row
    ws.cell(row=r, column=1, value="3")
    style_cell(ws.cell(row=r, column=1), font=font_bold, alignment=align_center,
                border=border_all)
    teal_label(ws, r, 2, "TOTAL")
    data_cell(ws, r, 3, "")
    data_cell(ws, r, 4, "")
    data_cell(ws, r, 5, is_input=True, fmt="#,##0.0")
    data_cell(ws, r, 6, "")
    data_cell(ws, r, 7, is_input=True, fmt="#,##0")
    r += 2

    # ── C. Coastal protection ────────────────────────────────────────
    r = write_section(ws, r, 1, "C. Coastal protection -- Replacement Cost method")
    r += 1

    coast_headers = ["Step", "Item", "Value"]
    header_row(ws, r, coast_headers)
    r += 1

    coast_steps = [
        ("1", "Reef coastline protected (m)", 12000, False, "#,##0"),
        ("2", "Mangrove coastline protected (m)", 3500, False, "#,##0"),
        ("3", "Total coastline (m)", None, True, "#,##0"),
        ("4", "Seawall cost per metre (USD)", 5000, False, "#,##0"),
        ("5", "Total replacement cost = total m x cost/m", None, True, "#,##0"),
        ("6", "Seawall lifespan (years)", 50, False, "#,##0"),
        ("7", "Annualized value = total cost / lifespan", None, True, "#,##0"),
    ]
    for step, item, val, is_inp, fmt in coast_steps:
        ws.cell(row=r, column=1, value=step)
        style_cell(ws.cell(row=r, column=1), font=font_bold, alignment=align_center,
                    border=border_all)
        ws.cell(row=r, column=2, value=item)
        style_cell(ws.cell(row=r, column=2), font=font_normal, alignment=align_wrap,
                    border=border_all)
        data_cell(ws, r, 3, val, is_input=is_inp, fmt=fmt)
        r += 1
    r += 2

    # ── D. Recreation ────────────────────────────────────────────────
    r = write_section(ws, r, 1, "D. Recreation -- Direct Expenditure method")
    r += 1

    rec_headers = ["Step", "Activity", "Visitors/yr", "Spend/visitor\n(USD)", "Value (USD)"]
    header_row(ws, r, rec_headers)
    r += 1

    rec_rows = [
        ("1", "Reef diving/snorkelling", 15000, 85),
        ("2", "Mangrove tours", 2400, 35),
    ]
    for step, act, vis, spend in rec_rows:
        ws.cell(row=r, column=1, value=step)
        style_cell(ws.cell(row=r, column=1), font=font_bold, alignment=align_center,
                    border=border_all)
        teal_label(ws, r, 2, act)
        data_cell(ws, r, 3, vis, fmt="#,##0")
        data_cell(ws, r, 4, spend, fmt="#,##0")
        data_cell(ws, r, 5, is_input=True, fmt="#,##0")
        r += 1

    # total
    ws.cell(row=r, column=1, value="3")
    style_cell(ws.cell(row=r, column=1), font=font_bold, alignment=align_center,
                border=border_all)
    teal_label(ws, r, 2, "Total recreation value")
    data_cell(ws, r, 3, "")
    data_cell(ws, r, 4, "")
    data_cell(ws, r, 5, is_input=True, fmt="#,##0")
    r += 1

    return ws


# =====================================================================
# SHEET 6: Step 5 - Final Tables
# =====================================================================
def build_step5(wb):
    ws = wb.create_sheet("Step 5 - Final Tables")
    ws.sheet_properties.tabColor = GREEN

    set_col_widths(ws, {
        "A": 26, "B": 16, "C": 18, "D": 18, "E": 18, "F": 18, "G": 20
    })

    r = 1
    r = write_title(ws, r, 1,
        "Step 5: Build the complete supply tables")
    r += 1

    # ── Physical Supply Table ────────────────────────────────────────
    r = write_section(ws, r, 1, "Physical Supply Table (Part A)")
    r = write_text(ws, r, 1,
        "Transfer your Step 2 answers into this formal SEEA EA table.",
        font_=font_italic)
    r += 1

    phys_headers = ["Service", "Unit", "Coral reefs\n(M1.3)",
                    "Seagrass\n(M1.1)", "Mangroves\n(MFT1.2)", "Total"]
    header_row(ws, r, phys_headers)
    r += 1

    phys_services = [
        ("Fish provisioning", "kg/yr"),
        ("Carbon sequestration", "Mg CO2/yr"),
        ("Coastal protection", "m coastline"),
        ("Recreation", "visitors/yr"),
    ]
    for svc, unit in phys_services:
        teal_label(ws, r, 1, svc)
        data_cell(ws, r, 2, unit, font_=font_normal)
        for col in range(3, 7):
            data_cell(ws, r, col, is_input=True)
        r += 1

    r += 2

    # ── Monetary Supply Table ────────────────────────────────────────
    r = write_section(ws, r, 1, "Monetary Supply Table (Part B)")
    r = write_text(ws, r, 1,
        "Transfer your Step 4 answers. Include the valuation method used.",
        font_=font_italic)
    r += 1

    mon_headers = ["Service", "Valuation method", "Coral reefs\n(M1.3)",
                   "Seagrass\n(M1.1)", "Mangroves\n(MFT1.2)", "Total (USD)"]
    header_row(ws, r, mon_headers)
    r += 1

    mon_services = [
        "Fish provisioning",
        "Carbon sequestration",
        "Coastal protection",
        "Recreation",
    ]
    for svc in mon_services:
        teal_label(ws, r, 1, svc)
        data_cell(ws, r, 2, is_input=True)
        for col in range(3, 7):
            data_cell(ws, r, col, is_input=True, fmt="#,##0")
        r += 1

    # Grand total row
    teal_label(ws, r, 1, "GRAND TOTAL")
    data_cell(ws, r, 2, "")
    for col in range(3, 7):
        data_cell(ws, r, col, is_input=True, fmt="#,##0")
    r += 2

    # ── Summary metrics ──────────────────────────────────────────────
    r = write_section(ws, r, 1, "Summary and Interpretation")
    r += 1

    metrics = [
        "Total ecosystem service value (USD/yr):",
        "Highest-value service:",
        "Percentage of total from regulating services\n(carbon + coastal protection):",
        "What does this tell you about\nmanagement priorities for this district?",
    ]
    for m in metrics:
        ws.cell(row=r, column=1, value=m)
        style_cell(ws.cell(row=r, column=1), font=font_bold, alignment=align_wrap,
                    border=border_all)
        data_cell(ws, r, 2, is_input=True)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
        r += 1

    return ws


# =====================================================================
# SHEET 7: Answers
# =====================================================================
def build_answers(wb):
    ws = wb.create_sheet("Answers")
    ws.sheet_properties.tabColor = TEAL

    set_col_widths(ws, {
        "A": 30, "B": 18, "C": 18, "D": 18, "E": 18, "F": 18, "G": 20
    })

    r = 1
    r = write_title(ws, r, 1, "Answer Key")
    r += 1
    r = write_text(ws, r, 1,
        "Check your work against the correct answers below.",
        font_=font_italic)
    r += 1

    # ── Step 1 answers ───────────────────────────────────────────────
    r = write_section(ws, r, 1, "Step 1: Service-Ecosystem Matrix")
    r += 1
    headers = ["Service", "Coral reefs", "Seagrass", "Mangroves"]
    header_row(ws, r, headers)
    r += 1

    s1_answers = [
        ("Fish provisioning", "Y", "Y", "Y"),
        ("Carbon sequestration", "N", "Y", "Y"),
        ("Coastal protection", "Y", "N", "Y"),
        ("Nursery habitat", "Y", "Y", "Y"),
        ("Recreation", "Y", "N", "Y"),
        ("Gleaning", "Y", "Y", "Y"),
    ]
    for svc, *vals in s1_answers:
        teal_label(ws, r, 1, svc)
        for i, v in enumerate(vals):
            c = ws.cell(row=r, column=2+i, value=v)
            style_cell(c, font=font_answer, fill=fill_answer,
                       alignment=align_center, border=border_all)
        r += 1
    r += 2

    # ── Step 2 answers ───────────────────────────────────────────────
    r = write_section(ws, r, 1, "Step 2: Physical Supply Table (Part A)")
    r += 1

    headers = ["Service", "Unit", "Coral reefs", "Seagrass", "Mangroves", "Total"]
    header_row(ws, r, headers)
    r += 1

    s2_answers = [
        ("Fish provisioning", "kg/yr", 120000, 45000, 15000, 180000),
        ("Carbon sequestration", "Mg CO2/yr", 0, 1040, 1295, 2335),
        ("Coastal protection", "m coastline", 12000, 0, 3500, 15500),
        ("Recreation", "visitors/yr", 15000, 0, 2400, 17400),
    ]
    for svc, unit, *vals in s2_answers:
        teal_label(ws, r, 1, svc)
        c = ws.cell(row=r, column=2, value=unit)
        style_cell(c, font=font_normal, alignment=align_center, border=border_all)
        for i, v in enumerate(vals):
            c = ws.cell(row=r, column=3+i, value=v)
            style_cell(c, font=font_answer, fill=fill_answer,
                       alignment=align_center, border=border_all,
                       number_format="#,##0")
        r += 1
    r += 2

    # ── Step 3 answers ───────────────────────────────────────────────
    r = write_section(ws, r, 1, "Step 3: Valuation Methods")
    r += 1

    headers = ["Service", "Valuation method", "Data needed"]
    header_row(ws, r, headers)
    r += 1

    s3_answers = [
        ("Fish provisioning", "Resource rent",
         "Catch, price, labour/fuel/gear/capital costs"),
        ("Carbon sequestration", "Social cost of carbon",
         "Physical supply (Mg CO2), SCC price ($51)"),
        ("Coastal protection", "Replacement cost",
         "Coastline length, seawall cost/m, lifespan"),
        ("Recreation", "Direct expenditure",
         "Visitor numbers, average spend per visitor"),
    ]
    for svc, method, data in s3_answers:
        teal_label(ws, r, 1, svc)
        for i, v in enumerate([method, data]):
            c = ws.cell(row=r, column=2+i, value=v)
            style_cell(c, font=font_answer, fill=fill_answer,
                       alignment=align_wrap, border=border_all)
        r += 1
    r += 2

    # ── Step 4 answers ───────────────────────────────────────────────
    r = write_section(ws, r, 1, "Step 4: Monetary Values")
    r += 1

    # Fish
    r = write_text(ws, r, 1, "A. Fish provisioning", font_=font_bold)
    fish_ans = [
        ("Gross revenue", "180,000 kg x $3.50", 630000),
        ("Total costs", "180,000 + 120,000 + 80,000 + 40,000", 420000),
        ("Resource rent", "630,000 - 420,000", 210000),
    ]
    for item, calc, val in fish_ans:
        ws.cell(row=r, column=1, value=item)
        style_cell(ws.cell(row=r, column=1), font=font_bold, border=border_all,
                    alignment=align_wrap)
        ws.cell(row=r, column=2, value=calc)
        style_cell(ws.cell(row=r, column=2), font=font_italic, border=border_all,
                    alignment=align_wrap)
        c = ws.cell(row=r, column=3, value=val)
        style_cell(c, font=font_answer, fill=fill_answer, alignment=align_center,
                    border=border_all, number_format="#,##0")
        r += 1
    r += 1

    # Carbon
    r = write_text(ws, r, 1, "B. Carbon sequestration", font_=font_bold)
    carb_ans = [
        ("Mangroves", "350 x 3.7 = 1,295 Mg CO2", "1,295 x $51", 66045),
        ("Seagrass", "800 x 1.3 = 1,040 Mg CO2", "1,040 x $51", 53040),
        ("Total", "2,335 Mg CO2", "", 119085),
    ]
    for eco, phys, calc, val in carb_ans:
        teal_label(ws, r, 1, eco)
        ws.cell(row=r, column=2, value=phys)
        style_cell(ws.cell(row=r, column=2), font=font_italic, border=border_all,
                    alignment=align_wrap)
        ws.cell(row=r, column=3, value=calc)
        style_cell(ws.cell(row=r, column=3), font=font_italic, border=border_all,
                    alignment=align_wrap)
        c = ws.cell(row=r, column=4, value=val)
        style_cell(c, font=font_answer, fill=fill_answer, alignment=align_center,
                    border=border_all, number_format="#,##0")
        r += 1
    r += 1

    # Coastal protection
    r = write_text(ws, r, 1, "C. Coastal protection", font_=font_bold)
    coast_ans = [
        ("Total coastline", "12,000 + 3,500", 15500),
        ("Total replacement cost", "15,500 x $5,000", 77500000),
        ("Annualized value", "77,500,000 / 50", 1550000),
    ]
    for item, calc, val in coast_ans:
        ws.cell(row=r, column=1, value=item)
        style_cell(ws.cell(row=r, column=1), font=font_bold, border=border_all,
                    alignment=align_wrap)
        ws.cell(row=r, column=2, value=calc)
        style_cell(ws.cell(row=r, column=2), font=font_italic, border=border_all,
                    alignment=align_wrap)
        c = ws.cell(row=r, column=3, value=val)
        style_cell(c, font=font_answer, fill=fill_answer, alignment=align_center,
                    border=border_all, number_format="#,##0")
        r += 1
    r += 1

    # Recreation
    r = write_text(ws, r, 1, "D. Recreation", font_=font_bold)
    rec_ans = [
        ("Reef diving/snorkelling", "15,000 x $85", 1275000),
        ("Mangrove tours", "2,400 x $35", 84000),
        ("Total", "", 1359000),
    ]
    for act, calc, val in rec_ans:
        ws.cell(row=r, column=1, value=act)
        style_cell(ws.cell(row=r, column=1), font=font_bold, border=border_all,
                    alignment=align_wrap)
        ws.cell(row=r, column=2, value=calc)
        style_cell(ws.cell(row=r, column=2), font=font_italic, border=border_all,
                    alignment=align_wrap)
        c = ws.cell(row=r, column=3, value=val)
        style_cell(c, font=font_answer, fill=fill_answer, alignment=align_center,
                    border=border_all, number_format="#,##0")
        r += 1
    r += 2

    # ── Step 5 answers ───────────────────────────────────────────────
    r = write_section(ws, r, 1, "Step 5: Complete Monetary Supply Table (Part B)")
    r += 1

    headers = ["Service", "Method", "Coral reefs", "Seagrass", "Mangroves", "Total (USD)"]
    header_row(ws, r, headers)
    r += 1

    # Allocations for Part B
    # Fish: we allocate resource rent proportionally to catch
    # 120/180 = 66.7%, 45/180 = 25%, 15/180 = 8.3%
    fish_total = 210000
    fish_coral = round(fish_total * 120000 / 180000)
    fish_sea   = round(fish_total * 45000 / 180000)
    fish_man   = fish_total - fish_coral - fish_sea

    s5_answers = [
        ("Fish provisioning", "Resource rent",
         fish_coral, fish_sea, fish_man, fish_total),
        ("Carbon sequestration", "Social cost of carbon",
         0, 53040, 66045, 119085),
        ("Coastal protection", "Replacement cost",
         # allocate by coastline: 12000/15500 and 3500/15500
         round(1550000*12000/15500), 0, round(1550000*3500/15500), 1550000),
        ("Recreation", "Direct expenditure",
         1275000, 0, 84000, 1359000),
    ]
    grand_coral = 0
    grand_sea = 0
    grand_man = 0
    grand_total = 0
    for svc, method, coral, sea, man, total in s5_answers:
        teal_label(ws, r, 1, svc)
        c = ws.cell(row=r, column=2, value=method)
        style_cell(c, font=font_normal, alignment=align_wrap, border=border_all)
        for i, v in enumerate([coral, sea, man, total]):
            c = ws.cell(row=r, column=3+i, value=v)
            style_cell(c, font=font_answer, fill=fill_answer,
                       alignment=align_center, border=border_all,
                       number_format="#,##0")
        grand_coral += coral
        grand_sea += sea
        grand_man += man
        grand_total += total
        r += 1

    # Grand total
    teal_label(ws, r, 1, "GRAND TOTAL")
    ws.cell(row=r, column=2, value="")
    style_cell(ws.cell(row=r, column=2), border=border_all)
    for i, v in enumerate([grand_coral, grand_sea, grand_man, grand_total]):
        c = ws.cell(row=r, column=3+i, value=v)
        style_cell(c, font=Font(name="Arial", size=10, bold=True, color=TEAL),
                   fill=fill_answer, alignment=align_center, border=border_all,
                   number_format="#,##0")
    r += 2

    # Summary
    r = write_section(ws, r, 1, "Summary Metrics")
    r += 1

    reg_total = 119085 + 1550000  # carbon + coastal
    pct_reg = round(reg_total / grand_total * 100, 1)

    summary_ans = [
        ("Total ecosystem service value (USD/yr)", f"${grand_total:,}"),
        ("Highest-value service", "Coastal protection ($1,550,000/yr)"),
        ("Percentage from regulating services",
         f"{pct_reg}% (${reg_total:,} out of ${grand_total:,})"),
        ("Management implication",
         "Coastal protection dominates the account, highlighting that "
         "maintaining reef and mangrove extent is critical. Degradation "
         "would require very expensive built infrastructure to replace."),
    ]
    for item, ans in summary_ans:
        ws.cell(row=r, column=1, value=item)
        style_cell(ws.cell(row=r, column=1), font=font_bold, border=border_all,
                    alignment=align_wrap)
        c = ws.cell(row=r, column=2, value=ans)
        style_cell(c, font=font_answer, fill=fill_answer, alignment=align_wrap,
                    border=border_all)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
        r += 1

    return ws


# =====================================================================
# MAIN
# =====================================================================
def main():
    wb = Workbook()

    build_readme(wb)
    build_step1(wb)
    build_step2(wb)
    build_step3(wb)
    build_step4(wb)
    build_step5(wb)
    build_answers(wb)

    out_dir = os.path.dirname(os.path.abspath(__file__))
    out_path = os.path.join(out_dir, "services_exercise.xlsx")
    wb.save(out_path)
    print(f"Saved: {out_path}")


if __name__ == "__main__":
    main()
