#!/usr/bin/env python3
"""Generate GOAP-styled OTSA Excel template and exercise."""

from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

GREEN = "3B9C7B"
TEAL = "0A5455"
GRAY = "404040"
TEXT = "30302F"
WHITE = "FFFFFF"
YELLOW = "FFF2CC"

green_fill = PatternFill(start_color=GREEN, end_color=GREEN, fill_type="solid")
teal_fill = PatternFill(start_color=TEAL, end_color=TEAL, fill_type="solid")
yellow_fill = PatternFill(start_color=YELLOW, end_color=YELLOW, fill_type="solid")
bdr = Border(left=Side("thin", GRAY), right=Side("thin", GRAY),
             top=Side("thin", GRAY), bottom=Side("thin", GRAY))
hdr_font = Font(name="Arial", size=10, bold=True, color=WHITE)
data_font = Font(name="Arial", size=10, color=TEXT)
label_font = Font(name="Arial", size=10, bold=True, color=TEAL)
title_font = Font(name="Arial", size=14, bold=True, color=TEAL)
subtitle_font = Font(name="Arial", size=11, color=GREEN)
input_font = Font(name="Arial", size=10, color=TEXT)
answer_font = Font(name="Arial", size=10, bold=True, color="006100")
answer_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")


def apply_header(cell):
    cell.font = hdr_font; cell.fill = green_fill; cell.border = bdr
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def apply_row_header(cell):
    cell.font = Font(name="Arial", size=10, bold=True, color=WHITE)
    cell.fill = teal_fill; cell.border = bdr
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

def apply_data(cell, align="right"):
    cell.font = data_font; cell.border = bdr
    cell.alignment = Alignment(horizontal=align, vertical="center")

def apply_input(cell):
    cell.fill = yellow_fill; cell.font = input_font; cell.border = bdr
    cell.alignment = Alignment(horizontal="right", vertical="center")

def apply_answer(cell):
    cell.font = answer_font; cell.fill = answer_fill; cell.border = bdr
    cell.alignment = Alignment(horizontal="right", vertical="center")

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def generate_otsa_template():
    wb = openpyxl.Workbook()

    # Sheet 1: OTSA Account
    ws = wb.active
    ws.title = "OTSA Account"
    ws.sheet_properties.tabColor = GREEN

    ws["A1"] = "Ocean Tourism Satellite Account"
    ws["A1"].font = title_font
    ws["A2"] = "Measuring the economic contribution of coastal and marine tourism"
    ws["A2"].font = subtitle_font

    ws["A4"] = "Country:"; ws["A4"].font = label_font
    ws["B4"].fill = yellow_fill; ws["B4"].border = bdr
    ws["A5"] = "Year:"; ws["A5"].font = label_font
    ws["B5"].fill = yellow_fill; ws["B5"].border = bdr

    # Tourism partial estimation
    ws["A7"] = "Step 1: Estimate Tourism Partial"
    ws["A7"].font = Font(name="Arial", size=12, bold=True, color=TEAL)

    for c, h in enumerate(["Method", "Value", "Source"], 1):
        apply_header(ws.cell(row=8, column=c, value=h))

    methods = [
        ("Visitor motivation survey (%)", "", "Tourism authority survey"),
        ("Geographic proxy (%)", "", "Coastal municipality accommodation share"),
        ("Activity-based proxy (%)", "", "Marine activity expenditure share"),
        ("Selected tourism partial", "", "Best estimate from above"),
    ]
    for r, (m, v, s) in enumerate(methods, 9):
        apply_row_header(ws.cell(row=r, column=1, value=m))
        apply_input(ws.cell(row=r, column=2))
        apply_data(ws.cell(row=r, column=3, value=s), align="left")

    # Main account table
    ws["A14"] = "Step 2: OTSA Indicators"
    ws["A14"].font = Font(name="Arial", size=12, bold=True, color=TEAL)

    for c, h in enumerate(["Indicator", "Total tourism\n(national)", "Tourism\npartial",
                           "Coastal/ocean\ntourism", "Unit"], 1):
        apply_header(ws.cell(row=15, column=c, value=h))

    indicators = [
        ("International arrivals", "", "persons/yr"),
        ("Domestic trips", "", "trips/yr"),
        ("Total tourism expenditure", "", "USD million"),
        ("Tourism direct GVA", "", "USD million"),
        ("Tourism employment", "", "persons"),
        ("National GDP", "", "USD million"),
        ("Coastal tourism % of GDP", "", "%"),
    ]
    for r, (ind, _, unit) in enumerate(indicators, 16):
        apply_row_header(ws.cell(row=r, column=1, value=ind))
        apply_input(ws.cell(row=r, column=2))  # Total tourism
        if r < 22:
            ws.cell(row=r, column=3).value = f"=$B$12"  # Link to selected partial
            apply_data(ws.cell(row=r, column=3))
            ws.cell(row=r, column=4).value = f"=B{r}*C{r}"  # Coastal = Total x Partial
        elif r == 22:  # GDP row
            apply_input(ws.cell(row=r, column=2))
            ws.cell(row=r, column=4).value = ""
        apply_data(ws.cell(row=r, column=4))
        apply_data(ws.cell(row=r, column=5, value=unit), align="left")

    # GDP % formula
    ws.cell(row=22, column=4).value = '=IF(B22=0,"",D19/B22*100)'

    set_col_widths(ws, [30, 18, 14, 18, 14])

    # Sheet 2: Tourism Sectors
    ws2 = wb.create_sheet("Tourism Sectors")
    ws2.sheet_properties.tabColor = TEAL

    ws2["A1"] = "Coastal Tourism by Sub-Sector"
    ws2["A1"].font = title_font

    for c, h in enumerate(["Sub-sector", "National output\n(USD M)", "Coastal partial",
                           "Coastal output\n(USD M)", "Employment\n(national)",
                           "Coastal\nemployment"], 1):
        apply_header(ws2.cell(row=3, column=c, value=h))

    sectors = ["Accommodation", "Food and beverage", "Recreation and entertainment",
               "Transport (coastal)", "Travel agencies and tour operators", "Other tourism services",
               "TOTAL"]
    for r, s in enumerate(sectors, 4):
        apply_row_header(ws2.cell(row=r, column=1, value=s))
        for c in range(2, 7):
            if s == "TOTAL":
                ws2.cell(row=r, column=c).value = f"=SUM({get_column_letter(c)}4:{get_column_letter(c)}9)"
                apply_data(ws2.cell(row=r, column=c))
                ws2.cell(row=r, column=c).font = Font(name="Arial", size=10, bold=True, color=TEXT)
            else:
                apply_input(ws2.cell(row=r, column=c))
        if s != "TOTAL":
            ws2.cell(row=r, column=4).value = f"=B{r}*C{r}"
            apply_data(ws2.cell(row=r, column=4))
            ws2.cell(row=r, column=6).value = f"=E{r}*C{r}"
            apply_data(ws2.cell(row=r, column=6))

    set_col_widths(ws2, [32, 16, 14, 16, 16, 14])

    # Sheet 3: Instructions
    ws3 = wb.create_sheet("Instructions")
    ws3["A1"] = "How to use this template"
    ws3["A1"].font = title_font
    instructions = [
        "1. Start with the OTSA Account sheet.",
        "2. Estimate your tourism partial using one of three methods (Step 1).",
        "3. Enter the selected partial in the 'Selected tourism partial' row.",
        "4. Enter your national tourism data in the 'Total tourism' column (Step 2).",
        "5. Coastal/ocean tourism values calculate automatically (Total x Partial).",
        "6. Use the Tourism Sectors sheet for a detailed breakdown by sub-sector.",
        "7. Yellow cells are for your input. White cells with formulas auto-calculate.",
        "",
        "Data sources:",
        "- National tourism authority or board for arrivals, expenditure",
        "- UNWTO Tourism Dashboard (unwto.org) for international arrivals",
        "- National statistical office for GDP, employment",
        "- Tourism Satellite Account (TSA) if your country has one",
        "- Visitor surveys for motivation-based partial estimation",
    ]
    for i, txt in enumerate(instructions, 3):
        ws3.cell(row=i, column=1, value=txt).font = data_font
    ws3.column_dimensions["A"].width = 70

    path = Path(__file__).parent / "otsa_template.xlsx"
    wb.save(path)
    print(f"  Created: {path}")


def generate_otsa_exercise():
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = "Exercise"
    ws.sheet_properties.tabColor = GREEN

    ws["A1"] = "Ocean Tourism Satellite Account Exercise"
    ws["A1"].font = title_font
    ws["A2"] = "Calculate the coastal tourism contribution for a coastal country"
    ws["A2"].font = subtitle_font

    # Given data
    ws["A4"] = "Given Data"
    ws["A4"].font = Font(name="Arial", size=12, bold=True, color=TEAL)
    given = [
        ("International arrivals", "500,000/yr"),
        ("Domestic tourism trips", "2,000,000/yr"),
        ("Total tourism expenditure", "USD 800 million/yr"),
        ("Tourism direct GVA", "USD 320 million/yr"),
        ("Tourism employment", "45,000"),
        ("National GDP", "USD 5,000 million"),
        ("", ""),
        ("From visitor survey:", ""),
        ("International visitors: coastal motivation", "35%"),
        ("Domestic visitors: coastal motivation", "20%"),
        ("International spend per trip vs domestic", "2.5x more"),
    ]
    for r, (lb, val) in enumerate(given, 5):
        if lb:
            ws.cell(row=r, column=1, value=lb).font = data_font
            ws.cell(row=r, column=1).border = bdr
            ws.cell(row=r, column=2, value=val).font = data_font
            ws.cell(row=r, column=2).border = bdr

    row = 18
    ws.cell(row=row, column=1, value="Part 1: Calculate coastal arrivals").font = Font(
        name="Arial", size=12, bold=True, color=TEAL)
    row += 1
    for lb in ["Coastal international arrivals (500,000 x 0.35)",
               "Coastal domestic trips (2,000,000 x 0.20)"]:
        ws.cell(row=row, column=1, value=lb).font = data_font
        ws.cell(row=row, column=1).border = bdr
        apply_input(ws.cell(row=row, column=2))
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Part 2: Calculate weighted expenditure partial").font = Font(
        name="Arial", size=12, bold=True, color=TEAL)
    row += 1
    for lb in ["International share of total spending (hint: 2.5x weight)",
               "Weighted partial for expenditure"]:
        ws.cell(row=row, column=1, value=lb).font = data_font
        ws.cell(row=row, column=1).border = bdr
        apply_input(ws.cell(row=row, column=2))
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Part 3-5: Calculate coastal indicators").font = Font(
        name="Arial", size=12, bold=True, color=TEAL)
    row += 1
    for lb in ["Coastal tourism expenditure (USD M)",
               "Coastal tourism GVA (USD M)",
               "Coastal tourism employment",
               "Coastal tourism % of national GDP"]:
        ws.cell(row=row, column=1, value=lb).font = data_font
        ws.cell(row=row, column=1).border = bdr
        apply_input(ws.cell(row=row, column=2))
        row += 1

    set_col_widths(ws, [48, 20])

    # Answer sheet
    ws_a = wb.create_sheet("Answers")
    ws_a.sheet_properties.tabColor = "006100"
    ws_a["A1"] = "Answer Key"
    ws_a["A1"].font = title_font

    answers = [
        ("Coastal international arrivals", "175,000"),
        ("Coastal domestic trips", "400,000"),
        ("", ""),
        ("International share of spending", "55.6%"),
        ("(500K x 2.5 = 1.25M weighted intl; 2M domestic; intl share = 1.25/3.25 = 38.5% of trips but 55.6% of spend)", ""),
        ("Weighted expenditure partial", "25.8%"),
        ("(0.556 x 0.35 + 0.444 x 0.20 = 0.194 + 0.089 = 0.258)", ""),
        ("", ""),
        ("Coastal tourism expenditure", "USD 206 million"),
        ("(800 x 0.258)", ""),
        ("Coastal tourism GVA", "USD 82.5 million"),
        ("(320 x 0.258)", ""),
        ("Coastal tourism employment", "11,610"),
        ("(45,000 x 0.258)", ""),
        ("Coastal tourism % of GDP", "1.65%"),
        ("(82.5 / 5,000 x 100)", ""),
    ]
    for r, (lb, val) in enumerate(answers, 3):
        ws_a.cell(row=r, column=1, value=lb).font = label_font if val else Font(
            name="Arial", size=9, italic=True, color="717171")
        ws_a.cell(row=r, column=1).border = bdr
        if val:
            cell = ws_a.cell(row=r, column=2, value=val)
            apply_answer(cell)

    set_col_widths(ws_a, [56, 20])

    path = Path(__file__).parent / "otsa_exercise.xlsx"
    wb.save(path)
    print(f"  Created: {path}")


if __name__ == "__main__":
    print("Generating OTSA Excel files...")
    generate_otsa_template()
    generate_otsa_exercise()
    print("Done!")
