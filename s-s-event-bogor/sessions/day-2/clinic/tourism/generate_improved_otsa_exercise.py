#!/usr/bin/env python3
"""Generate improved OTSA exercise workbook with step-by-step structure."""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# GOAP colour palette
# ---------------------------------------------------------------------------
GREEN = "3B9C7B"
TEAL = "0A5455"
GRAY = "404040"
TEXT = "30302F"
WHITE = "FFFFFF"
YELLOW = "FFF2CC"

# Reusable styles
fill_teal = PatternFill(start_color=TEAL, end_color=TEAL, fill_type="solid")
fill_green = PatternFill(start_color=GREEN, end_color=GREEN, fill_type="solid")
fill_gray = PatternFill(start_color=GRAY, end_color=GRAY, fill_type="solid")
fill_yellow = PatternFill(start_color=YELLOW, end_color=YELLOW, fill_type="solid")
fill_white = PatternFill(start_color=WHITE, end_color=WHITE, fill_type="solid")
fill_light_gray = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

font_title = Font(name="Calibri", size=16, bold=True, color=TEAL)
font_heading = Font(name="Calibri", size=13, bold=True, color=TEAL)
font_subheading = Font(name="Calibri", size=11, bold=True, color=GRAY)
font_body = Font(name="Calibri", size=11, color=TEXT)
font_bold = Font(name="Calibri", size=11, bold=True, color=TEXT)
font_white_bold = Font(name="Calibri", size=11, bold=True, color=WHITE)
font_white_title = Font(name="Calibri", size=14, bold=True, color=WHITE)
font_small = Font(name="Calibri", size=10, color=GRAY)
font_italic = Font(name="Calibri", size=11, italic=True, color=GRAY)

align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
align_wrap = Alignment(wrap_text=True, vertical="top")

thin_border = Border(
    left=Side(style="thin", color=GRAY),
    right=Side(style="thin", color=GRAY),
    top=Side(style="thin", color=GRAY),
    bottom=Side(style="thin", color=GRAY),
)


def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def header_row(ws, row, headers, fill=fill_teal, font=font_white_bold):
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = font
        c.fill = fill
        c.alignment = align_center
        c.border = thin_border


def data_cell(ws, row, col, value, yellow=False, bold=False, number_format=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font = font_bold if bold else font_body
    c.alignment = align_center
    c.border = thin_border
    if yellow:
        c.fill = fill_yellow
    if number_format:
        c.number_format = number_format
    return c


def title_banner(ws, row, text, col_span=6):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_span)
    c = ws.cell(row=row, column=1, value=text)
    c.font = font_white_title
    c.fill = fill_teal
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 36


def instruction_text(ws, row, col, text, col_span=None):
    if col_span:
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + col_span - 1)
    c = ws.cell(row=row, column=col, value=text)
    c.font = font_body
    c.alignment = align_wrap
    return c


# ===========================================================================
# BUILD WORKBOOK
# ===========================================================================
wb = Workbook()

# ---- Sheet 1: README -----------------------------------------------------
ws = wb.active
ws.title = "README"
set_col_widths(ws, [3, 60, 30])
ws.sheet_properties.tabColor = TEAL

title_banner(ws, 1, "  OTSA Exercise: Estimating the Coastal Tourism Economy", 3)

rows = [
    (3, "B", font_heading, "Overview"),
    (4, "B", font_body, "This workbook guides you through a 4-step process to build an Ocean Tourism Satellite Account (OTSA) from national tourism data."),
    (6, "B", font_heading, "The 4 Steps"),
    (7, "B", font_bold, "Step 1 - Define Boundary"),
    (8, "B", font_body, "   Identify which tourism activities and accommodation types are 'coastal/ocean' related."),
    (9, "B", font_bold, "Step 2 - Estimate Partial"),
    (10, "B", font_body, "   Calculate the share of national tourism attributable to coastal areas using three methods."),
    (11, "B", font_bold, "Step 3 - Calculate Indicators"),
    (12, "B", font_body, "   Apply the partial to national tourism data to derive coastal tourism indicators."),
    (13, "B", font_bold, "Step 4 - OTSA Table"),
    (14, "B", font_body, "   Compile results into the formal OTSA output table and interpret findings."),
    (16, "B", font_heading, "Instructions"),
    (17, "B", font_body, "- Yellow cells are for your answers."),
    (18, "B", font_body, "- Check your work against the 'Answers' sheet when finished."),
    (19, "B", font_body, "- The 'Glossary' sheet defines key terms."),
    (21, "B", font_heading, "Learning Objectives"),
    (22, "B", font_body, "1. Understand the boundary decisions behind an OTSA."),
    (23, "B", font_body, "2. Apply partial-accounting techniques to allocate national data to the coast."),
    (24, "B", font_body, "3. Produce OTSA-format output tables consistent with SEEA and TSA frameworks."),
]
for r, col_letter, fnt, val in rows:
    c = ws.cell(row=r, column=2 if col_letter == "B" else 1, value=val)
    c.font = fnt
    c.alignment = align_wrap

# ---- Sheet 2: Step 1 - Define Boundary -----------------------------------
ws = wb.create_sheet("Step 1 - Define Boundary")
ws.sheet_properties.tabColor = GREEN
set_col_widths(ws, [3, 40, 18, 40])

title_banner(ws, 1, "  Step 1: Define the Coastal Tourism Boundary", 4)

ws.cell(row=3, column=2, value="Task A: Which tourism ACTIVITIES are coastal/ocean?").font = font_heading
ws.cell(row=4, column=2, value="Mark Y (yes) or N (no) in the yellow column.").font = font_italic

activities = [
    "Beach visits", "Scuba diving", "Snorkeling", "Fishing charters",
    "Whale watching", "Coastal hiking", "Theme parks (inland)",
    "Business conferences (city centre)", "Mountain trekking", "Shopping malls (urban)"
]
header_row(ws, 6, ["", "Activity", "Coastal? (Y/N)", "Notes / justification"], fill=fill_green)
for i, act in enumerate(activities):
    r = 7 + i
    data_cell(ws, r, 2, act)
    data_cell(ws, r, 2, act).alignment = align_left
    data_cell(ws, r, 3, None, yellow=True)
    data_cell(ws, r, 4, None, yellow=True).alignment = align_left

r_start = 7 + len(activities) + 2
ws.cell(row=r_start, column=2, value="Task B: Which ACCOMMODATION types count as coastal?").font = font_heading
ws.cell(row=r_start + 1, column=2, value="Mark Y (yes) or N (no) in the yellow column.").font = font_italic

accom = [
    ("Hotels within 5 km of the coastline", ""),
    ("Inland resorts (>20 km from coast)", ""),
    ("Beach bungalows / homestays on coast", ""),
    ("Urban hotels in a coastal city", "Consider: is the city's economy ocean-linked?"),
    ("Mountain lodges", ""),
    ("Cruise ship berths", ""),
]
header_row(ws, r_start + 3, ["", "Accommodation Type", "Coastal? (Y/N)", "Hint"], fill=fill_green)
for i, (acc, hint) in enumerate(accom):
    r = r_start + 4 + i
    data_cell(ws, r, 2, acc).alignment = align_left
    data_cell(ws, r, 3, None, yellow=True)
    data_cell(ws, r, 4, hint).alignment = align_left
    ws.cell(row=r, column=4).font = font_small

r_end = r + 2
ws.cell(row=r_end, column=2, value="Reflection:").font = font_bold
ws.cell(row=r_end + 1, column=2, value="What challenges arise when defining the boundary between coastal and non-coastal tourism?").font = font_italic
data_cell(ws, r_end + 2, 2, None, yellow=True).alignment = align_wrap
ws.merge_cells(start_row=r_end + 2, start_column=2, end_row=r_end + 4, end_column=4)

# ---- Sheet 3: Step 2 - Estimate Partial ----------------------------------
ws = wb.create_sheet("Step 2 - Estimate Partial")
ws.sheet_properties.tabColor = GREEN
set_col_widths(ws, [3, 44, 22, 22, 22])

title_banner(ws, 1, "  Step 2: Estimate the Coastal Tourism Partial", 5)

ws.cell(row=3, column=2, value="You will estimate the share of national tourism that is coastal using three methods, then compute a weighted average.").font = font_body
ws.cell(row=3, column=2).alignment = align_wrap
ws.merge_cells("B3:E3")

# Method 1
ws.cell(row=5, column=2, value="Method 1: Visitor Motivation Survey").font = font_heading
ws.cell(row=6, column=2, value="Given data:").font = font_bold
ws.cell(row=7, column=2, value="   35% of international visitors report coastal/ocean as primary motivation").font = font_body
ws.cell(row=8, column=2, value="   20% of domestic visitors report coastal/ocean as primary motivation").font = font_body
ws.cell(row=9, column=2, value="   International visitors = 500,000  |  Domestic trips = 2,000,000").font = font_body

header_row(ws, 11, ["", "Visitor Type", "Total Visitors", "Coastal %", "Coastal Visitors"])
data_cell(ws, 12, 2, "International").alignment = align_left
data_cell(ws, 12, 3, 500000, number_format="#,##0")
data_cell(ws, 12, 4, "35%")
data_cell(ws, 12, 5, None, yellow=True)
data_cell(ws, 13, 2, "Domestic").alignment = align_left
data_cell(ws, 13, 3, 2000000, number_format="#,##0")
data_cell(ws, 13, 4, "20%")
data_cell(ws, 13, 5, None, yellow=True)
data_cell(ws, 14, 2, "Total", bold=True).alignment = align_left
data_cell(ws, 14, 3, 2500000, bold=True, number_format="#,##0")
data_cell(ws, 14, 4, None)
data_cell(ws, 14, 5, None, yellow=True)

ws.cell(row=16, column=2, value="Method 1 partial (coastal visitors / total visitors):").font = font_bold
data_cell(ws, 16, 5, None, yellow=True)

# Method 2
ws.cell(row=18, column=2, value="Method 2: Geographic Proxy").font = font_heading
ws.cell(row=19, column=2, value="Given: 28% of national accommodation capacity (rooms) is in coastal municipalities.").font = font_body
ws.cell(row=20, column=2, value="Method 2 partial:").font = font_bold
data_cell(ws, 20, 5, None, yellow=True)

# Method 3
ws.cell(row=22, column=2, value="Method 3: Activity-Based").font = font_heading
ws.cell(row=23, column=2, value="Given: 22% of total visitor expenditure is on marine/coastal activities (diving, boat tours, beach entry fees, seafood dining, etc.)").font = font_body
ws.merge_cells("B23:E23")
ws.cell(row=23, column=2).alignment = align_wrap
ws.cell(row=24, column=2, value="Method 3 partial:").font = font_bold
data_cell(ws, 24, 5, None, yellow=True)

# Weighted calculation
ws.cell(row=26, column=2, value="Weighted Expenditure Partial").font = font_heading
ws.cell(row=27, column=2, value="Important: international visitors spend 2.5x more per trip than domestic visitors.").font = font_bold
ws.cell(row=27, column=2).alignment = align_wrap
ws.merge_cells("B27:E27")

ws.cell(row=29, column=2, value="Task: Calculate the expenditure-weighted coastal partial.").font = font_subheading

header_row(ws, 31, ["", "Item", "Value", "Your Calculation", "Result"])
calc_rows = [
    ("Domestic spend index", "1.0", "", ""),
    ("International spend index", "2.5", "", ""),
    ("Weighted international coastal spend", "", "500,000 x 0.35 x 2.5 = ?", ""),
    ("Weighted domestic coastal spend", "", "2,000,000 x 0.20 x 1.0 = ?", ""),
    ("Total weighted coastal spend", "", "Sum of above", ""),
    ("Total weighted national spend", "", "500,000 x 2.5 + 2,000,000 x 1.0 = ?", ""),
    ("Expenditure-weighted partial", "", "Coastal / National", ""),
]
for i, (item, val, calc, res) in enumerate(calc_rows):
    r = 32 + i
    data_cell(ws, r, 2, item).alignment = align_left
    data_cell(ws, r, 3, val)
    data_cell(ws, r, 4, calc, yellow=True).alignment = align_left
    data_cell(ws, r, 5, None, yellow=True)

ws.cell(row=40, column=2, value="Which method do you think is most reliable and why?").font = font_italic
data_cell(ws, 41, 2, None, yellow=True).alignment = align_wrap
ws.merge_cells(start_row=41, start_column=2, end_row=43, end_column=5)

# ---- Sheet 4: Step 3 - Calculate Indicators ------------------------------
ws = wb.create_sheet("Step 3 - Calculate Indicators")
ws.sheet_properties.tabColor = GREEN
set_col_widths(ws, [3, 38, 22, 18, 24])

title_banner(ws, 1, "  Step 3: Apply the Partial to National Tourism Data", 5)

ws.cell(row=3, column=2, value="Use the expenditure-weighted partial from Step 2 to estimate coastal tourism indicators.").font = font_body
ws.merge_cells("B3:E3")
ws.cell(row=4, column=2, value="Hint: the weighted partial should be approximately 25.8% (check Answers sheet if unsure).").font = font_italic
ws.merge_cells("B4:E4")

header_row(ws, 6, ["", "Indicator", "Total Tourism", "Partial", "Coastal Tourism"])
indicators = [
    ("International arrivals", "500,000", True, True),
    ("Domestic trips", "2,000,000", True, True),
    ("Total expenditure (USD million)", "800", True, True),
    ("Direct GVA (USD million)", "320", True, True),
    ("Employment (persons)", "45,000", True, True),
    ("National GDP (USD million)", "5,000", False, False),
    ("Coastal tourism as % of GDP", "", False, True),
]
for i, (ind, total, partial_yellow, coastal_yellow) in enumerate(indicators):
    r = 7 + i
    data_cell(ws, r, 2, ind, bold=(i >= 5)).alignment = align_left
    c = data_cell(ws, r, 3, total)
    if total and "million" not in ind.lower() or "million" in ind.lower():
        pass  # keep as string for display
    data_cell(ws, r, 4, None, yellow=partial_yellow)
    data_cell(ws, r, 5, None, yellow=coastal_yellow)

ws.cell(row=15, column=2, value="Note: For arrivals/trips, apply the visitor-count partial (Method 1). For expenditure/GVA/employment, apply the expenditure-weighted partial.").font = font_small
ws.merge_cells("B15:E15")
ws.cell(row=15, column=2).alignment = align_wrap

ws.cell(row=17, column=2, value="Reflection questions:").font = font_heading
questions = [
    "1. Is it appropriate to use the same partial for employment as for GVA? Why or why not?",
    "2. What data would you need to produce a more accurate employment figure?",
    "3. How does the coastal tourism contribution compare to other sectors in your country?",
]
for i, q in enumerate(questions):
    ws.cell(row=18 + i, column=2, value=q).font = font_body
    ws.merge_cells(start_row=18 + i, start_column=2, end_row=18 + i, end_column=5)

# ---- Sheet 5: Step 4 - OTSA Table ----------------------------------------
ws = wb.create_sheet("Step 4 - OTSA Table")
ws.sheet_properties.tabColor = GREEN
set_col_widths(ws, [3, 36, 22, 22, 22, 22])

title_banner(ws, 1, "  Step 4: Compile the OTSA Output Table", 6)

ws.cell(row=3, column=2, value="Fill in the formal OTSA summary table using your results from Steps 1-3.").font = font_body
ws.merge_cells("B3:F3")

header_row(ws, 5, ["", "OTSA Indicator", "National Total", "Coastal Share (%)", "Coastal Value", "Unit"], fill=fill_teal)
otsa_rows = [
    "Inbound tourism expenditure",
    "Domestic tourism expenditure",
    "Internal tourism consumption",
    "Tourism direct GVA",
    "Tourism direct GDP",
    "Tourism employment",
    "Number of trips (international)",
    "Number of trips (domestic)",
    "Average length of stay (coastal)",
    "Tourism as % of national GDP",
]
for i, label in enumerate(otsa_rows):
    r = 6 + i
    data_cell(ws, r, 2, label).alignment = align_left
    for col in range(3, 7):
        data_cell(ws, r, col, None, yellow=True)

r_int = 6 + len(otsa_rows) + 2
ws.cell(row=r_int, column=2, value="Interpretation Questions").font = font_heading
interp = [
    "1. What does the OTSA tell policy-makers that the national TSA does not?",
    "2. If coastal tourism GVA is growing faster than national tourism GVA, what might that imply for ocean policy?",
    "3. What are the main limitations of partial-based OTSA estimates?",
    "4. How could this OTSA be improved with better data?",
]
for i, q in enumerate(interp):
    ws.cell(row=r_int + 1 + i, column=2, value=q).font = font_body
    ws.merge_cells(start_row=r_int + 1 + i, start_column=2, end_row=r_int + 1 + i, end_column=6)

# ---- Sheet 6: Answers ----------------------------------------------------
ws = wb.create_sheet("Answers")
ws.sheet_properties.tabColor = GRAY
set_col_widths(ws, [3, 48, 22, 22, 22])

title_banner(ws, 1, "  Answers: Full Calculations", 5)

ws.cell(row=3, column=2, value="Step 1 Answers - Boundary").font = font_heading
answers_act = [
    ("Beach visits", "Y"), ("Scuba diving", "Y"), ("Snorkeling", "Y"),
    ("Fishing charters", "Y"), ("Whale watching", "Y"), ("Coastal hiking", "Y"),
    ("Theme parks (inland)", "N"), ("Business conferences (city centre)", "N"),
    ("Mountain trekking", "N"), ("Shopping malls (urban)", "N"),
]
header_row(ws, 5, ["", "Activity", "Coastal?"], fill=fill_gray)
for i, (act, yn) in enumerate(answers_act):
    r = 6 + i
    data_cell(ws, r, 2, act).alignment = align_left
    data_cell(ws, r, 3, yn)

answers_accom = [
    ("Hotels within 5 km of coastline", "Y"),
    ("Inland resorts (>20 km from coast)", "N"),
    ("Beach bungalows / homestays on coast", "Y"),
    ("Urban hotels in a coastal city", "Y (arguable)"),
    ("Mountain lodges", "N"),
    ("Cruise ship berths", "Y"),
]
r_a = 6 + len(answers_act) + 1
header_row(ws, r_a, ["", "Accommodation", "Coastal?"], fill=fill_gray)
for i, (acc, yn) in enumerate(answers_accom):
    r = r_a + 1 + i
    data_cell(ws, r, 2, acc).alignment = align_left
    data_cell(ws, r, 3, yn)

r_s2 = r + 2
ws.cell(row=r_s2, column=2, value="Step 2 Answers - Partial Estimation").font = font_heading

calcs = [
    ("Method 1: Visitor motivation", "", ""),
    ("  International coastal visitors", "500,000 x 0.35", "175,000"),
    ("  Domestic coastal visitors", "2,000,000 x 0.20", "400,000"),
    ("  Total coastal visitors", "175,000 + 400,000", "575,000"),
    ("  Method 1 partial", "575,000 / 2,500,000", "23.0%"),
    ("", "", ""),
    ("Method 2: Geographic proxy", "", "28.0%"),
    ("Method 3: Activity-based", "", "22.0%"),
    ("", "", ""),
    ("Expenditure-weighted partial", "", ""),
    ("  Weighted intl coastal spend", "500,000 x 0.35 x 2.5", "437,500"),
    ("  Weighted domestic coastal spend", "2,000,000 x 0.20 x 1.0", "400,000"),
    ("  Total weighted coastal", "437,500 + 400,000", "837,500"),
    ("  Total weighted national", "(500,000 x 2.5) + (2,000,000 x 1.0)", "3,250,000"),
    ("  Expenditure-weighted partial", "837,500 / 3,250,000", "25.77% ~ 25.8%"),
]
header_row(ws, r_s2 + 1, ["", "Item", "Calculation", "Result"], fill=fill_gray)
for i, (item, calc, res) in enumerate(calcs):
    r = r_s2 + 2 + i
    data_cell(ws, r, 2, item).alignment = align_left
    data_cell(ws, r, 3, calc).alignment = align_left
    data_cell(ws, r, 4, res)
    if item and not item.startswith(" ") and item.startswith("Method"):
        ws.cell(row=r, column=2).font = font_bold

r_s3 = r + 2
ws.cell(row=r_s3, column=2, value="Step 3 Answers - Indicators").font = font_heading

header_row(ws, r_s3 + 1, ["", "Indicator", "Total", "Partial", "Coastal Value"], fill=fill_gray)
ind_answers = [
    ("International arrivals", "500,000", "35.0%", "175,000"),
    ("Domestic trips", "2,000,000", "20.0%", "400,000"),
    ("Total expenditure (USD M)", "800", "25.8%", "206.2"),
    ("Direct GVA (USD M)", "320", "25.8%", "82.6"),
    ("Employment", "45,000", "25.8%", "11,610"),
    ("National GDP (USD M)", "5,000", "", ""),
    ("Coastal tourism % of GDP", "", "", "82.6 / 5,000 = 1.65%"),
]
for i, (ind, tot, par, val) in enumerate(ind_answers):
    r = r_s3 + 2 + i
    data_cell(ws, r, 2, ind).alignment = align_left
    data_cell(ws, r, 3, tot)
    data_cell(ws, r, 4, par)
    data_cell(ws, r, 5, val)

r_sum = r + 2
ws.cell(row=r_sum, column=2, value="Key results:").font = font_bold
ws.cell(row=r_sum + 1, column=2, value="  Expenditure-weighted partial = 25.8%").font = font_body
ws.cell(row=r_sum + 2, column=2, value="  Coastal tourism GVA = USD 82.6 million (rounded to 82.5M)").font = font_body
ws.cell(row=r_sum + 3, column=2, value="  Coastal tourism as % of GDP = 1.65%").font = font_body

# ---- Sheet 7: Glossary ---------------------------------------------------
ws = wb.create_sheet("Glossary")
ws.sheet_properties.tabColor = TEAL
set_col_widths(ws, [3, 28, 65])

title_banner(ws, 1, "  Glossary of Key Terms", 3)

glossary = [
    ("TSA", "Tourism Satellite Account - a statistical framework (UN/UNWTO) that measures the economic contribution of tourism in a manner consistent with national accounts."),
    ("OTSA", "Ocean Tourism Satellite Account - an extension of the TSA that isolates the coastal/ocean component of tourism activity."),
    ("OESA", "Ocean Economy Satellite Account - a broader account covering all ocean-related economic activity, of which OTSA covers the tourism component."),
    ("Tourism partial", "The share (proportion) of national tourism activity attributable to the coastal/ocean economy, used to estimate ocean tourism from national TSA data."),
    ("GVA", "Gross Value Added - the value of goods and services produced minus the cost of intermediate inputs. Measures economic contribution of a sector."),
    ("GDP", "Gross Domestic Product - the total value of goods and services produced in a country. GDP = GVA + taxes on products - subsidies on products."),
    ("Internal tourism consumption", "Total spending by both domestic and inbound tourists within the country."),
    ("Inbound tourism", "Tourism by non-residents visiting a country (international visitors)."),
    ("Domestic tourism", "Tourism by residents within their own country."),
    ("Tourism direct GVA", "The GVA generated directly by industries serving visitors (accommodation, food, transport, recreation, etc.)."),
    ("SEEA", "System of Environmental-Economic Accounting - the international statistical standard for environmental-economic accounts, complementing the SNA."),
    ("SNA", "System of National Accounts - the international standard for compiling national economic statistics (GDP, GVA, etc.)."),
    ("Partial accounting", "A method of estimating a sub-sector's contribution by applying a ratio (partial) to aggregate data when detailed data is unavailable."),
    ("Expenditure-weighted partial", "A partial that accounts for differences in spending levels across visitor types, giving more weight to higher-spending segments."),
]
header_row(ws, 3, ["", "Term", "Definition"], fill=fill_teal)
for i, (term, defn) in enumerate(glossary):
    r = 4 + i
    data_cell(ws, r, 2, term, bold=True).alignment = align_left
    data_cell(ws, r, 3, defn).alignment = align_wrap
    ws.row_dimensions[r].height = 40

# ---------------------------------------------------------------------------
# Save
# ---------------------------------------------------------------------------
out_dir = os.path.dirname(os.path.abspath(__file__))
out_path = os.path.join(out_dir, "otsa_exercise.xlsx")
wb.save(out_path)
print(f"Saved: {out_path}")
