#!/usr/bin/env python3
"""
Generate social_scoping.xlsx -- a structured scoping workbook for
Social and Governance Accounts following the GOAP 9-stage process.
"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# GOAP colour palette
# ---------------------------------------------------------------------------
GREEN = "3B9C7B"
TEAL = "0A5455"
GRAY = "404040"
TEXT = "30302F"
WHITE = "FFFFFF"
YELLOW = "FFF2CC"
TAN = "B59F81"

# Reusable styles
fill_teal = PatternFill(start_color=TEAL, end_color=TEAL, fill_type="solid")
fill_green = PatternFill(start_color=GREEN, end_color=GREEN, fill_type="solid")
fill_yellow = PatternFill(start_color=YELLOW, end_color=YELLOW, fill_type="solid")
fill_gray = PatternFill(start_color=GRAY, end_color=GRAY, fill_type="solid")
fill_tan = PatternFill(start_color=TAN, end_color=TAN, fill_type="solid")
fill_white = PatternFill(start_color=WHITE, end_color=WHITE, fill_type="solid")

font_title = Font(name="Calibri", size=16, bold=True, color=WHITE)
font_subtitle = Font(name="Calibri", size=13, bold=True, color=TEAL)
font_section = Font(name="Calibri", size=12, bold=True, color=TEAL)
font_header = Font(name="Calibri", size=11, bold=True, color=WHITE)
font_body = Font(name="Calibri", size=11, color=TEXT)
font_body_bold = Font(name="Calibri", size=11, bold=True, color=TEXT)
font_italic = Font(name="Calibri", size=11, italic=True, color=GRAY)
font_white_bold = Font(name="Calibri", size=11, bold=True, color=WHITE)

thin_border = Border(
    left=Side(style="thin", color=GRAY),
    right=Side(style="thin", color=GRAY),
    top=Side(style="thin", color=GRAY),
    bottom=Side(style="thin", color=GRAY),
)

wrap = Alignment(wrap_text=True, vertical="top")
wrap_center = Alignment(wrap_text=True, vertical="center", horizontal="center")
wrap_left = Alignment(wrap_text=True, vertical="top", horizontal="left")


# ---------------------------------------------------------------------------
# Helper functions
# ---------------------------------------------------------------------------
def set_col_widths(ws, widths: dict):
    for col_letter, w in widths.items():
        ws.column_dimensions[col_letter].width = w


def title_banner(ws, row, text, last_col="I"):
    ws.merge_cells(f"A{row}:{last_col}{row}")
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = font_title
    cell.fill = fill_teal
    cell.alignment = Alignment(vertical="center")
    ws.row_dimensions[row].height = 40


def subtitle(ws, row, text, last_col="I"):
    ws.merge_cells(f"A{row}:{last_col}{row}")
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = font_subtitle
    cell.alignment = wrap


def section_label(ws, row, text, last_col="I"):
    ws.merge_cells(f"A{row}:{last_col}{row}")
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = font_section
    cell.alignment = wrap


def note_row(ws, row, text, last_col="I"):
    ws.merge_cells(f"A{row}:{last_col}{row}")
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = font_italic
    cell.alignment = wrap


def header_row(ws, row, headers, fill=fill_green):
    for col_idx, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=col_idx, value=h)
        c.font = font_header
        c.fill = fill
        c.alignment = wrap_center
        c.border = thin_border


def body_cell(ws, row, col, value="", yellow=False, bold=False):
    c = ws.cell(row=row, column=col, value=value)
    c.font = font_body_bold if bold else font_body
    c.alignment = wrap
    c.border = thin_border
    if yellow:
        c.fill = fill_yellow
    return c


def yellow_input_row(ws, row, label, start_col=1, end_col=2):
    body_cell(ws, row, start_col, label, bold=True)
    for col in range(start_col + 1, end_col + 1):
        body_cell(ws, row, col, "", yellow=True)


# ---------------------------------------------------------------------------
# Sheet 1: README
# ---------------------------------------------------------------------------
def build_readme(wb):
    ws = wb.active
    ws.title = "README"
    ws.sheet_properties.tabColor = TEAL
    set_col_widths(ws, {"A": 90})

    title_banner(ws, 1, "Social and Governance Accounts -- Scoping Workbook", "A")

    rows = [
        "",
        "PURPOSE",
        "Social accounts are an emerging domain within the Global Ocean Accounts Partnership (GOAP) draft guidance.",
        "This workbook helps you scope what social data exists in your country and plan a pilot social account.",
        "It follows Stages 1-3 of the 9-stage Social Account Development process.",
        "",
        "IMPORTANT NOTES",
        "There are no standard calculation tables yet for social accounts -- this is a data audit and planning tool.",
        "The workbook is designed as a structured scoping exercise, not a calculation exercise.",
        "",
        "FOUR PILLARS OF SOCIAL ACCOUNTS",
        "  1. Human Wellbeing -- health, education, food security, life satisfaction",
        "  2. Ocean-Based Livelihoods -- employment, income, working conditions",
        "  3. Equity & Access -- distribution of benefits, gender equity, indigenous rights",
        "  4. Cultural Connections -- traditional knowledge, cultural practices, heritage sites",
        "",
        "HOW TO USE THIS WORKBOOK",
        "  Sheet 2  'Stage 1 - Co-design'       Define your scope and stakeholders",
        "  Sheet 3  'Stage 2 - Data Audit'       Inventory existing social data",
        "  Sheet 4  'Stage 3 - Gap Analysis'     Identify gaps by pillar",
        "  Sheet 5  'Next Steps'                 Plan your pilot social account",
        "  Sheet 6  'Example Indicators'         Reference list of stock and flow indicators",
        "  Sheet 7  'Glossary'                   Key terms and definitions",
        "",
        "Work through the sheets in order. Yellow cells are for your inputs.",
    ]
    for i, text in enumerate(rows, 3):
        c = ws.cell(row=i, column=1, value=text)
        if text in ("PURPOSE", "IMPORTANT NOTES", "FOUR PILLARS OF SOCIAL ACCOUNTS", "HOW TO USE THIS WORKBOOK"):
            c.font = font_section
        elif text.startswith("  "):
            c.font = font_body
        else:
            c.font = font_body
        c.alignment = wrap


# ---------------------------------------------------------------------------
# Sheet 2: Stage 1 - Co-design
# ---------------------------------------------------------------------------
def build_stage1(wb):
    ws = wb.create_sheet("Stage 1 - Co-design")
    ws.sheet_properties.tabColor = GREEN
    set_col_widths(ws, {"A": 30, "B": 35, "C": 25, "D": 25})

    title_banner(ws, 1, "Stage 1: Co-design -- Define Your Scope", "D")

    r = 3
    section_label(ws, r, "Basic information", "D"); r += 1

    # Country/region
    body_cell(ws, r, 1, "Country / region:", bold=True)
    ws.merge_cells(f"B{r}:D{r}")
    body_cell(ws, r, 2, "", yellow=True); r += 1

    # Accounting area
    body_cell(ws, r, 1, "Accounting area:", bold=True)
    ws.merge_cells(f"B{r}:D{r}")
    body_cell(ws, r, 2, "", yellow=True); r += 2

    # Policy question
    section_label(ws, r, "Policy question to answer (choose one or write your own)", "D"); r += 1
    examples = [
        "How do fishing communities benefit from MPAs?",
        "What livelihoods depend on the reef?",
        "How are ocean resources distributed across income groups?",
        "What cultural connections exist with the ocean?",
        "Who participates in ocean governance decisions?",
    ]
    for ex in examples:
        note_row(ws, r, f'  Example: "{ex}"', "D"); r += 1
    r += 1
    body_cell(ws, r, 1, "Your policy question:", bold=True)
    ws.merge_cells(f"B{r}:D{r}")
    body_cell(ws, r, 2, "", yellow=True)
    ws.row_dimensions[r].height = 40; r += 2

    # Stakeholders
    section_label(ws, r, "Key stakeholders to involve", "D"); r += 1
    header_row(ws, r, ["Stakeholder type", "Name / organization", "Role", "Contact"]); r += 1
    for _ in range(8):
        for col in range(1, 5):
            body_cell(ws, r, col, "", yellow=True)
        r += 1
    r += 1

    # Expected outputs
    section_label(ws, r, "Expected outputs", "D"); r += 1
    header_row(ws, r, ["Output", "Description", "", ""]); r += 1
    ws.merge_cells(f"B{r}:D{r}")
    body_cell(ws, r, 1, "Agreed scope", bold=True)
    body_cell(ws, r, 2, "", yellow=True); r += 1
    ws.merge_cells(f"B{r}:D{r}")
    body_cell(ws, r, 1, "Priority indicators", bold=True)
    body_cell(ws, r, 2, "", yellow=True); r += 1
    ws.merge_cells(f"B{r}:D{r}")
    body_cell(ws, r, 1, "Accounting area boundary", bold=True)
    body_cell(ws, r, 2, "", yellow=True)


# ---------------------------------------------------------------------------
# Sheet 3: Stage 2 - Data Audit
# ---------------------------------------------------------------------------
def build_stage2(wb):
    ws = wb.create_sheet("Stage 2 - Data Audit")
    ws.sheet_properties.tabColor = GREEN
    cols = {
        "A": 28, "B": 22, "C": 14, "D": 22, "E": 20,
        "F": 14, "G": 18, "H": 18, "I": 22,
    }
    set_col_widths(ws, cols)

    title_banner(ws, 1, "Stage 2: Social Data Audit -- What Exists in Your Country?")

    r = 3
    note_row(ws, r, "Complete this inventory for each data source. Yellow cells are for your input."); r += 2

    headers = [
        "Data category", "Data source", "Available?\n(Y/N)",
        "Holder / agency", "Coverage\n(national/regional/local)",
        "Frequency", "Spatial resolution",
        "Relevance to your\nquestion (H/M/L)", "Notes",
    ]
    header_row(ws, r, headers); r += 1

    categories = [
        "National census",
        "Household income survey",
        "Labor force survey",
        "Time-use survey",
        "Health statistics",
        "Education statistics",
        "Fisheries employment records",
        "Tourism employment data",
        "Coastal population data",
        "MPA management records",
        "Fishing permits / licenses",
        "Cultural heritage database",
        "Community survey data",
        "Traditional knowledge records",
        "Water / sanitation data (WASH)",
        "Food security data",
        "Disaster / vulnerability data",
        "Governance / compliance records",
    ]
    for cat in categories:
        body_cell(ws, r, 1, cat, bold=True)
        for col in range(2, 10):
            body_cell(ws, r, col, "", yellow=True)
        ws.row_dimensions[r].height = 28
        r += 1


# ---------------------------------------------------------------------------
# Sheet 4: Stage 3 - Gap Analysis
# ---------------------------------------------------------------------------
def build_stage3(wb):
    ws = wb.create_sheet("Stage 3 - Gap Analysis")
    ws.sheet_properties.tabColor = GREEN
    set_col_widths(ws, {"A": 24, "B": 30, "C": 30, "D": 28, "E": 28})

    title_banner(ws, 1, "Stage 3: Gap Analysis by Pillar", "E")

    r = 3
    note_row(ws, r,
        "For each pillar, summarise the data you found in the audit and identify gaps.", "E"); r += 2

    headers = [
        "Pillar", "Available data\n(from audit)", "Data gaps",
        "Priority gap to fill", "Method to fill gap",
    ]
    header_row(ws, r, headers); r += 1

    pillars = [
        "Human Wellbeing",
        "Ocean-Based Livelihoods",
        "Equity & Access",
        "Cultural Connections",
    ]
    for p in pillars:
        body_cell(ws, r, 1, p, bold=True)
        for col in range(2, 6):
            body_cell(ws, r, col, "", yellow=True)
        ws.row_dimensions[r].height = 60
        r += 1

    r += 1
    section_label(ws, r, "Reflection questions", "E"); r += 1

    questions = [
        "Which pillar has the most data?",
        "Which pillar has the least data?",
        "What primary data collection would fill the biggest gap?",
    ]
    for q in questions:
        body_cell(ws, r, 1, q, bold=True)
        ws.merge_cells(f"B{r}:E{r}")
        body_cell(ws, r, 2, "", yellow=True)
        ws.row_dimensions[r].height = 36
        r += 1


# ---------------------------------------------------------------------------
# Sheet 5: Next Steps
# ---------------------------------------------------------------------------
def build_next_steps(wb):
    ws = wb.create_sheet("Next Steps")
    ws.sheet_properties.tabColor = TAN
    set_col_widths(ws, {"A": 42, "B": 60})

    title_banner(ws, 1, "Planning Your Pilot Social Account", "B")

    r = 3
    note_row(ws, r,
        "Use this sheet to plan a concrete pilot. Yellow cells are for your input.", "B"); r += 2

    header_row(ws, r, ["Item", "Your answer"]); r += 1

    items = [
        "Priority indicator for pilot",
        "Data already available for this indicator",
        "Data still needed",
        "Collection method (survey, interview, existing records)",
        "Key partners",
        "Timeline (months)",
        "Resources needed",
        "Who leads the pilot?",
        "How will you validate with communities?",
        "How will findings feed into policy?",
    ]
    for item in items:
        body_cell(ws, r, 1, item, bold=True)
        body_cell(ws, r, 2, "", yellow=True)
        ws.row_dimensions[r].height = 36
        r += 1


# ---------------------------------------------------------------------------
# Sheet 6: Example Indicators
# ---------------------------------------------------------------------------
def build_example_indicators(wb):
    ws = wb.create_sheet("Example Indicators")
    ws.sheet_properties.tabColor = GREEN
    set_col_widths(ws, {"A": 30, "B": 10, "C": 70})

    title_banner(ws, 1, "Reference: Stock and Flow Indicators (GOAP Circular)", "C")

    r = 3
    section_label(ws, r, "STOCK indicators -- measured at a point in time", "C"); r += 1
    header_row(ws, r, ["Category", "Type", "Example indicators"]); r += 1

    stocks = [
        ("Human capital", "Stock",
         "Educational attainment, fish consumption, life satisfaction, health status"),
        ("Social capital", "Stock",
         "Trust levels, cooperation, community management groups, partnerships"),
        ("Institutional capital", "Stock",
         "MPA coverage %, compliance rates, co-management agreements"),
    ]
    for cat, typ, examples in stocks:
        body_cell(ws, r, 1, cat, bold=True)
        body_cell(ws, r, 2, typ)
        body_cell(ws, r, 3, examples)
        ws.row_dimensions[r].height = 28
        r += 1

    r += 1
    section_label(ws, r, "FLOW indicators -- measured in hours over a period", "C"); r += 1
    header_row(ws, r, ["Category", "Type", "Example indicators"]); r += 1

    flows = [
        ("Unpaid household labour", "Flow",
         "Hours processing marine catch, caring for fishing families"),
        ("Social transfers", "Flow",
         "Hours sharing boats/equipment, sharing traditional knowledge"),
        ("Volunteering", "Flow",
         "Hours on beach cleanups, community monitoring"),
        ("Cultural / recreational", "Flow",
         "Hours on traditional ceremonies, ocean recreation"),
        ("Governance participation", "Flow",
         "Hours in fisheries management, coastal zone planning"),
    ]
    for cat, typ, examples in flows:
        body_cell(ws, r, 1, cat, bold=True)
        body_cell(ws, r, 2, typ)
        body_cell(ws, r, 3, examples)
        ws.row_dimensions[r].height = 28
        r += 1


# ---------------------------------------------------------------------------
# Sheet 7: Glossary
# ---------------------------------------------------------------------------
def build_glossary(wb):
    ws = wb.create_sheet("Glossary")
    ws.sheet_properties.tabColor = TEAL
    set_col_widths(ws, {"A": 26, "B": 80})

    title_banner(ws, 1, "Glossary of Key Terms", "B")

    r = 3
    header_row(ws, r, ["Term", "Definition"]); r += 1

    terms = [
        ("Social accounts",
         "Accounts that measure the social dimensions of human-ocean interactions, "
         "including wellbeing, livelihoods, equity, and cultural connections."),
        ("Governance accounts",
         "Accounts that capture institutional arrangements, participation in "
         "decision-making, compliance, and management effectiveness for ocean resources."),
        ("Stocks",
         "Social indicators measured at a point in time (e.g., educational attainment, "
         "trust levels, MPA coverage percentage)."),
        ("Flows",
         "Social indicators measured over a period, typically in hours "
         "(e.g., hours of unpaid labour, volunteering, governance participation)."),
        ("BSU (Basic Spatial Unit)",
         "The smallest geographic unit used in ocean accounts to link biophysical, "
         "economic, and social data to a specific marine or coastal area."),
        ("Co-design",
         "Stage 1 of the social accounts process where the scope, policy questions, "
         "and stakeholder roles are agreed collaboratively with communities and agencies."),
        ("FPIC (Free, Prior and Informed Consent)",
         "The principle that indigenous peoples and local communities have the right "
         "to give or withhold consent for activities that affect their lands, territories, "
         "or resources, with full information provided beforehand."),
        ("Time-use survey",
         "A survey instrument that records how individuals allocate their time across "
         "activities (paid work, unpaid household labour, recreation, governance, etc.) "
         "over a defined period. Key data source for flow indicators."),
        ("Human capital",
         "The stock of knowledge, skills, health, and nutrition embodied in people. "
         "In ocean social accounts this includes educational attainment, fish consumption, "
         "and life satisfaction of coastal communities."),
        ("Social capital",
         "The networks, norms, and trust that enable collective action. Measured through "
         "trust levels, cooperation indices, community management groups, and partnerships."),
        ("Institutional capital",
         "The formal and informal rules, organisations, and governance structures that "
         "manage ocean resources. Includes MPA coverage, compliance rates, and "
         "co-management agreements."),
        ("GOAP",
         "Global Ocean Accounts Partnership -- an international partnership that develops "
         "technical guidance for ocean accounts, including the emerging social and "
         "governance accounts framework."),
    ]
    for term, definition in terms:
        body_cell(ws, r, 1, term, bold=True)
        body_cell(ws, r, 2, definition)
        ws.row_dimensions[r].height = 45
        r += 1


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    wb = Workbook()
    build_readme(wb)
    build_stage1(wb)
    build_stage2(wb)
    build_stage3(wb)
    build_next_steps(wb)
    build_example_indicators(wb)
    build_glossary(wb)

    out_dir = os.path.dirname(os.path.abspath(__file__))
    out_path = os.path.join(out_dir, "social_scoping.xlsx")
    wb.save(out_path)
    print(f"Created: {out_path}")


if __name__ == "__main__":
    main()
